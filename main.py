import os
import requests
from flask import Flask, render_template, Response, request, redirect, url_for, jsonify, flash, session, send_file, send_from_directory
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
from flask_wtf import FlaskForm
from flask_wtf.csrf import CSRFProtect
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from PIL import Image
import stripe
from datetime import datetime, timedelta
from database import db
from decimal import Decimal
from models import get_utc_time, Product
import uuid
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.barcode import code128
from reportlab.graphics import renderPDF
import io
import random
import string
import midtransclient
import json
import jwt
import sys  # Import sys to check command line arguments

# Import Xendit and DOKU libraries
try:
    import xendit
    from xendit import EWallet, Invoice, VirtualAccount, QRCode
    XENDIT_AVAILABLE = True
except ImportError:
    XENDIT_AVAILABLE = False
    print("Warning: Xendit library not available")

try:
    from doku_python_library.src.snap import DokuSNAP
    from doku_python_library.src.model.va.virtual_account_request import VirtualAccountRequest
    from doku_python_library.src.model.va.total_amount import TotalAmount
    from doku_python_library.src.model.va.virtual_account_info import VirtualAccountInfo
    from doku_python_library.src.model.va.customer import Customer
    DOKU_AVAILABLE = True
except ImportError:
    DOKU_AVAILABLE = False
    print("Warning: DOKU library not available")

# Create the Flask app
app = Flask(__name__)

# Add ProxyFix middleware for Replit proxy support
from werkzeug.middleware.proxy_fix import ProxyFix

app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# Configure CORS for chat subdomain and cross-origin requests
CORS(
    app,
    resources={
        r"/api/*": {
            "origins": [
                "*"
            ],  # Allow all origins for API endpoints (includes chat subdomain)
            "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
            "allow_headers":
            ["Content-Type", "Authorization", "X-Requested-With"],
            "supports_credentials": False
        },
        r"/uploads/*": {
            "origins": ["*"],  # Allow all origins for media files
            "methods": ["GET", "OPTIONS"],
            "allow_headers": ["Content-Type"],
            "supports_credentials": False
        },
        r"/static/chat_media/*": {
            "origins": ["*"],  # Allow all origins for chat media
            "methods": ["GET", "OPTIONS"],
            "allow_headers": ["Content-Type"],
            "supports_credentials": False
        }
    })

# Configuration
if not os.environ.get("SESSION_SECRET"):
    raise ValueError("SESSION_SECRET environment variable is required")
app.secret_key = os.environ.get("SESSION_SECRET")
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL")
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_recycle": 300,
    "pool_pre_ping": True,
}

# Security configuration - Universal deployment ready
is_production = os.environ.get('IS_PRODUCTION', 'false').lower() == 'true'
debug_mode = os.environ.get('FLASK_DEBUG', '0') == '1'

app.config['SESSION_COOKIE_SECURE'] = is_production  # HTTPS only in production
app.config['SESSION_COOKIE_HTTPONLY'] = True  # No JS access
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # CSRF protection
app.config['REMEMBER_COOKIE_SECURE'] = is_production
app.config['REMEMBER_COOKIE_HTTPONLY'] = True

# CSRF trusted origins for Replit proxy
replit_domain = os.environ.get('REPLIT_DEV_DOMAIN', '')
if replit_domain:
    app.config['WTF_CSRF_TRUSTED_ORIGINS'] = [
        f'https://{replit_domain}', f'http://{replit_domain}',
        'http://localhost:5000', 'http://127.0.0.1:5000', 'http://0.0.0.0:5000'
    ]

# Initialize extensions
from database import configure_database
if not configure_database(app):
    print("[ERROR] Gagal mengkonfigurasi database, aplikasi akan berhenti")
    exit(1)

migrate = Migrate(app, db)
csrf = CSRFProtect(app)


# Context processor to make store profile available in all templates
@app.context_processor
def inject_store_profile():
    """Make store profile available to all templates"""
    try:
        profile = models.StoreProfile.get_active_profile()
        return dict(store_profile=profile)
    except Exception as e:
        print(f"[ERROR] Failed to inject store profile: {e}")
        db.session.rollback()
        return dict(store_profile=None)


# File upload configuration
app.config['UPLOAD_FOLDER'] = 'static/public/produk_images'
app.config['CHAT_MEDIA_FOLDER'] = 'static/chat_media'
app.config['UPLOADS_MEDIA_FOLDER'] = 'uploads/medias_sends'
app.config[
    'MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size for video support
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
ALLOWED_VIDEO_EXTENSIONS = {'mp4', 'webm', 'mov', 'avi'}
ALLOWED_CHAT_MEDIA = ALLOWED_EXTENSIONS | ALLOWED_VIDEO_EXTENSIONS


# Buat folder upload jika belum ada
def ensure_upload_folders():
    folders = [
        app.config['UPLOAD_FOLDER'], app.config['CHAT_MEDIA_FOLDER'],
        app.config['UPLOADS_MEDIA_FOLDER']
    ]
    for folder in folders:
        try:
            if not os.path.exists(folder):
                os.makedirs(folder, mode=0o755, exist_ok=True)
                print(f"[OK] Created folder: {folder}")

            # Ensure proper permissions
            if os.path.exists(folder):
                os.chmod(folder, 0o755)
                print(f"[OK] Set permissions for folder: {folder}")
        except Exception as e:
            print(
                f"[ERROR] Failed to create/set permissions for folder {folder}: {e}"
            )


ensure_upload_folders()


def allowed_file(filename):
    return '.' in filename and filename.rsplit(
        '.', 1)[1].lower() in ALLOWED_EXTENSIONS


def allowed_chat_media(filename):
    return '.' in filename and filename.rsplit(
        '.', 1)[1].lower() in ALLOWED_CHAT_MEDIA


def is_video_file(filename):
    return '.' in filename and filename.rsplit(
        '.', 1)[1].lower() in ALLOWED_VIDEO_EXTENSIONS


def get_image_orientation(image_path):
    """Determine if image is landscape or portrait"""
    try:
        img = Image.open(image_path)
        width, height = img.size
        if width > height:
            return 'landscape'
        elif height > width:
            return 'portrait'
        else:
            return 'square'
    except Exception as e:
        print(f"[ERROR] Could not determine orientation for {image_path}: {e}")
        return 'unknown'


def compress_image(image_path, max_size_mb=1):
    """Compress image to be under max_size_mb while preserving original format"""
    try:
        img = Image.open(image_path)
        original_format = img.format

        # Check if file is already under size limit
        file_size_mb = os.path.getsize(image_path) / (1024 * 1024)
        if file_size_mb <= max_size_mb:
            return image_path  # No compression needed

        # Don't compress animated GIFs - they lose animation
        if original_format == 'GIF' and getattr(img, 'is_animated', False):
            return image_path  # Keep animated GIFs as-is

        # Set format-specific parameters
        save_kwargs = {'optimize': True}
        quality = 85

        if original_format in ('JPEG', 'JPG'):
            save_format = 'JPEG'
            save_kwargs['quality'] = quality
        elif original_format == 'PNG':
            save_format = 'PNG'
            # PNG doesn't use quality, but we can optimize
            save_kwargs = {'optimize': True}
        elif original_format == 'WEBP':
            save_format = 'WEBP'
            save_kwargs['quality'] = quality
        else:
            # For other formats, convert to PNG to preserve quality
            save_format = 'PNG'
            save_kwargs = {'optimize': True}

        # Try compression with decreasing quality for lossy formats
        max_iterations = 10
        iteration = 0

        while iteration < max_iterations:
            iteration += 1

            # Update quality for lossy formats
            if save_format in ('JPEG', 'WEBP'):
                save_kwargs = {'optimize': True, 'quality': quality}

            img.save(image_path, save_format, **save_kwargs)

            # Check file size
            file_size_mb = os.path.getsize(image_path) / (1024 * 1024)

            if file_size_mb <= max_size_mb:
                break

            # For lossless formats (PNG), try downscaling dimensions
            current_width, current_height = img.size
            if current_width > 800 or current_height > 800:
                # Reduce by 80% each iteration
                new_width = int(current_width * 0.8)
                new_height = int(current_height * 0.8)
                img = img.resize((new_width, new_height),
                                 Image.Resampling.LANCZOS)
            else:
                # Can't compress further, break to avoid infinite loop
                break

        return image_path
    except Exception as e:
        print(f"[ERROR] Failed to compress image {image_path}: {e}")
        return image_path


# Login Manager setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Silakan login untuk mengakses halaman ini.'

# Stripe configuration
if not os.environ.get('STRIPE_SECRET_KEY'):
    print(
        "Warning: STRIPE_SECRET_KEY not set, using placeholder for development"
    )
# Stripe API key will be set dynamically per request from PaymentConfiguration
# stripe.api_key = os.environ.get('STRIPE_SECRET_KEY', 'sk_test_placeholder_for_development')


@login_manager.user_loader
def load_user(user_id):
    try:
        return db.session.get(models.User, int(user_id))
    except Exception as e:
        print(f"[ERROR] Failed to load user {user_id}: {e}")
        # Rollback the failed transaction
        db.session.rollback()
        return None


def admin_required(f):
    from functools import wraps

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash(
                'Akses ditolak. Anda harus admin untuk mengakses halaman ini.',
                'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)

    return decorated_function


def staff_required(f):
    from functools import wraps

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or (not current_user.is_admin and
                                                 not current_user.is_staff):
            flash(
                'Akses ditolak. Anda harus staff atau admin untuk mengakses halaman ini.',
                'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)

    return decorated_function


# JWT Configuration for chat service integration
JWT_SECRET_KEY = os.environ.get("JWT_SECRET_KEY") or os.environ.get(
    "SESSION_SECRET")  # Prefer separate JWT secret
JWT_ALGORITHM = 'HS256'
JWT_ACCESS_TOKEN_LIFETIME = 86400  # 24 hours


def generate_jwt_token(user):
    """Generate JWT token for chat service authentication"""
    payload = {
        'user_id': user.id,
        'email': user.email,
        'name': user.name,
        'role': user.role,
        'iat': datetime.utcnow(),
        'exp': datetime.utcnow() + timedelta(seconds=JWT_ACCESS_TOKEN_LIFETIME)
    }
    return jwt.encode(payload, JWT_SECRET_KEY, algorithm=JWT_ALGORITHM)


# Import models before routes
import models


def setup_django_chat_service():
    """Setup Django chat service and run migrations"""
    import subprocess
    import os
    import sys
    from pathlib import Path
    import time

    # Get absolute paths
    project_root = Path(__file__).resolve().parent
    chat_service_dir = project_root / 'chat_service'

    if not chat_service_dir.exists():
        print("[ERROR] Chat service directory not found")
        return False

    try:
        # Add paths to sys.path for Django imports
        sys.path.insert(0, str(chat_service_dir))
        sys.path.insert(0, str(project_root))

        # Set environment variables for Django
        os.environ.setdefault('DJANGO_SETTINGS_MODULE',
                              'chat_microservice.settings')
        os.environ.setdefault('DJANGO_SECRET_KEY', app.secret_key)

        # Copy Flask database URL to Django
        if app.config.get('SQLALCHEMY_DATABASE_URI'):
            os.environ.setdefault('DATABASE_URL',
                                  app.config['SQLALCHEMY_DATABASE_URI'])

        # Set domains for Django ALLOWED_HOSTS
        os.environ.setdefault('DOMAINS',
                              'kasir.fajarmandiri.store,fajarmandiri.store')

        # Change to chat service directory
        original_cwd = os.getcwd()
        os.chdir(str(chat_service_dir))

        print("[SETUP] Running Django chat service migrations...")

        # Get Python executable path
        python_exec = sys.executable
        manage_py = chat_service_dir / 'manage.py'

        # Clean up existing migrations more safely
        migrations_dir = chat_service_dir / 'chat' / 'migrations'
        if migrations_dir.exists():
            for migration_file in migrations_dir.glob('*.py'):
                if migration_file.name != '__init__.py':
                    try:
                        migration_file.unlink()
                        print(
                            f"[CLEANUP] Removed old migration: {migration_file.name}"
                        )
                    except Exception as e:
                        print(
                            f"[WARNING] Could not remove {migration_file.name}: {e}"
                        )

        # Create __init__.py if it doesn't exist
        init_file = migrations_dir / '__init__.py'
        if not init_file.exists():
            migrations_dir.mkdir(parents=True, exist_ok=True)
            init_file.touch()

        # Make fresh migrations with better error handling
        makemigrations_cmd = [
            python_exec, str(manage_py), 'makemigrations', 'chat'
        ]
        result = subprocess.run(makemigrations_cmd,
                                capture_output=True,
                                text=True,
                                timeout=30)

        if result.returncode == 0:
            print("[OK] Django migrations created")
        else:
            print(f"[WARNING] Makemigrations output: {result.stdout}")
            if result.stderr:
                print(f"[WARNING] Makemigrations stderr: {result.stderr}")

        # Apply migrations with --fake-initial to skip tables that already exist
        migrate_cmd = [
            python_exec,
            str(manage_py), 'migrate', '--fake-initial'
        ]
        result = subprocess.run(migrate_cmd,
                                capture_output=True,
                                text=True,
                                timeout=60)

        if result.returncode == 0:
            print("[OK] Django chat service migrations completed")
        else:
            print(f"[WARNING] Migration output: {result.stdout}")
            if result.stderr:
                print(f"[WARNING] Migration stderr: {result.stderr}")
            # Try without --fake-initial as fallback
            migrate_cmd = [python_exec, str(manage_py), 'migrate']
            result = subprocess.run(migrate_cmd,
                                    capture_output=True,
                                    text=True,
                                    timeout=60)
            if result.returncode == 0:
                print(
                    "[OK] Django chat service migrations completed (fallback)")

        # Test Django configuration
        test_cmd = [
            python_exec, '-c', '''
import os
import django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "chat_microservice.settings")
django.setup()
from chat.models import ChatRoom
print("[TEST] Django models accessible")
'''
        ]
        result = subprocess.run(test_cmd,
                                capture_output=True,
                                text=True,
                                timeout=15,
                                cwd=str(chat_service_dir))

        if result.returncode == 0:
            print("[OK] Django chat service configured")
        else:
            print(f"[WARNING] Django test failed: {result.stderr}")

        # Start Django service if not already running
        try:
            import requests
            response = requests.get('http://127.0.0.1:8000/health/', timeout=2)
            if response.status_code == 200:
                print("[OK] Django chat service already running")
            else:
                print("[INFO] Starting Django chat service...")
                start_django_service()
        except:
            print("[INFO] Starting Django chat service...")
            start_django_service()

        # Return to original directory
        os.chdir(original_cwd)
        return True

    except subprocess.TimeoutExpired:
        print("[ERROR] Django setup timeout")
        os.chdir(original_cwd)
        return False
    except Exception as e:
        print(f"[ERROR] Django setup error: {e}")
        if 'original_cwd' in locals():
            os.chdir(original_cwd)
        return False


def start_django_service():
    """Start Django chat service in background using Daphne ASGI server"""
    import subprocess
    import sys
    from pathlib import Path

    try:
        project_root = Path(__file__).resolve().parent
        chat_service_dir = project_root / 'chat_service'

        # Always use Daphne for WebSocket support (works best in Replit)
        subprocess.Popen([
            sys.executable, '-m', 'daphne', '-b', '0.0.0.0', '-p', '8000',
            'chat_microservice.asgi:application'
        ],
                         cwd=str(chat_service_dir),
                         stdout=subprocess.DEVNULL,
                         stderr=subprocess.DEVNULL)

        # Wait a moment for service to start
        import time
        time.sleep(3)

        print(
            "[OK] Django chat service started on port 8000 with Daphne ASGI server"
        )
        return True

    except Exception as e:
        print(f"[ERROR] Failed to start Django service: {e}")
        return False


def check_django_service():
    """Check if Django chat service is running"""
    try:
        import requests
        response = requests.get('http://127.0.0.1:8000/health/', timeout=2)
        return response.status_code == 200
    except:
        return False


def create_sample_data():
    """Create sample data for testing"""
    try:
        # Rollback any existing failed transaction first
        db.session.rollback()

        # Create categories
        categories_data = [
            {
                'name': 'Gitar',
                'description': 'Gitar akustik dan elektrik'
            },
            {
                'name': 'Bass',
                'description': 'Bass elektrik dan akustik'
            },
            {
                'name': 'Drum',
                'description': 'Drum kit dan perkusi'
            },
            {
                'name': 'Keyboard',
                'description': 'Keyboard dan piano digital'
            },
            {
                'name': 'Sound System',
                'description': 'Speaker, mixer, dan audio equipment'
            },
        ]

        for cat_data in categories_data:
            try:
                existing = models.Category.query.filter_by(
                    name=cat_data['name']).first()
                if not existing:
                    category = models.Category(**cat_data)
                    db.session.add(category)
            except Exception as e:
                print(
                    f"[WARNING] Could not check/create category {cat_data['name']}: {e}"
                )
                db.session.rollback()
                continue

        # Create suppliers
        suppliers_data = [{
            'name': 'Swelee Music Store',
            'contact_person': 'Swelee',
            'email': 'contact@swelee.com',
            'phone': '021-1234-5678',
            'company': 'PT Swelee Musik Indonesia'
        }, {
            'name': 'Media Recording Tech',
            'contact_person': 'Media Recording',
            'email': 'info@mediarecording.com',
            'phone': '021-8765-4321',
            'company': 'CV Media Recording Technology'
        }, {
            'name': 'Triple 3 Music Store',
            'contact_person': 'Triple Music',
            'email': 'sales@triple3music.com',
            'phone': '021-5555-3333',
            'company': 'Triple 3 Music Distribution'
        }]

        for sup_data in suppliers_data:
            try:
                existing = models.Supplier.query.filter_by(
                    name=sup_data['name']).first()
                if not existing:
                    supplier = models.Supplier(**sup_data)
                    db.session.add(supplier)
            except Exception as e:
                print(
                    f"[WARNING] Could not check/create supplier {sup_data['name']}: {e}"
                )
                db.session.rollback()
                continue

        # Create shipping services
        shipping_data = [{
            'name': 'JNE Regular',
            'code': 'jne_reg',
            'base_price': 15000,
            'price_per_kg': 5000,
            'min_days': 2,
            'max_days': 4
        }, {
            'name': 'JNE Express',
            'code': 'jne_exp',
            'base_price': 25000,
            'price_per_kg': 8000,
            'min_days': 1,
            'max_days': 2
        }, {
            'name': 'J&T Regular',
            'code': 'jnt_reg',
            'base_price': 12000,
            'price_per_kg': 4000,
            'min_days': 2,
            'max_days': 5
        }]

        for ship_data in shipping_data:
            try:
                existing = models.ShippingService.query.filter_by(
                    code=ship_data['code']).first()
                if not existing:
                    service = models.ShippingService(**ship_data)
                    db.session.add(service)
            except Exception as e:
                print(
                    f"[WARNING] Could not check/create shipping service {ship_data['code']}: {e}"
                )
                db.session.rollback()
                continue

        try:
            db.session.commit()
            print("[OK] Sample data created successfully")
        except Exception as commit_error:
            print(f"[ERROR] Failed to commit sample data: {commit_error}")
            db.session.rollback()

    except Exception as e:
        print(f"[ERROR] Error creating sample data: {e}")
        db.session.rollback()


# Global flag to prevent double initialization
_db_initialized = False


def initialize_database():
    """Initialize database with proper checks to prevent double initialization"""
    global _db_initialized

    if _db_initialized:
        print("[SKIP] Database already initialized")
        return

    try:
        # Ensure all Flask tables are created with current schema
        db.create_all()
        print("[OK] Flask database tables created")

        # Setup Django chat service
        if setup_django_chat_service():
            print("[OK] Django chat service setup completed")
        else:
            print(
                "[WARNING] Django chat service setup failed, continuing without chat"
            )

        # Create default admin user if it doesn't exist
        admin_email = "admin@hurtrock.com"

        # Check if admin user exists safely
        try:
            admin_user = models.User.query.filter_by(email=admin_email).first()
        except Exception as e:
            print(f"Database schema mismatch detected. Please run migration.")
            print(f"Error: {e}")
            admin_user = None

        if not admin_user:
            try:
                admin_user = models.User(
                    email=admin_email,
                    password_hash=generate_password_hash("admin123"),
                    name="Administrator",
                    role="admin")
                db.session.add(admin_user)
                db.session.commit()
                print(f"[OK] Default admin user created: {admin_email}")
            except Exception as e:
                print(f"[ERROR] Failed to create admin user: {e}")
        else:
            print(f"[OK] Admin user already exists: {admin_email}")

        # Create default store profile if it doesn't exist
        try:
            store_profile = models.StoreProfile.get_active_profile()
            if not store_profile:
                store_profile = models.StoreProfile(
                    store_name='Hurtrock Music Store',
                    store_tagline='Toko Alat Musik Terpercaya',
                    store_address=
                    'Jl Gegerkalong Girang complex Darut Tauhid Kav 22, Gegerkalong, Setiabudhi',
                    store_city='Kota Bandung',
                    store_postal_code='40153',
                    store_phone='0821-1555-8035',
                    store_email='info@hurtrock.com',
                    store_website='https://hurtrock-store.com',
                    whatsapp_number='6282115558035',
                    operating_hours=
                    'Senin - Sabtu: 09:00 - 21:00\nMinggu: 10:00 - 18:00',
                    branch_name='Cabang Pusat',
                    branch_code='HRT-001')
                db.session.add(store_profile)
                db.session.commit()
                print("[OK] Default store profile created")
            else:
                print("[OK] Store profile already exists")
        except Exception as e:
            print(f"[ERROR] Failed to create store profile: {e}")

        # Create sample data
        create_sample_data()

        # Mark as initialized
        _db_initialized = True
        print("[OK] Database initialization completed")

    except Exception as e:
        print(f"[ERROR] Database initialization error: {e}")
        db.session.rollback()


# Create database tables and setup everything
with app.app_context():
    # Initialize only Flask database for now, skip Django chat service during startup
    try:
        # Rollback any existing failed transaction first
        db.session.rollback()

        # Ensure all Flask tables are created with current schema
        db.create_all()
        print("[OK] Flask database tables created")

        # Create default admin user if it doesn't exist
        admin_email = "admin@hurtrock.com"

        # Check if admin user exists safely
        try:
            db.session.rollback()  # Clear any pending transaction
            admin_user = models.User.query.filter_by(email=admin_email).first()
        except Exception as e:
            print(f"Database schema issue detected: {e}")
            db.session.rollback()
            admin_user = None

        if not admin_user:
            try:
                admin_user = models.User(
                    email=admin_email,
                    password_hash=generate_password_hash("admin123"),
                    name="Administrator",
                    role="admin")
                db.session.add(admin_user)
                db.session.commit()
                print(f"[OK] Default admin user created: {admin_email}")
            except Exception as e:
                print(f"[ERROR] Failed to create admin user: {e}")
                db.session.rollback()
        else:
            print(f"[OK] Admin user already exists: {admin_email}")

        # Create default store profile if it doesn't exist
        try:
            db.session.rollback()  # Clear any pending transaction
            store_profile = models.StoreProfile.get_active_profile()
            if not store_profile:
                store_profile = models.StoreProfile(
                    store_name='Hurtrock Music Store',
                    store_tagline='Toko Alat Musik Terpercaya',
                    store_address=
                    'Jl. Musik Raya No. 123, RT/RW 001/002, Kelurahan Musik, Kecamatan Harmoni',
                    store_city='Jakarta Selatan',
                    store_postal_code='12345',
                    store_phone='0821-1555-8035',
                    store_email='info@hurtrock.com',
                    store_website='https://hurtrock.com',
                    whatsapp_number='6282115558035',
                    operating_hours=
                    'Senin - Sabtu: 09:00 - 21:00\nMinggu: 10:00 - 18:00',
                    branch_name='Cabang Pusat',
                    branch_code='HRT-001')
                db.session.add(store_profile)
                db.session.commit()
                print("[OK] Default store profile created")
            else:
                print("[OK] Store profile already exists")
        except Exception as e:
            print(f"[ERROR] Failed to create store profile: {e}")
            db.session.rollback()

        # Create sample data
        create_sample_data()
        print("[OK] Flask database initialization completed")

        # Setup and start Django chat service - DISABLED for now
        # if setup_django_chat_service():
        #     print("[OK] Django chat service setup and started successfully")
        # else:
        #     print(
        #         "[WARNING] Django chat service setup failed, chat may not work properly"
        #     )
        print("[INFO] Django chat service disabled - Flask-only mode")

    except Exception as e:
        print(f"[ERROR] Database initialization error: {e}")
        db.session.rollback()

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    return render_template(
        'error.html',
        error_code=404,
        error_message="Halaman yang Anda cari tidak ditemukan."), 404


@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template(
        'error.html',
        error_code=500,
        error_message="Terjadi kesalahan server. Silakan coba lagi nanti."
    ), 500


# SEO Routes
@app.route('/robots.txt')
def robots_txt():
    """Generate robots.txt for SEO"""
    return render_template('robots.txt'), 200, {'Content-Type': 'text/plain'}


@app.route('/favicon.ico')
def favicon():
    """Serve favicon.ico from static directory"""
    return send_from_directory('static',
                               'favicon.ico',
                               mimetype='image/vnd.microsoft.icon')


@app.route('/sitemap.xml')
def sitemap_xml():
    """Generate sitemap.xml for SEO"""
    try:
        # Get all active products with slugs
        products = models.Product.query.filter_by(is_active=True).all()
        categories = models.Category.query.filter_by(is_active=True).all()

        # Ensure all products have slugs
        for product in products:
            if not product.slug:
                product.ensure_slug()

        db.session.commit()

        current_date = datetime.utcnow().strftime('%Y-%m-%d')

        response = render_template('sitemap.xml',
                                   products=products,
                                   categories=categories,
                                   current_date=current_date)

        return response, 200, {'Content-Type': 'application/xml'}

    except Exception as e:
        print(f"[ERROR] Sitemap generation failed: {e}")
        db.session.rollback()
        return f"<?xml version='1.0' encoding='UTF-8'?><urlset xmlns='http://www.sitemaps.org/schemas/sitemap/0.9'></urlset>", 200, {
            'Content-Type': 'application/xml'
        }

@app.route('/product-feed.xml')
def product_feed():
    # Ambil semua produk aktif
    products = Product.query.filter_by(is_active=True).all()

    # Render ke template XML
    xml = render_template(
        'product-feed.xml',
        products=products,
        current_date=datetime.utcnow().strftime("%Y-%m-%d")
    )

    # Kembalikan response sebagai XML
    return Response(xml, mimetype='application/xml')

# Routes
@app.route('/')
def index():
    products = models.Product.query.filter_by(is_active=True).limit(8).all()
    categories = models.Category.query.filter_by(is_active=True).all()
    return render_template('index.html',
                           products=products,
                           categories=categories)

@app.route('/return-policy')
def return_policy():
    return render_template('return-policy.html')

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(
        'static',          # folder tempat file .exe disimpan
        filename,
        as_attachment=True # paksa download
    )

@app.route('/products')
def products():
    category_id = request.args.get('category')
    search_query = request.args.get('search', '')

    query = models.Product.query.filter_by(is_active=True)

    if category_id:
        query = query.filter_by(category_id=category_id)

    if search_query:
        query = query.filter(models.Product.name.contains(search_query))

    products = query.all()
    categories = models.Category.query.filter_by(is_active=True).all()

    # Safely convert category_id to int
    current_category = None
    if category_id:
        try:
            current_category = int(category_id)
        except (ValueError, TypeError):
            current_category = None

    return render_template('products.html',
                           products=products,
                           categories=categories,
                           current_category=current_category,
                           search_query=search_query)


@app.route('/produk/<slug>')
def product_detail(slug):
    product = models.Product.query.filter_by(slug=slug).first_or_404()
    valid_until = (datetime.utcnow() + timedelta(days=30)).strftime('%Y-%m-%d')
    return render_template('product_detail.html',
                           product=product,
                           valid_until=valid_until)


@app.route('/produk/id/<int:product_id>')
def product_detail_by_id(product_id):
    product = models.Product.query.get_or_404(product_id)
    if product.slug:
        return redirect(url_for('product_detail', slug=product.slug), code=301)
    valid_until = (datetime.utcnow() + timedelta(days=30)).strftime('%Y-%m-%d')
    return render_template('product_detail.html',
                           product=product,
                           valid_until=valid_until)


# This route was moved to avoid duplication - see line 1868-1884 for the actual implementation


@app.route('/search')
def search():
    query = request.args.get('q', '').strip()
    search_query_param = request.args.get('search',
                                          '').strip()  # Alternative parameter
    category_id = request.args.get('category', '').strip()

    # Use either 'q' or 'search' parameter
    search_term_text = query or search_query_param

    try:
        # Build base query for active products
        search_query = models.Product.query.filter(
            models.Product.is_active == True)

        # Smart keyword search - flexible word order matching
        if search_term_text and len(search_term_text) >= 2:
            # Split search term into keywords
            keywords = search_term_text.lower().split()
            
            # Build OR conditions for each keyword across multiple fields
            conditions = []
            for keyword in keywords:
                keyword_pattern = f"%{keyword}%"
                conditions.append(
                    db.or_(
                        models.Product.name.ilike(keyword_pattern),
                        models.Product.description.ilike(keyword_pattern),
                        models.Product.brand.ilike(keyword_pattern),
                        models.Product.model.ilike(keyword_pattern),
                        models.Product.gtin.ilike(keyword_pattern)
                    )
                )
            
            # Combine all keyword conditions with AND (all keywords must match somewhere)
            if conditions:
                search_query = search_query.filter(db.and_(*conditions))

        # Add category filter if specified
        if category_id and category_id.isdigit():
            search_query = search_query.filter(
                models.Product.category_id == int(category_id))

        # Order by relevance if text search is active, otherwise by name
        if search_term_text and len(search_term_text) >= 2:
            # Get all matching products for relevance scoring
            all_products = search_query.all()
            
            # Score products by relevance
            scored_products = []
            for p in all_products:
                score = 0
                search_lower = search_term_text.lower()
                name_lower = (p.name or '').lower()
                brand_lower = (p.brand or '').lower()
                model_lower = (p.model or '').lower()
                gtin_lower = (p.gtin or '').lower()
                
                # Exact match in name gets highest score
                if search_lower == name_lower:
                    score += 100
                # Starts with search term
                elif name_lower.startswith(search_lower):
                    score += 50
                # Contains search term
                elif search_lower in name_lower:
                    score += 30
                
                # Brand/model matching
                if search_lower in brand_lower:
                    score += 20
                if search_lower in model_lower:
                    score += 20
                
                # GTIN exact match
                if search_lower == gtin_lower:
                    score += 80
                    
                # Keyword matching bonus
                keywords = search_term_text.lower().split()
                for kw in keywords:
                    if kw in name_lower:
                        score += 10
                    if kw in brand_lower:
                        score += 5
                    if kw in model_lower:
                        score += 5
                
                scored_products.append((p, score))
            
            # Sort by score descending
            scored_products.sort(key=lambda x: x[1], reverse=True)
            products = [p for p, score in scored_products[:10]]
        else:
            # For category-only filter, show more results
            products = search_query.order_by(
                models.Product.name).limit(50).all()

        result = []
        for p in products:
            # Check if product has multiple images for hover effect
            has_multiple_images = len(p.images) > 0
            last_image_url = p.images[-1].image_url if has_multiple_images else p.image_url
            
            result.append({
                'id': p.id,
                'name': p.name,
                'slug': p.slug,
                'price': str(p.price),
                'image_url': p.image_url or '/static/images/placeholder.jpg',
                'brand': p.brand or '',
                'description': p.description[:100] + '...' if p.description and len(p.description) > 100 else p.description or '',
                'category': p.category.name if p.category else '',
                'url': url_for('product_detail', slug=p.slug) if p.slug else url_for('product_detail_by_id', product_id=p.id),
                'has_multiple_images': has_multiple_images,
                'last_image_url': last_image_url
            })
        
        return jsonify(result)
    except Exception as e:
        print(f"Search error: {e}")
        return jsonify({'error': 'Search failed'}), 500

@app.route('/api/chat/upload-media', methods=['POST'])
@login_required
@csrf.exempt
def upload_chat_media():
    """Upload media files for chat (images and videos)"""
    try:
        print(f"[DEBUG] Upload request received from user {current_user.id}")

        # Handle request parsing with better error handling
        try:
            if not request.files:
                print("[ERROR] No files in request")
                return jsonify({
                    'success': False,
                    'error': 'No file provided'
                }), 400

            print(f"[DEBUG] Request files: {list(request.files.keys())}")
            print(f"[DEBUG] Request form: {dict(request.form)}")
        except Exception as parse_error:
            print(f"[ERROR] Failed to parse request: {parse_error}")
            return jsonify({
                'success': False,
                'error': 'Invalid request format'
            }), 400

        if 'file' not in request.files:
            print("[ERROR] No 'file' key in request.files")
            return jsonify({
                'success': False,
                'error': 'No file provided'
            }), 400

        file = request.files['file']
        print(
            f"[DEBUG] File received: {file.filename}, Content-Type: {file.content_type}"
        )

        if not file or file.filename == '' or file.filename is None:
            print("[ERROR] Empty filename or no file object")
            return jsonify({
                'success': False,
                'error': 'No file selected'
            }), 400

        if not allowed_chat_media(file.filename):
            print(f"[ERROR] File type not allowed: {file.filename}")
            return jsonify({
                'success': False,
                'error': 'File type not allowed'
            }), 400

        # Check file size with better error handling
        try:
            file.seek(0, os.SEEK_END)
            file_size = file.tell()
            file.seek(0)  # Reset file pointer
        except Exception as size_error:
            print(f"[ERROR] Failed to check file size: {size_error}")
            return jsonify({
                'success': False,
                'error': 'Unable to process file'
            }), 400

        if file_size > 10 * 1024 * 1024:  # 10MB limit
            print(f"[ERROR] File too large: {file_size} bytes")
            return jsonify({
                'success': False,
                'error': 'File size too large (max 10MB)'
            }), 400

        print(f"[DEBUG] File size: {file_size} bytes")

        # Get file extension while preserving original format
        original_filename = secure_filename(file.filename)
        file_ext = original_filename.rsplit(
            '.', 1)[1].lower() if '.' in original_filename else 'jpg'

        # Generate consistent filename with timestamp
        timestamp = datetime.utcnow().strftime(
            '%Y%m%d_%H%M%S_%f')[:19]  # YYYYMMDD_HHMMSS_microseconds
        user_role = current_user.role if hasattr(current_user,
                                                 'role') else 'user'
        unique_filename = f"chat_{user_role}_{timestamp}_{current_user.id}.{file_ext}"

        # Ensure upload directory exists with proper permissions
        upload_dir = app.config['UPLOADS_MEDIA_FOLDER']
        try:
            if not os.path.exists(upload_dir):
                os.makedirs(upload_dir, mode=0o755, exist_ok=True)
                print(f"[DEBUG] Created directory: {upload_dir}")

            # Verify directory permissions
            if not os.access(upload_dir, os.W_OK):
                os.chmod(upload_dir, 0o755)
                print(f"[DEBUG] Fixed permissions for directory: {upload_dir}")

        except Exception as dir_error:
            print(f"[ERROR] Directory creation failed: {dir_error}")
            return jsonify({
                'success':
                False,
                'error':
                f'Failed to create upload directory: {str(dir_error)}'
            }), 500

        # Save file to uploads/medias_sends folder with better error handling
        file_path = os.path.join(upload_dir, unique_filename)
        try:
            # Save file in chunks to handle large files
            with open(file_path, 'wb') as f:
                file.seek(0)
                while True:
                    chunk = file.read(8192)  # 8KB chunks
                    if not chunk:
                        break
                    f.write(chunk)

            print(f"[DEBUG] File saved successfully: {file_path}")

            # Verify file was saved
            if not os.path.exists(file_path):
                raise Exception(f"File was not saved properly: {file_path}")

            # Check saved file size
            saved_size = os.path.getsize(file_path)
            print(f"[DEBUG] Saved file size: {saved_size} bytes")

        except Exception as save_error:
            print(f"[ERROR] File save failed: {save_error}")
            return jsonify({
                'success': False,
                'error': f'Failed to save file: {str(save_error)}'
            }), 500

        # Determine if it's image or video
        is_video = is_video_file(original_filename)
        media_type = 'video' if is_video else 'image'

        # Compress images if needed (for images only)
        if not is_video:
            try:
                compress_image(file_path, max_size_mb=2)
                print(f"[DEBUG] Image compressed: {file_path}")
            except Exception as e:
                print(f"[WARNING] Image compression failed: {e}")

        # Return media info dengan URL yang benar
        media_url = f"/uploads/medias_sends/{unique_filename}"

        result = {
            'success': True,
            'media_url': media_url,
            'media_type': media_type,
            'filename': unique_filename,
            'original_filename': original_filename
        }
        print(f"[DEBUG] Upload successful: {result}")
        print(f"[DEBUG] Media file saved at: {file_path}")
        print(f"[DEBUG] Media URL accessible at: {media_url}")

        return jsonify(result), 200

    except Exception as e:
        print(f"[ERROR] Media upload failed: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'Upload failed: {str(e)}'
        }), 500


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':

        email = request.form['email']
        password = request.form['password']
        name = request.form['name']

        # Check if user already exists
        if models.User.query.filter_by(email=email).first():
            flash('Email sudah terdaftar. Silakan login.', 'error')
            return redirect(url_for('register'))

        # Create new user
        hashed_password = generate_password_hash(password)
        user = models.User(email=email,
                           password_hash=hashed_password,
                           name=name)

        db.session.add(user)
        db.session.commit()

        flash('Akun berhasil dibuat! Silakan login.', 'success')
        return redirect(url_for('login'))

    return render_template('register.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        user = models.User.query.filter_by(email=email).first()

        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            
            # --- LOGIKA PENGALIHAN BERDASARKAN ROLE DITAMBAHKAN DI SINI ---
            
            # 1. Prioritaskan pengalihan 'next' jika ada
            next_page = request.args.get('next')
            if next_page:
                return redirect(next_page)

            # 2. Pengalihan berdasarkan peran (role)
            if user.role == 'admin':
                # Alihkan ke dashboard admin (misal: /admin/dashboard)
                return redirect(url_for('admin_dashboard'))
            elif user.role == 'cashier':
                # Alihkan ke dashboard cashier/kasir (misal: /cashier/dashboard)
                return redirect(url_for('cashier_dashboard'))
            elif user.role == 'user':
                # Alihkan ke dashboard user/pelanggan (misal: /user/dashboard)
                return redirect(url_for('user_dashboard'))
            else:
                # Fallback jika peran tidak dikenal
                return redirect(url_for('index'))
            
            # --- AKHIR LOGIKA PENGALIHAN BERDASARKAN ROLE ---

        else:
            flash('Email atau password salah.', 'error')

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Anda telah logout.', 'info')
    return redirect(url_for('index'))


@app.route('/cart')
@login_required
def cart():
    cart_items = models.CartItem.query.filter_by(user_id=current_user.id).all()
    total = sum(item.quantity * item.product.price for item in cart_items)
    return render_template('cart.html', cart_items=cart_items, total=total)


@app.route('/add_to_cart/<int:product_id>', methods=['POST'])
@login_required
def add_to_cart(product_id):
    product = models.Product.query.get_or_404(product_id)
    quantity = int(request.form.get('quantity', 1))

    # Validate quantity is positive
    if quantity <= 0:
        flash('Jumlah tidak valid!', 'error')
        return redirect(
            url_for('product_detail', slug=product.slug) if product.
            slug else url_for('product_detail_by_id', product_id=product_id))

    # Check if item already in cart
    cart_item = models.CartItem.query.filter_by(user_id=current_user.id,
                                                product_id=product_id).first()

    # Calculate total quantity that would be in cart
    total_quantity = quantity
    if cart_item:
        total_quantity += cart_item.quantity

    # Check stock availability
    if total_quantity > product.stock_quantity:
        flash(
            f'Stok tidak mencukupi! Stok tersedia: {product.stock_quantity}, di keranjang: {cart_item.quantity if cart_item else 0}',
            'error')
        return redirect(
            url_for('product_detail', slug=product.slug) if product.
            slug else url_for('product_detail_by_id', product_id=product_id))

    if cart_item:
        cart_item.quantity += quantity
    else:
        cart_item = models.CartItem(user_id=current_user.id,
                                    product_id=product_id,
                                    quantity=quantity)
        db.session.add(cart_item)

    db.session.commit()
    flash(f'{product.name} ditambahkan ke keranjang!', 'success')

    return redirect(
        url_for('product_detail', slug=product.slug) if product.
        slug else url_for('product_detail_by_id', product_id=product_id))


@app.route('/update_cart/<int:item_id>', methods=['POST'])
@login_required
def update_cart(item_id):
    cart_item = models.CartItem.query.filter_by(
        id=item_id, user_id=current_user.id).first_or_404()

    quantity = int(request.form.get('quantity', 1))

    if quantity > 0:
        # Check stock availability for the new quantity
        if quantity > cart_item.product.stock_quantity:
            flash(
                f'Stok tidak mencukupi untuk {cart_item.product.name}! Stok tersedia: {cart_item.product.stock_quantity}',
                'error')
            return redirect(url_for('cart'))
        cart_item.quantity = quantity
    else:
        db.session.delete(cart_item)

    db.session.commit()
    return redirect(url_for('cart'))


@app.route('/remove_from_cart/<int:item_id>')
@login_required
def remove_from_cart(item_id):
    cart_item = models.CartItem.query.filter_by(
        id=item_id, user_id=current_user.id).first_or_404()

    db.session.delete(cart_item)
    db.session.commit()
    flash('Item dihapus dari keranjang.', 'info')

    return redirect(url_for('cart'))


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        current_user.name = request.form['name']
        current_user.phone = request.form['phone']
        current_user.address = request.form['address']

        db.session.commit()
        flash('Profile berhasil diperbarui!', 'success')

        # Redirect to checkout if came from there
        next_url = request.args.get('next')
        if next_url:
            return redirect(next_url)
        return redirect(url_for('profile'))

    return render_template('profile.html')


@app.route('/checkout')
@login_required
def checkout():
    # Check if buyer has complete profile (for buyers only)
    if current_user.is_buyer and (not current_user.address
                                  or not current_user.phone):
        flash(
            'Silakan lengkapi profile Anda terlebih dahulu sebelum melakukan pembelian.',
            'warning')
        return redirect(url_for('profile', next=url_for('checkout')))

    cart_items = models.CartItem.query.filter_by(user_id=current_user.id).all()

    if not cart_items:
        flash('Keranjang kosong!', 'error')
        return redirect(url_for('cart'))

    subtotal = sum(item.quantity * item.product.price for item in cart_items)

    # Calculate shipping weight and volume
    total_weight = sum(item.quantity * (item.product.weight or 0)
                       for item in cart_items)
    total_volume = sum(item.quantity * (item.product.volume_cm3 or 0)
                       for item in cart_items)

    # Get available shipping services
    shipping_services = models.ShippingService.query.filter_by(
        is_active=True).all()

    # Calculate shipping costs for each service
    shipping_options = []
    for service in shipping_services:
        cost = service.calculate_shipping_cost(total_weight, total_volume)
        shipping_options.append({
            'service':
            service,
            'cost':
            cost,
            'delivery_estimate':
            f"{service.min_days}-{service.max_days} hari"
        })

    # Get active payment configurations
    payment_configs = models.PaymentConfiguration.query.filter_by(
        is_active=True).all()

    # If no active payment config, show error
    if not payment_configs:
        flash(
            'Tidak ada metode pembayaran yang tersedia. Silakan hubungi admin.',
            'error')
        return redirect(url_for('cart'))

    return render_template(
        'checkout.html',
        cart_items=cart_items,
        subtotal=subtotal,
        total=
        subtotal,  # Add total variable (will be updated by JS when shipping is selected)
        total_weight=total_weight,
        total_volume=total_volume,
        shipping_options=shipping_options,
        payment_configs=payment_configs)


@app.route('/create-checkout-session', methods=['POST'])
@login_required
def create_checkout_session():
    cart_items = models.CartItem.query.filter_by(user_id=current_user.id).all()

    if not cart_items:
        return jsonify({'error': 'Keranjang kosong'}), 400

    # Get selected payment configuration
    payment_config_id = request.form.get('payment_config_id')
    if not payment_config_id:
        flash('Silakan pilih metode pembayaran!', 'error')
        return redirect(url_for('checkout'))

    payment_config = models.PaymentConfiguration.query.get_or_404(
        int(payment_config_id))

    if not payment_config.is_active:
        flash('Metode pembayaran yang dipilih tidak aktif!', 'error')
        return redirect(url_for('checkout'))

    # Get selected shipping service
    shipping_service_id = request.form.get('shipping_service_id')
    if not shipping_service_id:
        flash('Silakan pilih jasa kirim!', 'error')
        return redirect(url_for('checkout'))

    shipping_service = models.ShippingService.query.get_or_404(
        int(shipping_service_id))

    # Calculate costs
    subtotal = sum(item.quantity * item.product.price for item in cart_items)
    total_weight = sum(item.quantity * (item.product.weight or 0)
                       for item in cart_items)
    total_volume = sum(item.quantity * (item.product.volume_cm3 or 0)
                       for item in cart_items)
    shipping_cost = shipping_service.calculate_shipping_cost(
        total_weight, total_volume)
    total_amount = float(subtotal) + float(shipping_cost)

    # Store order info in session
    session['shipping_service_id'] = shipping_service_id
    session['shipping_cost'] = float(shipping_cost)
    session['payment_config_id'] = payment_config_id

    try:
        YOUR_DOMAIN = os.environ.get('DOMAINS', 'localhost:5000')

        if payment_config.provider == 'stripe':
            return _create_stripe_checkout(cart_items, shipping_service,
                                           shipping_cost, total_amount,
                                           YOUR_DOMAIN, payment_config)
        elif payment_config.provider == 'midtrans':
            return _create_midtrans_checkout(cart_items, shipping_service,
                                             shipping_cost, total_amount,
                                             YOUR_DOMAIN, payment_config)
        else:
            flash('Metode pembayaran tidak didukung!', 'error')
            return redirect(url_for('checkout'))

    except Exception as e:
        flash(f'Error dalam memproses pembayaran: {str(e)}', 'error')
        return redirect(url_for('cart'))


def _create_stripe_checkout(cart_items, shipping_service, shipping_cost,
                            total_amount, domain, payment_config):
    """Create Stripe checkout session"""

    # Set Stripe API key from config with fallback
    api_key = payment_config.stripe_secret_key or os.environ.get(
        'STRIPE_SECRET_KEY')
    if not api_key:
        raise ValueError(
            "No Stripe API key configured. Please configure payment settings in admin panel."
        )

    stripe.api_key = api_key

    # Build line items
    line_items = []
    for item in cart_items:
        line_items.append({
            'price_data': {
                'currency': 'idr',
                'product_data': {
                    'name': item.product.name,
                },
                'unit_amount':
                int(item.product.price * 100),  # Convert to cents
            },
            'quantity': item.quantity,
        })

    # Add shipping as line item
    line_items.append({
        'price_data': {
            'currency': 'idr',
            'product_data': {
                'name': f'Ongkos Kirim - {shipping_service.name}',
            },
            'unit_amount': int(shipping_cost * 100),  # Convert to cents
        },
        'quantity': 1,
    })

    checkout_session = stripe.checkout.Session.create(
        line_items=line_items,
        mode='payment',
        success_url=f'https://{domain}/payment-success',
        cancel_url=f'https://{domain}/cart',
        customer_email=current_user.email,
    )

    return redirect(checkout_session.url, code=303)


def _create_midtrans_checkout(cart_items, shipping_service, shipping_cost,
                              total_amount, domain, payment_config):
    """Create Midtrans checkout session"""
    import uuid

    # Create Snap API instance
    snap = midtransclient.Snap(is_production=not payment_config.is_sandbox,
                               server_key=payment_config.midtrans_server_key,
                               client_key=payment_config.midtrans_client_key)

    # Generate unique order ID
    order_id = f"ORDER-{current_user.id}-{int(datetime.utcnow().timestamp())}-{str(uuid.uuid4())[:8]}"

    # Build item details
    item_details = []
    for item in cart_items:
        item_details.append({
            'id': str(item.product.id),
            'price': int(item.product.price),
            'quantity': item.quantity,
            'name': item.product.name[:50]  # Midtrans has character limit
        })

    # Add shipping cost
    item_details.append({
        'id': 'shipping',
        'price': int(shipping_cost),
        'quantity': 1,
        'name': f'Ongkir {shipping_service.name}'
    })

    # Customer details
    customer_details = {
        'first_name':
        current_user.name.split()[0] if current_user.name else 'Customer',
        'last_name':
        ' '.join(current_user.name.split()[1:])
        if current_user.name and len(current_user.name.split()) > 1 else '',
        'email':
        current_user.email,
        'phone':
        current_user.phone or '081234567890',
        'billing_address': {
            'address': current_user.address or 'Jakarta',
            'city': 'Jakarta',
            'postal_code': '12345',
            'country_code': 'IDN'
        },
        'shipping_address': {
            'address': current_user.address or 'Jakarta',
            'city': 'Jakarta',
            'postal_code': '12345',
            'country_code': 'IDN'
        }
    }

    # Transaction details
    transaction_details = {
        'order_id': order_id,
        'gross_amount': int(total_amount)
    }

    # Parameter for Snap API
    param = {
        'transaction_details': transaction_details,
        'item_details': item_details,
        'customer_details': customer_details,
        'callbacks': {
            'finish':
            payment_config.callback_finish_url
            or f'https://{domain}/payment/finish',
            'unfinish':
            payment_config.callback_unfinish_url
            or f'https://{domain}/payment/unfinish',
            'error':
            payment_config.callback_error_url
            or f'https://{domain}/payment/error'
        }
    }

    # Create transaction
    transaction = snap.create_transaction(param)

    # Store order ID in session for callback handling
    session['midtrans_order_id'] = order_id

    # Redirect to Snap payment page
    return redirect(transaction['redirect_url'], code=303)


@app.route('/payment-success')
@login_required
def payment_success():
    # Create order from cart items
    cart_items = models.CartItem.query.filter_by(user_id=current_user.id).all()

    if cart_items:
        subtotal = sum(item.quantity * item.product.price
                       for item in cart_items)

        # Get shipping info from session
        shipping_service_id = session.get('shipping_service_id')
        shipping_cost = session.get('shipping_cost', 0)

        total_amount = float(subtotal) + float(shipping_cost)

        # Calculate estimated delivery days
        estimated_delivery_days = 0
        if shipping_service_id:
            shipping_service = models.ShippingService.query.get(
                int(shipping_service_id))
            if shipping_service:
                estimated_delivery_days = shipping_service.max_days

        order = models.Order(user_id=current_user.id,
                             total_amount=total_amount,
                             shipping_cost=shipping_cost,
                             shipping_service_id=int(shipping_service_id)
                             if shipping_service_id else None,
                             estimated_delivery_days=estimated_delivery_days,
                             shipping_address=current_user.address,
                             status='paid',
                             created_at=datetime.utcnow())
        db.session.add(order)
        db.session.flush()  # To get the order ID

        # Add order items and reduce stock quantities
        for cart_item in cart_items:
            # Check stock availability before reducing
            product = cart_item.product
            if product.stock_quantity < cart_item.quantity:
                db.session.rollback()
                flash(
                    f'Stok tidak mencukupi untuk produk {product.name}. Stok tersedia: {product.stock_quantity}',
                    'error')
                return redirect(url_for('cart'))

            # Create order item
            order_item = models.OrderItem(order_id=order.id,
                                          product_id=cart_item.product_id,
                                          quantity=cart_item.quantity,
                                          price=cart_item.product.price)
            db.session.add(order_item)

            # Reduce stock quantity
            product.stock_quantity -= cart_item.quantity
            print(
                f"[STOCK] Reduced stock for {product.name}: {product.stock_quantity + cart_item.quantity} -> {product.stock_quantity}"
            )

        # Clear cart
        for cart_item in cart_items:
            db.session.delete(cart_item)

        db.session.commit()

        flash('Pembayaran berhasil! Terima kasih atas pesanan Anda.',
              'success')

    return render_template('payment_success.html',
                           current_datetime=datetime.utcnow())


@app.route('/orders')
@login_required
def orders():
    user_orders = models.Order.query.filter_by(
        user_id=current_user.id).order_by(
            models.Order.created_at.desc()).all()
    return render_template('orders.html', orders=user_orders)


@app.route('/store-info')
def store_info():
    return render_template('store_info.html')


@app.route('/uploads/medias_sends/<path:filename>')
def serve_media_file(filename):
    """Serve uploaded media files"""
    try:
        print(f"[MEDIA] Serving media file: {filename}")
        print(f"[MEDIA] From directory: {app.config['UPLOADS_MEDIA_FOLDER']}")
        print(
            f"[MEDIA] Full path: {os.path.join(app.config['UPLOADS_MEDIA_FOLDER'], filename)}"
        )

        if not os.path.exists(
                os.path.join(app.config['UPLOADS_MEDIA_FOLDER'], filename)):
            print(f"[MEDIA] File not found: {filename}")
            return jsonify({'error': 'File not found'}), 404

        response = send_from_directory(app.config['UPLOADS_MEDIA_FOLDER'],
                                       filename)
        # Add CORS headers for cross-origin access
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        # Add cache headers for better performance
        response.headers['Cache-Control'] = 'public, max-age=86400'  # 1 day
        print(f"[MEDIA] Successfully served: {filename}")
        return response
    except Exception as e:
        print(f"[MEDIA] Error serving file {filename}: {e}")
        return jsonify({'error': 'File access error'}), 500

@app.route('/static/chat_media/<path:filename>')
def serve_chat_media(filename):
    """Serve chat media files (alternative endpoint)"""
    try:
        response = send_from_directory(app.config['UPLOADS_MEDIA_FOLDER'],
                                       filename)
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        response.headers['Cache-Control'] = 'public, max-age=31536000'
        return response
    except FileNotFoundError:
        return jsonify({'error': 'File not found'}), 404


# WebSocket proxy routes for Cloudflare tunnel compatibility
@app.route('/ws/chat/<path:path>', methods=['GET'])
def websocket_proxy(path):
    """Proxy WebSocket connections to Django service"""
    try:
        import requests
        from flask import Response

        # Check if this is a WebSocket upgrade request
        if request.headers.get('Upgrade', '').lower() == 'websocket':
            # This should be handled by Cloudflare tunnel routing directly to Django
            return Response(
                'WebSocket connections should be routed directly to Django service',
                status=426,
                headers={
                    'Upgrade':
                    'websocket',
                    'Connection':
                    'Upgrade',
                    'Access-Control-Allow-Origin':
                    '*',
                    'Access-Control-Allow-Methods':
                    'GET, POST, OPTIONS',
                    'Access-Control-Allow-Headers':
                    'Origin, Content-Type, Accept, Authorization, X-Requested-With'
                })

        # For non-WebSocket requests, check Django service availability
        try:
            health_check = requests.get('http://127.0.0.1:8000/health/',
                                        timeout=2)
            if health_check.status_code != 200:
                return jsonify({'error': 'Chat service unavailable'}), 503
        except:
            return jsonify({'error': 'Chat service unavailable'}), 503

        return jsonify({
            'status': 'WebSocket endpoint available',
            'message': 'Use WebSocket protocol to connect',
            'django_url': f"ws://127.0.0.1:8000/ws/chat/{path}"
        }), 200

    except Exception as e:
        return jsonify({'error': f'WebSocket proxy error: {str(e)}'}), 500


# Chat API proxy routes
@app.route('/api/rooms/<path:path>', methods=['GET', 'POST'])
def chat_api_proxy(path):
    """Proxy chat API requests to Django service"""
    try:
        import requests

        # Try multiple Django service endpoints
        django_endpoints = [
            'http://127.0.0.1:8000', 'http://localhost:8000',
            'http://0.0.0.0:8000'
        ]

        django_url = None
        for endpoint in django_endpoints:
            try:
                health_check = requests.get(f'{endpoint}/health/', timeout=2)
                if health_check.status_code == 200:
                    django_url = f"{endpoint}/api/rooms/{path}"
                    break
            except:
                continue

        if not django_url:
            return jsonify({
                'error': 'Chat service not available',
                'results': []
            }), 503

        # Add JWT token if user is authenticated
        headers = {'Content-Type': 'application/json'}

        if current_user.is_authenticated:
            headers[
                'Authorization'] = f'Bearer {generate_jwt_token(current_user)}'
        elif request.headers.get('Authorization'):
            headers['Authorization'] = request.headers.get('Authorization')

        if request.method == 'GET':
            # Forward GET parameters
            params = request.args.to_dict()
            response = requests.get(django_url,
                                    params=params,
                                    headers=headers,
                                    timeout=10)
        else:  # POST
            response = requests.post(
                django_url,
                json=request.get_json() if request.is_json else None,
                data=request.form if not request.is_json else None,
                headers=headers,
                timeout=10)

        return jsonify(response.json()), response.status_code

    except requests.RequestException as e:
        print(f"[ERROR] Chat API proxy error: {e}")
        return jsonify({
            'error': 'Chat service unavailable',
            'results': []
        }), 503
    except Exception as e:
        print(f"[ERROR] Unexpected chat API error: {e}")
        return jsonify({
            'error': f'API proxy error: {str(e)}',
            'results': []
        }), 500


# Admin API proxy routes
@app.route('/api/admin/<path:path>', methods=['GET', 'POST'])
def admin_api_proxy(path):
    """Proxy admin API requests to Django service"""
    try:
        import requests

        # Generate JWT token for current user
        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        }

        if current_user.is_authenticated:
            jwt_token = generate_jwt_token(current_user)
            headers['Authorization'] = f'Bearer {jwt_token}'
            print(f"[DEBUG] Admin API proxy for {current_user.email} - {path}")

        django_url = f"http://127.0.0.1:8000/api/admin/{path}"

        if request.method == 'GET':
            # Forward GET parameters
            params = request.args.to_dict()
            response = requests.get(django_url,
                                    params=params,
                                    headers=headers,
                                    timeout=10)
        else:  # POST
            response = requests.post(
                django_url,
                json=request.get_json() if request.is_json else None,
                data=request.form if not request.is_json else None,
                headers=headers,
                timeout=10)

        print(f"[DEBUG] Admin API response: {response.status_code}")
        return jsonify(response.json()), response.status_code

    except requests.RequestException as e:
        print(f"[ERROR] Admin API proxy error: {e}")
        return jsonify({
            'error': 'Chat service unavailable',
            'results': []
        }), 503
    except Exception as e:
        print(f"[ERROR] Unexpected admin API error: {e}")
        return jsonify({'error': f'Admin API proxy error: {str(e)}'}), 500


@app.route('/api/admin/buyer-rooms/', methods=['GET'])
@login_required
@admin_required
def admin_buyer_rooms_proxy():
    """Proxy admin buyer rooms API to Django service"""
    try:
        import requests

        # Generate JWT token for current user
        jwt_token = generate_jwt_token(current_user)
        print(
            f"[DEBUG] Generated JWT token for admin {current_user.email}: {jwt_token[:50]}..."
        )

        # Forward request to Django service
        search_query = request.args.get('search', '')
        django_url = f"http://127.0.0.1:8000/api/admin/buyer-rooms/"
        if search_query:
            django_url += f"?search={search_query}"

        headers = {
            'Authorization': f'Bearer {jwt_token}',
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        }

        print(f"[DEBUG] Forwarding admin request to: {django_url}")
        print(f"[DEBUG] Headers: {headers}")

        response = requests.get(django_url, headers=headers, timeout=10)

        print(f"[DEBUG] Django response status: {response.status_code}")
        if response.status_code != 200:
            print(f"[DEBUG] Django response error: {response.text}")

        return jsonify(response.json()), response.status_code

    except requests.RequestException as e:
        print(f"[ERROR] Admin rooms API proxy error: {e}")
        return jsonify({
            'error': 'Chat service unavailable',
            'rooms': [],
            'total_count': 0
        }), 503
    except Exception as e:
        print(f"[ERROR] Unexpected admin API error: {e}")
        return jsonify({'error': f'Admin API proxy error: {str(e)}'}), 500

# New: Rute baru untuk mengambil semua data produk (ringan) untuk pencarian di klien
@app.route('/api/admin/products/search')
@login_required
@staff_required
def api_admin_search_products():
    """API endpoint for searching products (for manual invoice) - supports ID, GTIN, name"""
    try:
        query = request.args.get('q', '').strip()
        
        if not query or len(query) < 2:
            # Return empty if query too short
            return jsonify({'success': True, 'products': []})
        
        # Sederhanakan Kondisi Pencarian (Fokus pada ID, GTIN, dan Name)
        search_conditions = []
        
        # 1. Exact ID match (prioritize)
        if query.isdigit():
            search_conditions.append(models.Product.id == int(query))
        
        # 2. GTIN exact or partial match
        search_conditions.append(models.Product.gtin.ilike(f'{query}%')) # Diubah: Cari yang berawal dari query (lebih cepat)
        
        # 3. Name partial match
        search_conditions.append(models.Product.name.ilike(f'%{query}%'))
        
        # Hapus Brand dan Description dari pencarian
        
        # Lakukan Query dengan Pengurutan Sederhana
        products = models.Product.query.filter(
            models.Product.is_active == True,
            db.or_(*search_conditions)
        ).order_by(
            models.Product.name # Urutkan berdasarkan Nama Produk (Lebih Cepat)
        ).limit(50).all() # Batasan 50 tetap dipertahankan
        
        # Format response (tetap sama)
        result = []
        for product in products:
            result.append({
                'id': product.id,
                'name': product.name,
                'brand': product.brand or '',
                'price': float(product.price),
                'stock': product.stock_quantity,
                'gtin': product.gtin or '',
                # Kurangi informasi yang tidak terlalu penting di modal jika memang masih lambat
                'description': (product.description or '')[:100]  
            })
        
        return jsonify({'success': True, 'products': result})
        
    except Exception as e:
        print(f"[ERROR] Failed to search products: {str(e)}")
        # Coba kembalikan respons 500, ini membantu Anda melihat error di log server
        return jsonify({'success': False, 'error': str(e)}), 500
        
# Admin Routes
@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    total_products = models.Product.query.count()
    total_orders = models.Order.query.count()
    total_users = models.User.query.count()

    recent_orders = models.Order.query.order_by(
        models.Order.created_at.desc()).limit(5).all()

    # Analisis penjualan
    from sqlalchemy import func, extract
    from decimal import Decimal

    # Total penjualan hari ini
    today = datetime.utcnow().date()
    today_sales = db.session.query(func.sum(models.Order.total_amount)).filter(
        func.date(models.Order.created_at) == today,
        models.Order.status.in_(['paid', 'shipped', 'delivered'
                                 ])).scalar() or Decimal('0')

    # Total penjualan bulan ini
    current_month = datetime.utcnow().month
    current_year = datetime.utcnow().year
    monthly_sales = db.session.query(func.sum(
        models.Order.total_amount)).filter(
            extract('month', models.Order.created_at) == current_month,
            extract('year', models.Order.created_at) == current_year,
            models.Order.status.in_(['paid', 'shipped', 'delivered'
                                     ])).scalar() or Decimal('0')

    # Produk terlaris dengan explicit join
    best_selling_products = db.session.query(
        models.Product.name,
        func.sum(models.OrderItem.quantity).label('total_sold')
    ).select_from(models.Product)\
    .join(models.OrderItem, models.OrderItem.product_id == models.Product.id)\
    .join(models.Order, models.Order.id == models.OrderItem.order_id)\
    .filter(
        models.Order.status.in_(['paid', 'shipped', 'delivered'])
    ).group_by(models.Product.id, models.Product.name).order_by(
        func.sum(models.OrderItem.quantity).desc()
    ).limit(5).all()

    return render_template('admin/dashboard.html',
                           total_products=total_products,
                           total_orders=total_orders,
                           total_users=total_users,
                           recent_orders=recent_orders,
                           current_date=datetime.utcnow(),
                           today_sales=today_sales,
                           monthly_sales=monthly_sales,
                           best_selling_products=best_selling_products)


@app.route('/admin/products')
@login_required
@admin_required
def admin_products():
    # Get all products ordered by stock status (critical first)
    products = models.Product.query.order_by(
        db.case(
            (models.Product.stock_quantity <= 0, 1),
            (models.Product.stock_quantity <= models.Product.minimum_stock, 2),
            (models.Product.stock_quantity
             <= models.Product.low_stock_threshold, 3),
            else_=4), models.Product.name).all()

    # Ensure all products have slugs
    slug_updated = False
    for product in products:
        if not product.slug:
            product.ensure_slug()
            slug_updated = True

    if slug_updated:
        try:
            db.session.commit()
            print(f"[INFO] Updated slugs for products without slugs")
        except Exception as e:
            print(f"[ERROR] Failed to update slugs: {e}")
            db.session.rollback()

    categories = models.Category.query.filter_by(is_active=True).all()
    suppliers = models.Supplier.query.filter_by(is_active=True).all()

    # Get products with different stock levels
    out_of_stock_products = models.Product.query.filter(
        models.Product.stock_quantity <= 0).all()
    critical_stock_products = models.Product.query.filter(
        models.Product.stock_quantity > 0, models.Product.stock_quantity
        <= models.Product.minimum_stock).all()
    low_stock_products = models.Product.query.filter(
        models.Product.stock_quantity > models.Product.minimum_stock,
        models.Product.stock_quantity
        <= models.Product.low_stock_threshold).all()

    # Combined alert products (critical + low stock)
    alert_products = critical_stock_products + low_stock_products

    return render_template('admin/products.html',
                           products=products,
                           categories=categories,
                           suppliers=suppliers,
                           out_of_stock_products=out_of_stock_products,
                           critical_stock_products=critical_stock_products,
                           low_stock_products=low_stock_products,
                           alert_products=alert_products)


@app.route('/admin/products/add', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_add_product():
    if request.method == 'GET':
        # Show add product form
        categories = models.Category.query.filter_by(is_active=True).all()
        suppliers = models.Supplier.query.filter_by(is_active=True).all()
        return render_template('admin/add_product.html',
                               categories=categories,
                               suppliers=suppliers)

    try:
        # Validate required fields
        if not request.form.get('name') or not request.form.get(
                'price') or not request.form.get('category_id'):
            flash('Nama produk, harga, dan kategori wajib diisi!', 'error')
            categories = models.Category.query.filter_by(is_active=True).all()
            suppliers = models.Supplier.query.filter_by(is_active=True).all()
            return render_template('admin/add_product.html',
                                   categories=categories,
                                   suppliers=suppliers)

        name = request.form['name'].strip()
        description = request.form.get('description', '').strip()
        price = float(request.form['price'])
        stock_quantity = int(request.form.get('stock_quantity', 0))
        brand = request.form.get('brand', '').strip()
        model = request.form.get('model', '').strip()
        gtin = request.form.get('gtin', '').strip()
        category_id = int(request.form['category_id'])
        supplier_id = int(request.form['supplier_id']) if request.form.get(
            'supplier_id') else None
        is_featured = 'is_featured' in request.form
        is_active = 'is_active' in request.form
        
        # Auto-generate GTIN if not provided
        if not gtin:
            import time
            timestamp = str(int(time.time() * 1000))
            gtin = f"hrtbrg{timestamp}"

        # Handle dimensions and weight
        weight = float(request.form.get('weight', 0)) if request.form.get('weight') else 0
        shipping_weight = float(request.form.get('shipping_weight', 0)) if request.form.get('shipping_weight') else weight
        length = float(request.form.get('length', 0)) if request.form.get('length') else 0
        width = float(request.form.get('width', 0)) if request.form.get('width') else 0
        height = float(request.form.get('height', 0)) if request.form.get('height') else 0

        # Handle stock management thresholds
        minimum_stock = int(request.form.get('minimum_stock', 5))
        low_stock_threshold = int(request.form.get('low_stock_threshold', 10))

        # Create new product
        new_product = models.Product(name=name,
                                     description=description,
                                     price=price,
                                     stock_quantity=stock_quantity,
                                     minimum_stock=minimum_stock,
                                     low_stock_threshold=low_stock_threshold,
                                     brand=brand,
                                     model=model,
                                     gtin=gtin,
                                     category_id=category_id,
                                     supplier_id=supplier_id,
                                     weight=weight,
                                     length=length,
                                     width=width,
                                     height=height,
                                     is_featured=is_featured,
                                     is_active=is_active)

        db.session.add(new_product)
        db.session.flush()  # Get product ID

        # Generate slug for SEO-friendly URL
        new_product.ensure_slug()

        # Handle multiple images upload if provided
        uploaded_images = []
        processed_images = []  # Store image info for sorting
        selected_thumbnail_index = int(
            request.form.get('selected_thumbnail', 0))

        print(f"[DEBUG] Processing images for product {new_product.id}")
        print(f"[DEBUG] Selected thumbnail index: {selected_thumbnail_index}")

        if 'images' in request.files:
            files = request.files.getlist('images')
            print(f"[DEBUG] Found {len(files)} files in request")

            # Filter out empty files
            valid_files = [
                f for f in files if f and f.filename and f.filename.strip()
            ]
            print(f"[DEBUG] {len(valid_files)} valid files to process")

            # First pass: save and process all images
            for i, file in enumerate(valid_files):
                if allowed_file(file.filename):
                    try:
                        # Generate unique filename
                        filename = secure_filename(file.filename)
                        filename = f"{uuid.uuid4()}_{filename}"
                        filepath = os.path.join(app.config['UPLOAD_FOLDER'],
                                                filename)

                        # Create directory if it doesn't exist
                        os.makedirs(os.path.dirname(filepath), exist_ok=True)

                        # Save file
                        file.save(filepath)
                        print(f"[DEBUG] File saved: {filepath}")

                        # Compress image
                        compress_image(filepath)

                        # Detect orientation
                        orientation = get_image_orientation(filepath)

                        image_url = f"/static/public/produk_images/{filename}"

                        # Store image info for sorting
                        processed_images.append({
                            'url': image_url,
                            'orientation': orientation,
                            'original_index': i,
                            'filepath': filepath
                        })

                        print(
                            f"[DEBUG] Image processed: {filename}, orientation: {orientation}"
                        )

                    except Exception as img_error:
                        print(
                            f"[ERROR] Failed to process image {file.filename}: {str(img_error)}"
                        )
                        continue

            # Sort images: landscape first, then portrait/square
            def sort_key(img):
                orientation_priority = {
                    'landscape': 0,
                    'square': 1,
                    'portrait': 2,
                    'unknown': 3
                }
                return orientation_priority.get(img['orientation'], 3)

            sorted_images = sorted(processed_images, key=sort_key)
            print(
                f"[DEBUG] Images sorted by orientation: {[img['orientation'] for img in sorted_images]}"
            )

            # Second pass: create ProductImage records with sorted order
            for display_order, img_info in enumerate(sorted_images):
                try:
                    image_url = img_info['url']
                    original_index = img_info['original_index']
                    uploaded_images.append(image_url)

                    # Check if this was the selected thumbnail based on original index
                    is_thumbnail = (original_index == selected_thumbnail_index)

                    product_image = models.ProductImage(
                        product_id=new_product.id,
                        image_url=image_url,
                        is_thumbnail=is_thumbnail,
                        display_order=display_order)
                    db.session.add(product_image)
                    print(
                        f"[DEBUG] ProductImage created: {image_url}, orientation: {img_info['orientation']}, display_order: {display_order}, is_thumbnail: {is_thumbnail}"
                    )

                    # Set the selected thumbnail as the main image_url
                    if is_thumbnail:
                        new_product.image_url = image_url
                        print(f"[DEBUG] Set main image_url: {image_url}")

                except Exception as img_error:
                    print(
                        f"[ERROR] Failed to create ProductImage record: {str(img_error)}"
                    )
                    continue

        # If no thumbnail was selected but images were uploaded, use the first one
        if uploaded_images and not new_product.image_url:
            new_product.image_url = uploaded_images[0]
            # Update the first ProductImage to be thumbnail
            first_image = models.ProductImage.query.filter_by(
                product_id=new_product.id).first()
            if first_image:
                first_image.is_thumbnail = True
            print(
                f"[DEBUG] Using first image as thumbnail: {new_product.image_url}"
            )

        db.session.commit()
        print(
            f"[DEBUG] Product {new_product.name} saved successfully with {len(uploaded_images)} images"
        )
        flash(
            f'Produk {new_product.name} berhasil ditambahkan dengan {len(uploaded_images)} gambar!',
            'success')

    except ValueError as ve:
        db.session.rollback()
        print(f"[ERROR] Validation error: {str(ve)}")
        flash(f'Data tidak valid: {str(ve)}', 'error')
        categories = models.Category.query.filter_by(is_active=True).all()
        suppliers = models.Supplier.query.filter_by(is_active=True).all()
        return render_template('admin/add_product.html',
                               categories=categories,
                               suppliers=suppliers)
    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] Failed to add product: {str(e)}")
        flash(f'Gagal menambahkan produk: {str(e)}', 'error')

    return redirect(url_for('admin_products'))


@app.route('/admin/products/<int:product_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_product(product_id):
    try:
        product = models.Product.query.get_or_404(product_id)
        print(f"[DEBUG] Editing product {product_id}: {product.name}")
    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] Product {product_id} not found: {e}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': 'Product not found'}), 404
        else:
            flash('Produk tidak ditemukan!', 'error')
            return redirect(url_for('admin_products'))

    try:
        # Validate required fields first
        name = request.form.get('name', '').strip()
        price_str = request.form.get('price', '')
        category_id_str = request.form.get('category_id', '')

        if not name:
            raise ValueError('Nama produk harus diisi')
        if not price_str:
            raise ValueError('Harga harus diisi')
        if not category_id_str:
            raise ValueError('Kategori harus dipilih')

        old_name = product.name
        product.name = name
        product.description = request.form.get('description', '')
        product.price = float(price_str)
        product.stock_quantity = int(request.form.get('stock_quantity', 0))
        product.minimum_stock = int(
            request.form.get('minimum_stock', product.minimum_stock or 5))
        product.low_stock_threshold = int(
            request.form.get('low_stock_threshold', product.low_stock_threshold
                             or 10))
        product.brand = request.form.get('brand', '')
        product.model = request.form.get('model', '')
        
        # Handle GTIN - allow manual edit or auto-generate if empty
        gtin = request.form.get('gtin', '').strip()
        if gtin:
            product.gtin = gtin
        elif not product.gtin:
            # Auto-generate only if product doesn't have GTIN yet
            import time
            timestamp = str(int(time.time() * 1000))
            product.gtin = f"hrtbrg{timestamp}"
        
        product.category_id = int(category_id_str)
        product.supplier_id = int(
            request.form['supplier_id']) if request.form.get(
                'supplier_id') else None
        product.is_featured = 'is_featured' in request.form
        product.is_active = 'is_active' in request.form

        # Always regenerate slug to ensure it's current (like add product)
        product.slug = None  # Reset slug to force regeneration
        product.ensure_slug()
        print(
            f"[DEBUG] Edit: Slug regenerated for product {product.name}: {product.slug}"
        )

        # Handle dimensions and weight
        product.weight = float(request.form.get('weight', 0)) if request.form.get('weight') else 0
        product.shipping_weight = float(request.form.get('shipping_weight', 0)) if request.form.get('shipping_weight') else product.weight
        product.length = float(request.form.get('length', 0)) if request.form.get('length') else 0
        product.width = float(request.form.get('width', 0)) if request.form.get('width') else 0
        product.height = float(request.form.get('height', 0)) if request.form.get('height') else 0

        # Handle deleted images
        deleted_images = request.form.getlist('deleted_images[]')
        if deleted_images:
            for img_id in deleted_images:
                try:
                    image = models.ProductImage.query.filter_by(id=int(img_id), product_id=product.id).first()
                    if image:
                        # Hapus file fisik
                        file_path = os.path.join(app.root_path, image.image_url.lstrip('/'))
                        if os.path.exists(file_path):
                            os.remove(file_path)
                            print(f"[DEBUG] Deleted image file: {file_path}")

                        # Hapus dari database
                        db.session.delete(image)
                        print(f"[DEBUG] Deleted ProductImage record ID {img_id}")
                except Exception as del_error:
                    print(f"[ERROR] Failed to delete image ID {img_id}: {str(del_error)}")

        # Handle multiple images upload if provided
        uploaded_images = []
        processed_images = []  # Store image info for sorting
        selected_thumbnail_index = int(
            request.form.get('selected_thumbnail', -1))

        # Check for new_images field (from edit form)
        files_to_process = []
        if 'new_images' in request.files:
            files_to_process = request.files.getlist('new_images')
        elif 'images' in request.files:
            files_to_process = request.files.getlist('images')

        if files_to_process:
            # First pass: save and process all images
            for i, file in enumerate(files_to_process):
                if file and file.filename and allowed_file(file.filename):
                    try:
                        filename = secure_filename(file.filename)
                        filename = f"{uuid.uuid4()}_{filename}"
                        filepath = os.path.join(app.config['UPLOAD_FOLDER'],
                                                filename)

                        # Create directory if it doesn't exist
                        os.makedirs(os.path.dirname(filepath), exist_ok=True)

                        file.save(filepath)
                        compress_image(filepath)

                        # Detect orientation
                        orientation = get_image_orientation(filepath)

                        image_url = f"/static/public/produk_images/{filename}"

                        # Store image info for sorting
                        processed_images.append({
                            'url': image_url,
                            'orientation': orientation,
                            'original_index': i,
                            'filepath': filepath
                        })

                        print(
                            f"[DEBUG] Edit: Image processed: {filename}, orientation: {orientation}"
                        )

                    except Exception as img_error:
                        print(
                            f"[ERROR] Failed to process image {file.filename}: {str(img_error)}"
                        )
                        continue

            # Sort images: landscape first, then portrait/square
            def sort_key(img):
                orientation_priority = {
                    'landscape': 0,
                    'square': 1,
                    'portrait': 2,
                    'unknown': 3
                }
                return orientation_priority.get(img['orientation'], 3)

            sorted_images = sorted(processed_images, key=sort_key)
            print(
                f"[DEBUG] Edit: Images sorted by orientation: {[img['orientation'] for img in sorted_images]}"
            )

            # Get current max display_order
            existing_max_order = max(
                [img.display_order for img in product.images], default=-1)

            # Second pass: create ProductImage records with sorted order
            for display_order_offset, img_info in enumerate(sorted_images):
                try:
                    image_url = img_info['url']
                    original_index = img_info['original_index']
                    uploaded_images.append(image_url)

                    # Check if this was the selected thumbnail based on original index
                    is_thumbnail = (original_index == selected_thumbnail_index)

                    product_image = models.ProductImage(
                        product_id=product.id,
                        image_url=image_url,
                        is_thumbnail=is_thumbnail,
                        display_order=existing_max_order + 1 +
                        display_order_offset  # Add after existing images
                    )
                    db.session.add(product_image)
                    print(
                        f"[DEBUG] Edit: ProductImage created: {image_url}, orientation: {img_info['orientation']}, display_order: {existing_max_order + 1 + display_order_offset}, is_thumbnail: {is_thumbnail}"
                    )

                    # Set the selected thumbnail as the main image_url
                    if is_thumbnail:
                        product.image_url = image_url

                except Exception as img_error:
                    print(
                        f"[ERROR] Failed to create ProductImage record: {str(img_error)}"
                    )
                    continue

        # Handle thumbnail selection from existing images
        current_thumbnail_id = request.form.get('current_thumbnail_selection')
        if current_thumbnail_id:
            # Update existing thumbnails
            for image in product.images:
                image.is_thumbnail = (str(image.id) == current_thumbnail_id)
                if image.is_thumbnail:
                    product.image_url = image.image_url

        # If new images were uploaded but no thumbnail selected, use first new image
        if uploaded_images and selected_thumbnail_index == -1:
            product.image_url = uploaded_images[0]
            # Find and update the first new image to be thumbnail
            newest_images = models.ProductImage.query.filter_by(product_id=product.id)\
                                                     .filter(models.ProductImage.image_url.in_(uploaded_images))\
                                                     .first()
            if newest_images:
                newest_images.is_thumbnail = True

        db.session.commit()
        print(f"[SUCCESS] Product {product.name} updated successfully")

        # Always return success response for AJAX requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': f'Produk {product.name} berhasil diperbarui!',
                'product_id': product.id
            }), 200
        else:
            flash(f'Produk {product.name} berhasil diperbarui!', 'success')
            return redirect(url_for('admin_products'))

    except ValueError as ve:
        db.session.rollback()
        error_msg = str(ve)
        print(f"[ERROR] Validation error: {error_msg}")

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': error_msg}), 400
        else:
            flash(f'Error: {error_msg}', 'error')
            return redirect(url_for('admin_products'))

    except Exception as e:
        db.session.rollback()
        error_msg = f'Gagal memperbarui produk: {str(e)}'
        print(f"[ERROR] Edit product error: {error_msg}")

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': error_msg}), 500
        else:
            flash(error_msg, 'error')
            return redirect(url_for('admin_products'))

@app.route('/admin/products/<int:product_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_product(product_id):
    try:
        product = models.Product.query.get_or_404(product_id)

        # Check if product has been ordered
        if product.order_items:
            flash('Tidak bisa menghapus produk yang sudah pernah dipesan!',
                  'error')
            return redirect(url_for('admin_products'))

        # Check if product is in cart
        if product.cart_items:
            # Remove from all carts
            for cart_item in product.cart_items:
                db.session.delete(cart_item)

        # Delete product images first
        for image in product.images:
            db.session.delete(image)

        product_name = product.name
        db.session.delete(product)
        db.session.commit()

        flash(f'Produk {product_name} berhasil dihapus!', 'success')

    except Exception as e:
        db.session.rollback()
        print(f"Error deleting product: {str(e)}")
        flash(f'Gagal menghapus produk: {str(e)}', 'error')

    return redirect(url_for('admin_products'))


@app.route('/admin/product/<int:product_id>/edit', methods=['GET'])
@login_required
@admin_required
def admin_get_product(product_id):
    """Get product data for editing"""
    product = models.Product.query.get_or_404(product_id)

    return jsonify({
        'id': product.id,
        'name': product.name,
        'description': product.description or '',
        'price': float(product.price),
        'stock_quantity': product.stock_quantity,
        'minimum_stock': product.minimum_stock,
        'low_stock_threshold': product.low_stock_threshold,
        'brand': product.brand or '',
        'model': product.model or '',
        'category_id': product.category_id,
        'supplier_id': product.supplier_id,
        'is_active': product.is_active,
        'is_featured': product.is_featured,
        'image_url': product.image_url or '',
        'weight': float(product.weight or 0),
        'length': float(product.length or 0),
        'width': float(product.width or 0),
        'height': float(product.height or 0)
    })


@app.route('/admin/product/<int:product_id>/images', methods=['GET'])
@login_required
@admin_required
def admin_get_product_images(product_id):
    """Get product images for editing"""
    product = models.Product.query.get_or_404(product_id)

    # Get all product images ordered by display_order
    images = models.ProductImage.query.filter_by(product_id=product_id)\
                                      .order_by(models.ProductImage.display_order)\
                                      .all()

    image_data = []
    for image in images:
        image_data.append({
            'id': image.id,
            'image_url': image.image_url,
            'is_thumbnail': image.is_thumbnail,
            'display_order': image.display_order
        })

    return jsonify(image_data)

@app.route('/admin/product/image/<int:image_id>/delete', methods=['DELETE'])
@login_required
@admin_required
@csrf.exempt
def delete_product_image(image_id):
    try:
        image = models.ProductImage.query.get_or_404(image_id)
        product_id = image.product_id
        
        # Delete physical file
        if image.image_url:
            file_path = os.path.join(app.root_path, image.image_url.lstrip('/'))
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"[DEBUG] Deleted image file: {file_path}")
        
        # Delete from database
        db.session.delete(image)
        db.session.commit()
        
        print(f"[SUCCESS] Image {image_id} deleted successfully")
        return jsonify({'success': True, 'message': 'Gambar berhasil dihapus'})
    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] Failed to delete image: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/admin/product/<int:product_id>/thumbnail/<int:image_id>', methods=['PUT'])
@login_required
@admin_required
@csrf.exempt
def set_product_thumbnail(product_id, image_id):
    try:
        product = models.Product.query.get_or_404(product_id)
        image = models.ProductImage.query.get_or_404(image_id)
        
        if image.product_id != product_id:
            return jsonify({'success': False, 'message': 'Image tidak ditemukan pada produk ini'}), 404
        
        # Reset all thumbnails for this product
        for img in product.images:
            img.is_thumbnail = False
        
        # Set new thumbnail
        image.is_thumbnail = True
        product.image_url = image.image_url
        
        db.session.commit()
        
        print(f"[SUCCESS] Thumbnail set for product {product_id}, image {image_id}")
        return jsonify({'success': True, 'message': 'Gambar utama berhasil diubah'})
    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] Failed to set thumbnail: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/admin/product/<int:product_id>/barcode')
@login_required
@staff_required
def generate_product_barcode(product_id):
    """Generate barcode untuk produk (hanya admin/staff)"""
    try:
        import barcode
        from barcode.writer import ImageWriter
        
        product = models.Product.query.get_or_404(product_id)
        
        if not product.gtin:
            return jsonify({'error': 'Produk tidak memiliki GTIN'}), 400
        
        # Generate barcode menggunakan Code128
        code128 = barcode.get_barcode_class('code128')
        barcode_image = code128(product.gtin, writer=ImageWriter())
        
        # Save to BytesIO
        buffer = io.BytesIO()
        barcode_image.write(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='image/png',
            as_attachment=True,
            download_name=f'barcode_{product.gtin}.png'
        )
        
    except Exception as e:
        print(f"[ERROR] Failed to generate barcode: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/admin/product/<int:product_id>/qrcode')
@login_required
@staff_required
def generate_product_qrcode(product_id):
    """Generate QR code untuk produk (hanya admin/staff)"""
    try:
        import qrcode
        from qrcode.image.pure import PyPNGImage
        
        product = models.Product.query.get_or_404(product_id)
        
        if not product.gtin:
            return jsonify({'error': 'Produk tidak memiliki GTIN'}), 400
        
        # Create QR data dengan info produk
        qr_data = f"GTIN: {product.gtin}\nNama: {product.name}\nHarga: Rp {product.price:,.0f}"
        
        # Generate QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Save to BytesIO
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='image/png',
            as_attachment=True,
            download_name=f'qrcode_{product.gtin}.png'
        )
        
    except Exception as e:
        print(f"[ERROR] Failed to generate QR code: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/admin/categories/<int:category_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_category(category_id):
    category = models.Category.query.get_or_404(category_id)

    try:
        category.name = request.form['name']
        category.description = request.form.get('description', '')
        category.is_active = 'is_active' in request.form

        db.session.commit()
        flash(f'Kategori {category.name} berhasil diperbarui!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_categories'))


@app.route('/admin/categories/<int:category_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_category(category_id):
    category = models.Category.query.get_or_404(category_id)

    # Check if category has products
    if category.products:
        flash('Tidak bisa menghapus kategori yang memiliki produk!', 'error')
        return redirect(url_for('admin_categories'))

    category_name = category.name
    db.session.delete(category)
    db.session.commit()

    flash(f'Kategori {category_name} berhasil dihapus!', 'success')
    return redirect(url_for('admin_categories'))


@app.route('/admin/categories')
@login_required
@admin_required
def admin_categories():
    categories = models.Category.query.all()
    return render_template('admin/categories.html', categories=categories)


@app.route('/admin/categories/add', methods=['POST'])
@login_required
@admin_required
def admin_add_category():
    name = request.form['name']
    description = request.form['description']

    category = models.Category(name=name, description=description)
    db.session.add(category)
    db.session.commit()

    flash('Kategori berhasil ditambahkan!', 'success')
    return redirect(url_for('admin_categories'))


# Payment Configuration Routes
@app.route('/admin/payment-config')
@login_required
@admin_required
def admin_payment_config():
    configurations = models.PaymentConfiguration.query.all()
    return render_template('admin/payment_config.html',
                           configurations=configurations)


@app.route('/admin/payment-config/create', methods=['GET', 'POST'])
@login_required
@admin_required
@csrf.exempt
def admin_create_payment_config():
    if request.method == 'POST':
        try:
            provider = request.form.get('provider')
            is_sandbox = request.form.get('is_sandbox') == 'true'

            # Validasi provider
            if not provider or provider not in [
                    'midtrans', 'stripe', 'xendit', 'doku'
            ]:
                flash('Provider pembayaran tidak valid!', 'error')
                return render_template('admin/create_payment_config.html')

            # Cek apakah sudah ada konfigurasi aktif untuk provider ini
            existing_config = models.PaymentConfiguration.query.filter_by(
                provider=provider, is_active=True).first()

            if existing_config:
                flash(
                    f'Sudah ada konfigurasi {provider} yang aktif! Nonaktifkan terlebih dahulu sebelum menambah yang baru.',
                    'warning')
                return render_template('admin/create_payment_config.html')

            config = models.PaymentConfiguration(
                provider=provider,
                is_sandbox=is_sandbox,
                is_active=False  # Start as inactive
            )

            if provider == 'midtrans':
                client_key = request.form.get('midtrans_client_key',
                                              '').strip()
                server_key = request.form.get('midtrans_server_key',
                                              '').strip()
                merchant_id = request.form.get('midtrans_merchant_id',
                                               '').strip()

                if not client_key or not server_key:
                    flash(
                        'Client Key dan Server Key harus diisi untuk Midtrans!',
                        'error')
                    return render_template('admin/create_payment_config.html')

                config.midtrans_client_key = client_key
                config.midtrans_server_key = server_key
                config.midtrans_merchant_id = merchant_id if merchant_id else None

            elif provider == 'stripe':
                publishable_key = request.form.get('stripe_publishable_key',
                                                   '').strip()
                secret_key = request.form.get('stripe_secret_key', '').strip()

                if not publishable_key or not secret_key:
                    flash(
                        'Publishable Key dan Secret Key harus diisi untuk Stripe!',
                        'error')
                    return render_template('admin/create_payment_config.html')

                config.stripe_publishable_key = publishable_key
                config.stripe_secret_key = secret_key

            elif provider == 'xendit':
                api_key = request.form.get('xendit_api_key', '').strip()
                webhook_token = request.form.get('xendit_webhook_token',
                                                 '').strip()
                public_key = request.form.get('xendit_public_key', '').strip()

                if not api_key:
                    flash('API Key harus diisi untuk Xendit!', 'error')
                    return render_template('admin/create_payment_config.html')

                config.xendit_api_key = api_key
                config.xendit_webhook_token = webhook_token if webhook_token else None
                config.xendit_public_key = public_key if public_key else None

            elif provider == 'doku':
                client_id = request.form.get('doku_client_id', '').strip()
                secret_key = request.form.get('doku_secret_key', '').strip()
                private_key = request.form.get('doku_private_key', '').strip()
                public_key = request.form.get('doku_public_key', '').strip()

                if not client_id or not secret_key:
                    flash('Client ID dan Secret Key harus diisi untuk DOKU!',
                          'error')
                    return render_template('admin/create_payment_config.html')

                config.doku_client_id = client_id
                config.doku_secret_key = secret_key
                config.doku_private_key = private_key if private_key else None
                config.doku_public_key = public_key if public_key else None

            # Set callback URLs
            base_url = request.host_url.rstrip('/')
            config.callback_finish_url = f"{base_url}/payment/finish"
            config.callback_unfinish_url = f"{base_url}/payment/unfinish"
            config.callback_error_url = f"{base_url}/payment/error"
            config.notification_url = f"{base_url}/notification/handling"

            # Set additional URLs for Midtrans
            if provider == 'midtrans':
                config.recurring_notification_url = f"{base_url}/notification/recurring"
                config.account_linking_url = f"{base_url}/notification/account-linking"

            db.session.add(config)
            db.session.commit()

            environment_text = "Sandbox" if is_sandbox else "Production"
            flash(
                f'Konfigurasi pembayaran {provider.title()} ({environment_text}) berhasil ditambahkan!',
                'success')
            return redirect(url_for('admin_payment_config'))

        except Exception as e:
            db.session.rollback()
            print(f"Error creating payment config: {str(e)}")
            flash(f'Terjadi kesalahan saat menyimpan konfigurasi: {str(e)}',
                  'error')
            return render_template('admin/create_payment_config.html')

    return render_template('admin/create_payment_config.html')


@app.route('/admin/payment-config/<int:config_id>/toggle', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def admin_toggle_payment_config(config_id):
    config = models.PaymentConfiguration.query.get_or_404(config_id)

    # Deactivate all other configs of the same provider
    if not config.is_active:
        models.PaymentConfiguration.query.filter_by(
            provider=config.provider).update({'is_active': False})

    config.is_active = not config.is_active
    config.updated_at = datetime.utcnow()

    db.session.commit()

    status = 'diaktifkan' if config.is_active else 'dinonaktifkan'
    flash(f'Konfigurasi {config.provider} berhasil {status}!', 'success')
    return redirect(url_for('admin_payment_config'))


# Midtrans Payment Callback Endpoints
@app.route('/payment/finish')
def payment_finish():
    order_id = request.args.get('order_id')
    status_code = request.args.get('status_code')
    transaction_status = request.args.get('transaction_status')

    if status_code == '200' and transaction_status == 'settlement':
        flash('Pembayaran berhasil! Terima kasih atas pesanan Anda.',
              'success')
        return redirect(url_for('payment_success'))
    else:
        flash('Pembayaran gagal atau dibatalkan.', 'error')
        return redirect(url_for('cart'))


@app.route('/payment/notification', methods=['POST'])
@csrf.exempt
def payment_notification():
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                'status': 'error',
                'message': 'No data received'
            }), 400

        # Get active Midtrans configuration
        midtrans_config = models.PaymentConfiguration.query.filter_by(
            provider='midtrans', is_active=True).first()

        if not midtrans_config:
            return jsonify({
                'status': 'error',
                'message': 'Midtrans not configured'
            }), 400

        # Verify signature here (implementation depends on your requirements)
        order_id = data.get('order_id')
        transaction_status = data.get('transaction_status')
        fraud_status = data.get('fraud_status')

        if order_id:
            # Find the transaction
            midtrans_transaction = models.MidtransTransaction.query.filter_by(
                transaction_id=order_id).first()

            if midtrans_transaction:
                midtrans_transaction.transaction_status = transaction_status
                midtrans_transaction.fraud_status = fraud_status
                midtrans_transaction.midtrans_response = json.dumps(data)
                midtrans_transaction.updated_at = datetime.utcnow()

                # Update order status based on transaction status
                order = midtrans_transaction.order
                previous_status = order.status

                if transaction_status == 'settlement' and fraud_status == 'accept':
                    order.status = 'paid'

                    # Reduce stock only if order was not previously paid (avoid double reduction)
                    if previous_status not in ['paid', 'shipped', 'delivered']:
                        print(
                            f"[MIDTRANS] Processing stock reduction for order {order.id}"
                        )
                        for order_item in order.order_items:
                            product = order_item.product
                            if product.stock_quantity >= order_item.quantity:
                                product.stock_quantity -= order_item.quantity
                                print(
                                    f"[MIDTRANS] Reduced stock for {product.name}: {product.stock_quantity + order_item.quantity} -> {product.stock_quantity}"
                                )
                            else:
                                print(
                                    f"[MIDTRANS] WARNING: Insufficient stock for {product.name}. Available: {product.stock_quantity}, Required: {order_item.quantity}"
                                )

                elif transaction_status in ['deny', 'cancel', 'expire']:
                    order.status = 'cancelled'

                db.session.commit()

        return jsonify({'status': 'ok'})

    except Exception as e:
        print(f"Payment notification error: {str(e)}")
        return jsonify({'status': 'error', 'message': 'Internal error'}), 500


@app.route('/notification/handling', methods=['POST'])
@csrf.exempt
def notification_handling():
    """
    URL Callback utama untuk notifikasi pembayaran Midtrans
    Menangani semua jenis notifikasi dari Midtrans
    """
    try:
        # Try to get JSON data, fallback to form data
        data = request.get_json()
        if not data:
            data = request.form.to_dict()

        if not data:
            print("No data received in notification")
            return jsonify({
                'status': 'error',
                'message': 'No data received'
            }), 400

        # Log notifikasi untuk debugging
        print(f"Midtrans notification received: {json.dumps(data)}")

        # Get active Midtrans configuration
        midtrans_config = models.PaymentConfiguration.query.filter_by(
            provider='midtrans', is_active=True).first()

        if not midtrans_config:
            print("No active Midtrans configuration found")
            # Return OK to prevent Midtrans from retrying
            return jsonify({
                'status': 'ok',
                'message': 'No active Midtrans config'
            }), 200

        # Extract data from notification
        order_id = data.get('order_id', '')
        transaction_status = data.get('transaction_status', '')
        fraud_status = data.get('fraud_status', 'accept')
        payment_type = data.get('payment_type', '')
        gross_amount = data.get('gross_amount', '0')
        settlement_time = data.get('settlement_time')

        print(f"Processing order_id: {order_id}, status: {transaction_status}")

        if not order_id:
            print("No order_id in notification")
            return jsonify({
                'status': 'ok',
                'message': 'No order_id provided'
            }), 200

        # Find or create the transaction record
        midtrans_transaction = models.MidtransTransaction.query.filter_by(
            transaction_id=order_id).first()

        if not midtrans_transaction:
            print(f"Creating new transaction record for {order_id}")
            # Try to find order by various patterns
            order = None

            # Pattern 1: ORDER-{user_id}-{timestamp}-{uuid}
            try:
                order_parts = order_id.split('-')
                if len(order_parts) >= 3 and order_parts[0] == 'ORDER':
                    user_id = int(order_parts[1])
                    # Find recent unpaid order for this user
                    order = models.Order.query.filter_by(
                        user_id=user_id, status='pending').order_by(
                            models.Order.created_at.desc()).first()
            except (ValueError, IndexError):
                pass

            # Pattern 2: Direct order ID
            if not order and order_id.isdigit():
                try:
                    order = models.Order.query.get(int(order_id))
                except ValueError:
                    pass

            # Pattern 3: Find by session data or recent orders
            if not order:
                # Find the most recent pending order
                order = models.Order.query.filter_by(
                    status='pending').order_by(
                        models.Order.created_at.desc()).first()

            if order:
                midtrans_transaction = models.MidtransTransaction(
                    order_id=order.id,
                    transaction_id=order_id,
                    gross_amount=float(gross_amount) if gross_amount else 0,
                    payment_type=payment_type,
                    transaction_status=transaction_status,
                    fraud_status=fraud_status,
                    midtrans_response=json.dumps(data))
                db.session.add(midtrans_transaction)
                db.session.flush()
                print(f"Created transaction record for order {order.id}")
            else:
                print(f"No matching order found for transaction {order_id}")
                return jsonify({
                    'status': 'ok',
                    'message': 'No matching order found'
                }), 200

        if midtrans_transaction:
            # Update transaction details
            old_status = midtrans_transaction.transaction_status
            midtrans_transaction.transaction_status = transaction_status
            midtrans_transaction.fraud_status = fraud_status
            midtrans_transaction.payment_type = payment_type
            midtrans_transaction.midtrans_response = json.dumps(data)
            midtrans_transaction.updated_at = datetime.utcnow()

            if settlement_time:
                try:
                    midtrans_transaction.settlement_time = datetime.strptime(
                        settlement_time, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    print(f"Invalid settlement_time format: {settlement_time}")

            # Update order status based on transaction status
            order = midtrans_transaction.order
            old_order_status = order.status

            if transaction_status == 'settlement' and fraud_status == 'accept':
                order.status = 'paid'
                print(f"Order {order.id} marked as paid")
            elif transaction_status in ['deny', 'cancel', 'expire', 'failure']:
                order.status = 'cancelled'
                print(f"Order {order.id} marked as cancelled")
            elif transaction_status == 'pending':
                order.status = 'pending'
                print(f"Order {order.id} kept as pending")

            db.session.commit()

            print(
                f"Transaction {order_id} updated: {old_status} -> {transaction_status}, Order {order.id}: {old_order_status} -> {order.status}"
            )

        return jsonify({
            'status': 'ok',
            'message': 'Notification processed successfully'
        }), 200

    except Exception as e:
        print(f"Notification handling error: {str(e)}")
        import traceback
        traceback.print_exc()
        # Return OK to prevent Midtrans from retrying indefinitely
        return jsonify({
            'status': 'ok',
            'message': 'Error processed, please check logs'
        }), 200


@app.route('/notification/recurring', methods=['POST'])
@csrf.exempt
def notification_recurring():
    """
    URL untuk notifikasi pembayaran berulang (subscription)
    """
    try:
        # Try to get JSON data, fallback to form data
        data = request.get_json()
        if not data:
            data = request.form.to_dict()

        if not data:
            print("No data received in recurring notification")
            return jsonify({
                'status': 'error',
                'message': 'No data received'
            }), 400

        print(f"Recurring payment notification: {json.dumps(data)}")

        # Get active Midtrans configuration
        midtrans_config = models.PaymentConfiguration.query.filter_by(
            provider='midtrans', is_active=True).first()

        if not midtrans_config:
            print("No active Midtrans configuration found for recurring")
            return jsonify({
                'status': 'ok',
                'message': 'No active Midtrans config'
            }), 200

        # Extract recurring payment data
        subscription_id = data.get('subscription_id', '')
        transaction_id = data.get('transaction_id', '')
        transaction_status = data.get('transaction_status', '')
        payment_type = data.get('payment_type', '')
        gross_amount = data.get('gross_amount', '0')
        order_id = data.get('order_id', '')

        print(
            f"Processing recurring payment - subscription_id: {subscription_id}, transaction_id: {transaction_id}, status: {transaction_status}"
        )

        # Log recurring payment attempt
        recurring_data = {
            'subscription_id': subscription_id,
            'transaction_id': transaction_id,
            'transaction_status': transaction_status,
            'payment_type': payment_type,
            'gross_amount': gross_amount,
            'order_id': order_id,
            'notification_data': data,
            'created_at': datetime.utcnow().isoformat()
        }

        # For now, we'll just log the recurring payment notification
        # In the future, this can be extended to handle subscription logic
        print(f"Recurring payment logged: {json.dumps(recurring_data)}")

        # If this is related to an existing order, try to update it
        if order_id:
            try:
                # Find transaction by order_id
                midtrans_transaction = models.MidtransTransaction.query.filter_by(
                    transaction_id=order_id).first()

                if midtrans_transaction:
                    # Update with recurring payment info
                    response_data = json.loads(
                        midtrans_transaction.midtrans_response or '{}')
                    response_data['recurring_info'] = data
                    midtrans_transaction.midtrans_response = json.dumps(
                        response_data)
                    midtrans_transaction.updated_at = datetime.utcnow()

                    db.session.commit()
                    print(
                        f"Updated transaction {order_id} with recurring payment info"
                    )
                else:
                    print(
                        f"No transaction found for recurring payment order_id: {order_id}"
                    )

            except Exception as e:
                print(
                    f"Error updating transaction with recurring info: {str(e)}"
                )

        # TODO: Implement subscription management logic here
        # This could include:
        # 1. Creating subscription records
        # 2. Managing recurring billing cycles
        # 3. Handling subscription renewals
        # 4. Sending renewal notifications to customers

        return jsonify({
            'status': 'ok',
            'message': 'Recurring payment notification processed',
            'subscription_id': subscription_id,
            'transaction_status': transaction_status
        }), 200

    except Exception as e:
        print(f"Recurring notification error: {str(e)}")
        import traceback
        traceback.print_exc()
        # Return OK to prevent Midtrans from retrying indefinitely
        return jsonify({
            'status': 'ok',
            'message': 'Error processed, please check logs'
        }), 200


@app.route('/notification/account-linking', methods=['POST'])
@csrf.exempt
def notification_account_linking():
    """
    URL untuk notifikasi menghubungkan akun (account linking)
    """
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                'status': 'error',
                'message': 'No data received'
            }), 400

        print(f"Account linking notification: {json.dumps(data)}")

        # Handle account linking logic here
        # This can be extended based on your account linking needs

        return jsonify({'status': 'ok'})

    except Exception as e:
        print(f"Account linking notification error: {str(e)}")
        return jsonify({'status': 'error', 'message': 'Internal error'}), 500


# ===========================
# XENDIT & DOKU PAYMENT ROUTES
# ===========================


# Xendit Payment Processing Routes
@app.route('/payment/xendit/create', methods=['POST'])
@login_required
def create_xendit_payment():
    """Create Xendit payment (eWallet, QR Code, Virtual Account)"""
    if not XENDIT_AVAILABLE:
        flash('Xendit payment tidak tersedia!', 'error')
        return redirect(url_for('cart'))

    try:
        data = request.get_json() or request.form.to_dict()
        payment_method = data.get('payment_method')  # 'ewallet', 'qr', 'va'
        channel_code = data.get(
            'channel_code')  # 'OVO', 'DANA', 'LINKAJA', etc.

        # Get cart items and calculate total
        cart_items = models.CartItem.query.filter_by(
            user_id=current_user.id).all()
        if not cart_items:
            return jsonify({
                'success': False,
                'error': 'Keranjang kosong'
            }), 400

        subtotal = sum(item.quantity * item.product.price
                       for item in cart_items)
        shipping_cost = session.get('shipping_cost', 0)
        total_amount = float(subtotal) + float(shipping_cost)

        # Get Xendit configuration
        xendit_config = models.PaymentConfiguration.query.filter_by(
            provider='xendit', is_active=True).first()

        if not xendit_config or not xendit_config.xendit_api_key:
            return jsonify({
                'success': False,
                'error': 'Xendit belum dikonfigurasi'
            }), 400

        # Set Xendit API key
        xendit.api_key = xendit_config.xendit_api_key

        # Generate unique external ID
        external_id = f"ORDER-{current_user.id}-{int(datetime.utcnow().timestamp())}"

        # Create order first
        order = models.Order(
            user_id=current_user.id,
            total_amount=total_amount,
            status='pending',
            shipping_service_id=session.get('shipping_service_id'),
            shipping_cost=shipping_cost)
        db.session.add(order)
        db.session.flush()  # Get order ID

        # Create payment based on method
        if payment_method == 'ewallet':
            payment_response = create_xendit_ewallet_payment(
                external_id, total_amount, channel_code, current_user,
                order.id)
        elif payment_method == 'qr':
            payment_response = create_xendit_qr_payment(
                external_id, total_amount, current_user, order.id)
        elif payment_method == 'va':
            payment_response = create_xendit_va_payment(
                external_id, total_amount, channel_code, current_user,
                order.id)
        else:
            return jsonify({
                'success': False,
                'error': 'Metode pembayaran tidak valid'
            }), 400

        if payment_response['success']:
            # Add order items
            for cart_item in cart_items:
                order_item = models.OrderItem(order_id=order.id,
                                              product_id=cart_item.product_id,
                                              quantity=cart_item.quantity,
                                              price=cart_item.product.price)
                db.session.add(order_item)

            db.session.commit()
            return jsonify(payment_response)
        else:
            db.session.rollback()
            return jsonify(payment_response), 400

    except Exception as e:
        db.session.rollback()
        print(f"Xendit payment error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


def create_xendit_ewallet_payment(external_id, amount, channel_code, user,
                                  order_id):
    """Create eWallet payment with Xendit"""
    try:
        ewallet_charge = EWallet.create_ewallet_charge(
            reference_id=external_id,
            currency="IDR",
            amount=int(amount),
            checkout_method="ONE_TIME_PAYMENT",
            channel_code=channel_code,
            channel_properties={
                "mobile_number": getattr(user, 'phone', "+628123456789")
            },
            customer={
                "reference_id": f"customer-{user.id}",
                "given_names": user.username,
                "email": user.email,
                "mobile_number": getattr(user, 'phone', "+628123456789")
            },
            metadata={
                "order_id": str(order_id),
                "user_id": str(user.id)
            })

        # Store transaction
        transaction = models.XenditTransaction(
            order_id=order_id,
            transaction_id=ewallet_charge['id'],
            external_id=external_id,
            amount=amount,
            payment_method='EWALLET',
            channel_code=channel_code,
            status=ewallet_charge['status'],
            checkout_url=ewallet_charge.get(
                'actions', {}).get('desktop_web_checkout_url'),
            mobile_deeplink=ewallet_charge.get(
                'actions', {}).get('mobile_deeplink_checkout_url'),
            xendit_response=json.dumps(ewallet_charge))
        db.session.add(transaction)

        return {
            'success':
            True,
            'payment_id':
            ewallet_charge['id'],
            'checkout_url':
            ewallet_charge.get('actions', {}).get('desktop_web_checkout_url'),
            'mobile_deeplink':
            ewallet_charge.get('actions',
                               {}).get('mobile_deeplink_checkout_url'),
            'status':
            ewallet_charge['status']
        }

    except Exception as e:
        return {'success': False, 'error': str(e)}


def create_xendit_qr_payment(external_id, amount, user, order_id):
    """Create QR Code payment with Xendit"""
    try:
        qr_code = QRCode.create_qr_code(
            external_id=external_id,
            type="DYNAMIC",
            callback_url=f"{request.host_url}webhook/xendit",
            amount=int(amount),
            currency="IDR")

        # Store transaction
        transaction = models.XenditTransaction(
            order_id=order_id,
            transaction_id=qr_code['id'],
            external_id=external_id,
            amount=amount,
            payment_method='QR_CODE',
            status=qr_code['status'],
            qr_code=qr_code['qr_string'],
            xendit_response=json.dumps(qr_code))
        db.session.add(transaction)

        return {
            'success': True,
            'payment_id': qr_code['id'],
            'qr_code': qr_code['qr_string'],
            'status': qr_code['status']
        }

    except Exception as e:
        return {'success': False, 'error': str(e)}


def create_xendit_va_payment(external_id, amount, bank_code, user, order_id):
    """Create Virtual Account payment with Xendit"""
    try:
        va = VirtualAccount.create_fixed_virtual_account(
            external_id=external_id,
            bank_code=bank_code,
            name=user.username,
            expected_amount=int(amount),
            currency="IDR",
            is_closed=True,
            expiration_date=(datetime.utcnow() +
                             timedelta(hours=24)).isoformat())

        # Store transaction
        transaction = models.XenditTransaction(
            order_id=order_id,
            transaction_id=va['id'],
            external_id=external_id,
            amount=amount,
            payment_method='VIRTUAL_ACCOUNT',
            channel_code=bank_code,
            status=va['status'],
            va_number=va['account_number'],
            bank_code=bank_code,
            expired_at=datetime.fromisoformat(va['expiration_date'].replace(
                'Z', '+00:00')),
            xendit_response=json.dumps(va))
        db.session.add(transaction)

        return {
            'success': True,
            'payment_id': va['id'],
            'va_number': va['account_number'],
            'bank_code': bank_code,
            'amount': amount,
            'expired_at': va['expiration_date'],
            'status': va['status']
        }

    except Exception as e:
        return {'success': False, 'error': str(e)}


# DOKU Payment Processing Routes
@app.route('/payment/doku/create', methods=['POST'])
@login_required
def create_doku_payment():
    """Create DOKU payment (Virtual Account, Direct Debit)"""
    if not DOKU_AVAILABLE:
        flash('DOKU payment tidak tersedia!', 'error')
        return redirect(url_for('cart'))

    try:
        data = request.get_json() or request.form.to_dict()
        payment_method = data.get('payment_method')  # 'va', 'direct_debit'
        channel_code = data.get('channel_code')  # 'VIRTUAL_ACCOUNT_BNI', etc.

        # Get cart items and calculate total
        cart_items = models.CartItem.query.filter_by(
            user_id=current_user.id).all()
        if not cart_items:
            return jsonify({
                'success': False,
                'error': 'Keranjang kosong'
            }), 400

        subtotal = sum(item.quantity * item.product.price
                       for item in cart_items)
        shipping_cost = session.get('shipping_cost', 0)
        total_amount = float(subtotal) + float(shipping_cost)

        # Get DOKU configuration
        doku_config = models.PaymentConfiguration.query.filter_by(
            provider='doku', is_active=True).first()

        if not doku_config or not doku_config.doku_client_id:
            return jsonify({
                'success': False,
                'error': 'DOKU belum dikonfigurasi'
            }), 400

        # Initialize DOKU SNAP
        snap = DokuSNAP(private_key=doku_config.doku_private_key,
                        client_id=doku_config.doku_client_id,
                        is_production=not doku_config.is_sandbox,
                        public_key=doku_config.doku_public_key,
                        secret_key=doku_config.doku_secret_key,
                        issuer="Hurtrock Music Store")

        # Generate unique invoice number
        invoice_number = f"INV-{current_user.id}-{int(datetime.utcnow().timestamp())}"

        # Create order first
        order = models.Order(
            user_id=current_user.id,
            total_amount=total_amount,
            status='pending',
            shipping_service_id=session.get('shipping_service_id'),
            shipping_cost=shipping_cost)
        db.session.add(order)
        db.session.flush()  # Get order ID

        if payment_method == 'va':
            payment_response = create_doku_va_payment(snap, invoice_number,
                                                      total_amount,
                                                      channel_code,
                                                      current_user, order.id)
        else:
            return jsonify({
                'success': False,
                'error': 'Metode pembayaran tidak valid'
            }), 400

        if payment_response['success']:
            # Add order items
            for cart_item in cart_items:
                order_item = models.OrderItem(order_id=order.id,
                                              product_id=cart_item.product_id,
                                              quantity=cart_item.quantity,
                                              price=cart_item.product.price)
                db.session.add(order_item)

            db.session.commit()
            return jsonify(payment_response)
        else:
            db.session.rollback()
            return jsonify(payment_response), 400

    except Exception as e:
        db.session.rollback()
        print(f"DOKU payment error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


def create_doku_va_payment(snap, invoice_number, amount, channel_code, user,
                           order_id):
    """Create Virtual Account payment with DOKU"""
    try:
        va_request = VirtualAccountRequest(
            order={
                "invoice_number": invoice_number,
                "amount": int(amount)
            },
            virtual_account_info={
                "billing_type": "FIX_BILL",
                "expired_time": 60,  # 60 minutes
                "reusable_status": False,
                "info1": "Hurtrock Music Store",
                "info2": "Thank you for shopping",
                "info3": "with us"
            },
            customer={
                "name": user.username,
                "email": user.email
            })

        response = snap.create_virtual_account(va_request)

        # Store transaction
        transaction = models.DokuTransaction(
            order_id=order_id,
            transaction_id=response.get('virtual_account_number',
                                        invoice_number),
            invoice_number=invoice_number,
            amount=amount,
            payment_method='VIRTUAL_ACCOUNT',
            channel_code=channel_code,
            status='PENDING',
            va_number=response.get('virtual_account_number'),
            expired_at=datetime.utcnow() + timedelta(hours=1),
            doku_response=json.dumps(response))
        db.session.add(transaction)

        return {
            'success': True,
            'invoice_number': invoice_number,
            'va_number': response.get('virtual_account_number'),
            'amount': amount,
            'expired_at': (datetime.utcnow() + timedelta(hours=1)).isoformat(),
            'status': 'PENDING'
        }

    except Exception as e:
        return {'success': False, 'error': str(e)}


# Webhook Handlers
@app.route('/webhook/xendit', methods=['POST'])
@csrf.exempt
def xendit_webhook():
    """Handle Xendit webhook notifications"""
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                'status': 'error',
                'message': 'No data received'
            }), 400

        print(f"Xendit webhook received: {json.dumps(data)}")

        # Get webhook token for verification (optional but recommended)
        webhook_token = request.headers.get('x-callback-token')

        # Process different event types
        event_type = data.get('event_type')

        if event_type == 'ewallet.charge.succeeded':
            handle_xendit_ewallet_success(data['data'])
        elif event_type == 'virtual_account.paid':
            handle_xendit_va_success(data)
        elif event_type == 'qr_code.paid':
            handle_xendit_qr_success(data)
        elif event_type in [
                'ewallet.charge.failed', 'virtual_account.expired',
                'qr_code.expired'
        ]:
            handle_xendit_payment_failed(data)

        return jsonify({'status': 'received'}), 200

    except Exception as e:
        print(f"Xendit webhook error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


def handle_xendit_ewallet_success(data):
    """Handle successful Xendit eWallet payment"""
    try:
        payment_id = data['id']
        transaction = models.XenditTransaction.query.filter_by(
            transaction_id=payment_id).first()

        if transaction:
            transaction.status = 'SUCCEEDED'
            transaction.paid_at = datetime.utcnow()
            transaction.order.status = 'paid'

            # Reduce stock quantities
            for order_item in transaction.order.order_items:
                product = order_item.product
                if product.stock_quantity >= order_item.quantity:
                    product.stock_quantity -= order_item.quantity
                    print(
                        f"[XENDIT] Reduced stock for {product.name}: {product.stock_quantity + order_item.quantity} -> {product.stock_quantity}"
                    )
                else:
                    print(
                        f"[XENDIT] WARNING: Insufficient stock for {product.name}"
                    )

            db.session.commit()
            print(
                f"[XENDIT] eWallet payment {payment_id} processed successfully"
            )

    except Exception as e:
        print(f"Error processing Xendit eWallet success: {str(e)}")
        db.session.rollback()


def handle_xendit_va_success(data):
    """Handle successful Xendit Virtual Account payment"""
    try:
        external_id = data.get('external_id')
        transaction = models.XenditTransaction.query.filter_by(
            external_id=external_id).first()

        if transaction:
            transaction.status = 'SUCCEEDED'
            transaction.paid_at = datetime.utcnow()
            transaction.order.status = 'paid'

            # Reduce stock quantities
            for order_item in transaction.order.order_items:
                product = order_item.product
                if product.stock_quantity >= order_item.quantity:
                    product.stock_quantity -= order_item.quantity
                    print(
                        f"[XENDIT] Reduced stock for {product.name}: {product.stock_quantity + order_item.quantity} -> {product.stock_quantity}"
                    )
                else:
                    print(
                        f"[XENDIT] WARNING: Insufficient stock for {product.name}"
                    )

            db.session.commit()
            print(f"[XENDIT] VA payment {external_id} processed successfully")

    except Exception as e:
        print(f"Error processing Xendit VA success: {str(e)}")
        db.session.rollback()


def handle_xendit_qr_success(data):
    """Handle successful Xendit QR Code payment"""
    try:
        external_id = data.get('external_id')
        transaction = models.XenditTransaction.query.filter_by(
            external_id=external_id).first()

        if transaction:
            transaction.status = 'SUCCEEDED'
            transaction.paid_at = datetime.utcnow()
            transaction.order.status = 'paid'

            # Reduce stock quantities
            for order_item in transaction.order.order_items:
                product = order_item.product
                if product.stock_quantity >= order_item.quantity:
                    product.stock_quantity -= order_item.quantity
                    print(
                        f"[XENDIT] Reduced stock for {product.name}: {product.stock_quantity + order_item.quantity} -> {product.stock_quantity}"
                    )
                else:
                    print(
                        f"[XENDIT] WARNING: Insufficient stock for {product.name}"
                    )

            db.session.commit()
            print(f"[XENDIT] QR payment {external_id} processed successfully")

    except Exception as e:
        print(f"Error processing Xendit QR success: {str(e)}")
        db.session.rollback()


def handle_xendit_payment_failed(data):
    """Handle failed Xendit payment"""
    try:
        payment_id = data.get('data', {}).get('id') or data.get('id')
        external_id = data.get('external_id')

        transaction = None
        if payment_id:
            transaction = models.XenditTransaction.query.filter_by(
                transaction_id=payment_id).first()
        elif external_id:
            transaction = models.XenditTransaction.query.filter_by(
                external_id=external_id).first()

        if transaction:
            transaction.status = 'FAILED'
            transaction.order.status = 'cancelled'
            db.session.commit()
            print(
                f"[XENDIT] Payment {payment_id or external_id} marked as failed"
            )

    except Exception as e:
        print(f"Error processing Xendit payment failure: {str(e)}")
        db.session.rollback()


@app.route('/webhook/doku', methods=['POST'])
@csrf.exempt
def doku_webhook():
    """Handle DOKU webhook notifications"""
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                'status': 'error',
                'message': 'No data received'
            }), 400

        print(f"DOKU webhook received: {json.dumps(data)}")

        # Process payment notification
        transaction_status = data.get('transaction', {}).get('status')
        invoice_number = data.get('order', {}).get('invoice_number')

        if transaction_status == 'SUCCESS' and invoice_number:
            handle_doku_payment_success(invoice_number)
        elif transaction_status in ['FAILED', 'EXPIRED']:
            handle_doku_payment_failed(invoice_number)

        return jsonify({'status': 'received'}), 200

    except Exception as e:
        print(f"DOKU webhook error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


def handle_doku_payment_success(invoice_number):
    """Handle successful DOKU payment"""
    try:
        transaction = models.DokuTransaction.query.filter_by(
            invoice_number=invoice_number).first()

        if transaction:
            transaction.status = 'SUCCESS'
            transaction.paid_at = datetime.utcnow()
            transaction.order.status = 'paid'

            # Reduce stock quantities
            for order_item in transaction.order.order_items:
                product = order_item.product
                if product.stock_quantity >= order_item.quantity:
                    product.stock_quantity -= order_item.quantity
                    print(
                        f"[DOKU] Reduced stock for {product.name}: {product.stock_quantity + order_item.quantity} -> {product.stock_quantity}"
                    )
                else:
                    print(
                        f"[DOKU] WARNING: Insufficient stock for {product.name}"
                    )

            db.session.commit()
            print(f"[DOKU] Payment {invoice_number} processed successfully")

    except Exception as e:
        print(f"Error processing DOKU payment success: {str(e)}")
        db.session.rollback()


def handle_doku_payment_failed(invoice_number):
    """Handle failed DOKU payment"""
    try:
        transaction = models.DokuTransaction.query.filter_by(
            invoice_number=invoice_number).first()

        if transaction:
            transaction.status = 'FAILED'
            transaction.order.status = 'cancelled'
            db.session.commit()
            print(f"[DOKU] Payment {invoice_number} marked as failed")

    except Exception as e:
        print(f"Error processing DOKU payment failure: {str(e)}")
        db.session.rollback()


# ===========================
# END XENDIT & DOKU PAYMENT ROUTES
# ===========================

# ===========================
# OFFLINE CASHIER SYSTEM ROUTES
# ===========================


@app.route('/cashier')
@login_required
@staff_required
def cashier_dashboard():
    """Main cashier interface for POS operations"""
    # Double check untuk memastikan hanya admin dan staff
    if current_user.role not in ['admin', 'staff']:
        flash(
            'Akses ditolak. Hanya admin dan staff yang dapat menggunakan kasir.',
            'error')
        return redirect(url_for('index'))

    # Check if cashier has active session
    active_session = models.CashierSession.query.filter_by(
        cashier_user_id=current_user.id, status='active').first()

    # Get products for POS interface
    products = models.Product.query.filter_by(is_active=True).all()
    categories = models.Category.query.all()

    return render_template('cashier/dashboard.html',
                           active_session=active_session,
                           products=products,
                           categories=categories)


@app.route('/cashier/session/start', methods=['POST'])
@login_required
@staff_required
@csrf.exempt
def start_cashier_session():
    """Start a new cashier session"""
    if current_user.role not in ['admin', 'staff']:
        return jsonify({
            'success':
            False,
            'error':
            'Akses ditolak. Hanya admin dan staff yang dapat menggunakan kasir.'
        }), 403

    # Check if user already has active session
    active_session = models.CashierSession.query.filter_by(
        cashier_user_id=current_user.id, status='active').first()

    if active_session:
        return jsonify({
            'success': False,
            'error': 'Anda sudah memiliki sesi kasir aktif'
        }), 400

    try:
        # Handle both JSON and form data
        if request.is_json:
            data = request.get_json() or {}
        else:
            data = request.form.to_dict()

        opening_cash = float(data.get('opening_cash', 0))

        session = models.CashierSession(cashier_user_id=current_user.id,
                                        opening_cash=opening_cash,
                                        status='active')

        db.session.add(session)
        db.session.commit()

        return jsonify({
            'success': True,
            'session': {
                'id': session.id,
                'cashier_user_id': session.cashier_user_id,
                'opening_cash': float(session.opening_cash),
                'status': session.status,
                'session_start': session.session_start.isoformat()
            },
            'message': 'Sesi kasir dimulai'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/cashier/session/close', methods=['POST'])
@login_required
@staff_required
@csrf.exempt
def close_cashier_session():
    """Close the current cashier session"""
    if current_user.role not in ['admin', 'staff']:
        return jsonify({
            'success':
            False,
            'error':
            'Akses ditolak. Hanya admin dan staff yang dapat menggunakan kasir.'
        }), 403

    try:
        # Handle both JSON and form data
        if request.is_json:
            data = request.get_json() or {}
        else:
            data = request.form.to_dict()

        closing_cash = float(data.get('closing_cash', 0))
        notes = data.get('notes', '')

        # Get active session
        session = models.CashierSession.query.filter_by(
            cashier_user_id=current_user.id, status='active').first()

        if not session:
            return jsonify({
                'success': False,
                'error': 'Tidak ada sesi kasir aktif'
            }), 400

        # Calculate expected cash and difference
        expected_cash = session.opening_cash + session.cash_sales
        cash_difference = closing_cash - expected_cash

        session.session_end = datetime.utcnow()
        session.closing_cash = closing_cash
        session.expected_cash = expected_cash
        session.cash_difference = cash_difference
        session.status = 'closed'
        session.notes = notes

        db.session.commit()

        return jsonify({
            'success': True,
            'session': {
                'id': session.id,
                'status': session.status,
                'closing_cash': float(session.closing_cash),
                'expected_cash': float(session.expected_cash),
                'cash_difference': float(session.cash_difference)
            },
            'message': 'Sesi kasir ditutup'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/cashier/connectivity')
@login_required
def api_cashier_connectivity():
    """Simple connectivity check for cashier system"""
    if current_user.role not in ['admin', 'staff']:
        return jsonify({'error': 'Akses ditolak'}), 403

    return jsonify({
        'status': 'online',
        'timestamp': datetime.utcnow().isoformat(),
        'server': 'ready'
    })


@app.route('/api/cashier/products')
@login_required
@staff_required
def api_cashier_products():
    """Get products for cashier interface"""
    if current_user.role not in ['admin', 'staff']:
        return jsonify({
            'error':
            'Akses ditolak. Hanya admin dan staff yang dapat menggunakan kasir.'
        }), 403

    try:
        category_id = request.args.get('category_id')
        search = request.args.get('search', '').strip()

        query = models.Product.query.filter_by(is_active=True)

        if category_id:
            query = query.filter_by(category_id=int(category_id))

        if search:
            query = query.filter(
                db.or_(
                    models.Product.id == int(search) if search.isdigit() else False,
                    models.Product.name.ilike(f'%{search}%'),
                    models.Product.description.ilike(f'%{search}%'),
                    models.Product.brand.ilike(f'%{search}%'),
                    models.Product.gtin.ilike(f'%{search}%')
                )
            )

        products = query.all()

        products_data = []
        for product in products:
            products_data.append({
                'id':
                product.id,
                'name':
                product.name,
                'price':
                float(product.price),
                'stock_quantity':
                product.stock_quantity,
                'image_url':
                product.image_url or '/static/images/placeholder.jpg',
                'brand':
                product.brand or '',
                'category':
                product.category.name if product.category else '',
                'formatted_price':
                f"Rp {product.price:,.0f}".replace(',', '.')
            })

        return jsonify({'products': products_data})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/cashier/transaction/save', methods=['POST'])
@login_required
@staff_required
@csrf.exempt
def save_cashier_transaction():
    """Save cashier POS transaction as offline order - NO SESSION REQUIRED"""
    if current_user.role not in ['admin', 'staff']:
        return jsonify({'success': False, 'error': 'Akses ditolak'}), 403

    try:
        data = request.get_json()

        # Validate required fields
        required_fields = [
            'local_transaction_id', 'total_amount', 'payment_method',
            'buyer_name', 'items'
        ]
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False,
                    'error': f'Field {field} wajib diisi'
                }), 400

        # Validate payment method (whitelist)
        valid_payment_methods = [
            'cash', 'debit', 'qris', 'transfer', 'ewallet', 'lainnya'
        ]
        if data['payment_method'].lower() not in valid_payment_methods:
            return jsonify({
                'success': False,
                'error': 'Metode pembayaran tidak valid'
            }), 400

        # Check for duplicate transaction (idempotency)
        existing_order = models.Order.query.filter_by(
            local_transaction_id=data['local_transaction_id']).first()

        if existing_order:
            return jsonify({
                'success': True,
                'order_id': existing_order.id,
                'message': 'Transaksi sudah tersimpan sebelumnya',
                'duplicate': True
            })

        # Calculate server-side total from authoritative prices
        server_total = 0
        order_items_data = []

        for item_data in data['items']:
            product = models.Product.query.get(item_data['product_id'])
            if not product:
                return jsonify({
                    'success':
                    False,
                    'error':
                    f'Produk dengan ID {item_data["product_id"]} tidak ditemukan'
                }), 400

            # Check stock availability
            quantity = int(item_data['quantity'])
            if product.stock_quantity < quantity:
                return jsonify({
                    'success':
                    False,
                    'error':
                    f'Stok tidak mencukupi untuk {product.name}. Stok tersedia: {product.stock_quantity}'
                }), 400

            # Use server's authoritative price, not client's
            item_total = float(product.price) * quantity
            server_total += item_total

            order_items_data.append({
                'product': product,
                'quantity': quantity,
                'price': float(product.price)
            })

        # Create Order with server-calculated total
        order = models.Order(
            user_id=current_user.id,  # Staff user sebagai penginput
            total_amount=server_total,  # Use SERVER-calculated total
            status='paid',  # Kasir langsung paid
            source_type='offline',  # Transaksi dari kasir/POS
            buyer_name=data['buyer_name'],
            payment_method=data['payment_method'].lower(),
            pos_user_id=current_user.id,  # Staff kasir yang melayani
            paid_at=get_utc_time(),  # Waktu pembayaran
            local_transaction_id=data[
                'local_transaction_id'],  # ID untuk idempotency
            shipping_cost=0,  # Kasir tidak ada ongkir
            created_at=get_utc_time())

        db.session.add(order)
        db.session.flush()  # Get order ID

        # Create order items and reduce stock atomically
        for item_info in order_items_data:
            order_item = models.OrderItem(order_id=order.id,
                                          product_id=item_info['product'].id,
                                          quantity=item_info['quantity'],
                                          price=item_info['price'])
            db.session.add(order_item)

            # Reduce stock quantity atomically
            item_info['product'].stock_quantity -= item_info['quantity']
            print(
                f"[KASIR] Reduced stock for {item_info['product'].name}: {item_info['product'].stock_quantity + item_info['quantity']} -> {item_info['product'].stock_quantity}"
            )

        db.session.commit()

        return jsonify({
            'success': True,
            'order_id': order.id,
            'message': 'Transaksi kasir berhasil disimpan',
            'receipt_number': f'POS-{order.id:06d}'
        })

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] Cashier transaction save failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/cashier/connectivity')
@login_required
@staff_required
def check_connectivity():
    """Simple endpoint to check if server is reachable"""
    if current_user.role not in ['admin', 'staff']:
        return jsonify({
            'error':
            'Akses ditolak. Hanya admin dan staff yang dapat menggunakan kasir.'
        }), 403

    return jsonify({
        'status':
        'online',
        'timestamp':
        datetime.utcnow().isoformat(),
        'server_time':
        datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    })


@app.route('/cashier/print-receipt')
@login_required
@staff_required
def print_receipt():
    """Generate thermal receipt PDF dari transaksi nyata"""
    transaction_id = request.args.get('transaction_id')
    paper_size = request.args.get('size', '80')

    if not transaction_id:
        return "Transaction ID required", 400

    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm as unit_mm
    import io

    # Lookup order by local_transaction_id (bukan id)
    order = models.Order.query.filter_by(local_transaction_id=transaction_id).first()
    if not order:
        return f"Transaksi {transaction_id} tidak ditemukan", 404

    # Ambil items dari order
    items = models.OrderItem.query.filter_by(order_id=order.id).all()

    # Hitung total dan pembayaran
    total = float(order.total_amount)
    tunai = float(request.args.get('tunai', total))  # bisa dari query ?tunai=
    kembali = max(tunai - total, 0)

    # Ukuran kertas
    paper_widths = {'58': 58 * unit_mm, '80': 80 * unit_mm, '120': 120 * unit_mm}
    width = paper_widths.get(paper_size, 80 * unit_mm)

    store_profile = models.StoreProfile.get_active_profile()

    buffer = io.BytesIO()
    base_height = 150 * unit_mm + len(items) * 8 * unit_mm
    p = canvas.Canvas(buffer, pagesize=(width, base_height))

    y = base_height - 10 * unit_mm

    # Header toko
    p.setFont("Helvetica-Bold", 12)
    p.drawCentredString(width / 2, y, store_profile.store_name if store_profile else "HURTROCK MUSIC STORE")
    y -= 5 * unit_mm

    p.setFont("Helvetica", 8)
    if store_profile:
        p.drawCentredString(width / 2, y, store_profile.formatted_address[:40])
        y -= 4 * unit_mm
        p.drawCentredString(width / 2, y, f"Telp: {store_profile.store_phone}")
        y -= 4 * unit_mm

    p.line(5 * unit_mm, y, width - 5 * unit_mm, y)
    y -= 5 * unit_mm

    from datetime import datetime
    now = datetime.now()
    p.drawString(5 * unit_mm, y, f"Tanggal: {now.strftime('%d/%m/%Y %H:%M')}")
    y -= 4 * unit_mm
    p.drawString(5 * unit_mm, y, f"Kasir: {current_user.name}")
    y -= 4 * unit_mm
    p.drawString(5 * unit_mm, y, f"No. Transaksi: {transaction_id[-10:]}")
    y -= 5 * unit_mm

    p.line(5 * unit_mm, y, width - 5 * unit_mm, y)
    y -= 5 * unit_mm

    # Header item
    p.setFont("Helvetica-Bold", 8)
    p.drawString(5 * unit_mm, y, "Item")
    p.drawRightString(width - 5 * unit_mm, y, "Subtotal")
    y -= 4 * unit_mm

    p.setFont("Helvetica", 8)
    for item in items:
        name = item.product.name if item.product else f"Produk #{item.product_id}"
        p.drawString(5 * unit_mm, y, f"{name[:25]}")
        y -= 3 * unit_mm
        p.drawString(7 * unit_mm, y, f"{item.quantity} x Rp {item.price:,.0f}".replace(",", "."))
        subtotal = item.quantity * item.price
        p.drawRightString(width - 5 * unit_mm, y, f"Rp {subtotal:,.0f}".replace(",", "."))
        y -= 4 * unit_mm

    # Garis total
    p.line(5 * unit_mm, y, width - 5 * unit_mm, y)
    y -= 5 * unit_mm

    # Total
    p.setFont("Helvetica-Bold", 10)
    p.drawString(5 * unit_mm, y, "TOTAL")
    p.drawRightString(width - 5 * unit_mm, y, f"Rp {total:,.0f}".replace(",", "."))
    y -= 5 * unit_mm

    # Tunai & Kembali
    p.setFont("Helvetica", 8)
    p.drawString(5 * unit_mm, y, f"Tunai: Rp {tunai:,.0f}".replace(",", "."))
    y -= 3 * unit_mm
    p.drawString(5 * unit_mm, y, f"Kembali: Rp {kembali:,.0f}".replace(",", "."))
    y -= 6 * unit_mm

    p.line(5 * unit_mm, y, width - 5 * unit_mm, y)
    y -= 5 * unit_mm

    p.setFont("Helvetica-Bold", 9)
    p.drawCentredString(width / 2, y, "TERIMA KASIH")
    y -= 4 * unit_mm
    p.setFont("Helvetica", 8)
    p.drawCentredString(width / 2, y, "Barang yang sudah dibeli")
    y -= 3 * unit_mm
    p.drawCentredString(width / 2, y, "tidak dapat dikembalikan")

    p.showPage()
    p.save()

    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf',
                     as_attachment=True,
                     download_name=f'struk_{transaction_id}.pdf')


@app.route('/cashier/transactions')
@login_required
@staff_required
def cashier_transactions():
    """View cashier transaction history"""
    if current_user.role not in ['admin', 'staff']:
        flash(
            'Akses ditolak. Hanya admin dan staff yang dapat melihat transaksi kasir.',
            'error')
        return redirect(url_for('index'))

    # Get filter parameters
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')

    # Base query
    query = models.OfflineTransaction.query.filter_by(
        cashier_user_id=current_user.id)

    # Apply date filters
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(
                models.OfflineTransaction.transaction_date >= date_from_obj)
        except:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to,
                                            '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(
                models.OfflineTransaction.transaction_date < date_to_obj)
        except:
            pass

    transactions = query.order_by(
        models.OfflineTransaction.transaction_date.desc()).all()

    # Calculate totals
    total_amount = sum(t.total_amount for t in transactions)
    total_count = len(transactions)

    return render_template('cashier/transactions.html',
                           transactions=transactions,
                           total_amount=total_amount,
                           total_count=total_count,
                           date_from=date_from,
                           date_to=date_to)


@app.route('/cashier/sessions')
@login_required
@staff_required
def cashier_sessions():
    """View cashier session history"""
    if current_user.role not in ['admin', 'staff']:
        flash(
            'Akses ditolak. Hanya admin dan staff yang dapat melihat sesi kasir.',
            'error')
        return redirect(url_for('index'))

    sessions = models.CashierSession.query.filter_by(
        cashier_user_id=current_user.id).order_by(
            models.CashierSession.session_start.desc()).all()

    return render_template('cashier/sessions.html', sessions=sessions)


@app.route('/cashier/transactions/export')
@login_required
@staff_required
def export_cashier_transactions():
    """Export cashier transactions to Excel with full analytics"""
    if current_user.role not in ['admin', 'staff']:
        flash(
            'Akses ditolak. Hanya admin dan staff yang dapat mengekspor transaksi.',
            'error')
        return redirect(url_for('index'))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from openpyxl.utils import get_column_letter

        # Get filter parameters
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')

        # Base query
        query = models.OfflineTransaction.query.filter_by(
            cashier_user_id=current_user.id)

        # Apply date filters
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                query = query.filter(models.OfflineTransaction.transaction_date
                                     >= date_from_obj)
            except:
                pass

        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to,
                                                '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(
                    models.OfflineTransaction.transaction_date < date_to_obj)
            except:
                pass

        transactions = query.order_by(
            models.OfflineTransaction.transaction_date.desc()).all()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Transaksi Kasir"

        # Add store logo/header
        profile = models.StoreProfile.get_active_profile()
        ws.merge_cells('A1:H1')
        header_cell = ws['A1']
        header_cell.value = profile.store_name if profile else "HURTROCK MUSIC STORE"
        header_cell.font = Font(size=18, bold=True, color="FF6B35")
        header_cell.alignment = Alignment(horizontal='center',
                                          vertical='center')

        ws.merge_cells('A2:H2')
        subheader_cell = ws['A2']
        subheader_cell.value = "LAPORAN TRANSAKSI KASIR"
        subheader_cell.font = Font(size=14, bold=True)
        subheader_cell.alignment = Alignment(horizontal='center',
                                             vertical='center')

        # Add report info
        ws.merge_cells('A3:H3')
        info_cell = ws['A3']
        date_range = ""
        if date_from and date_to:
            date_range = f"Periode: {date_from} s/d {date_to}"
        elif date_from:
            date_range = f"Mulai: {date_from}"
        elif date_to:
            date_range = f"Sampai: {date_to}"
        else:
            date_range = "Semua Transaksi"
        info_cell.value = f"{date_range} | Kasir: {current_user.name}"
        info_cell.font = Font(size=10)
        info_cell.alignment = Alignment(horizontal='center')

        # Blank row
        ws.append([])

        # Define column headers
        headers = [
            'No', 'ID Transaksi', 'Tanggal & Waktu', 'Nama Pembeli',
            'Metode Pembayaran', 'Daftar Barang', 'Total', 'Status'
        ]

        # Style for headers
        header_fill = PatternFill(start_color='FF6B35',
                                  end_color='FF6B35',
                                  fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Write headers
        header_row = ws.max_row + 1
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write transaction data
        for idx, transaction in enumerate(transactions, start=1):
            # Compile items list
            items_list = []
            for item in transaction.offline_items:
                items_list.append(
                    f"{item.product_name} x{item.quantity} (@Rp {item.product_price:,.0f})"
                )
            items_str = "; ".join(items_list)

            row_data = [
                idx, transaction.local_transaction_id,
                transaction.transaction_date.strftime('%d/%m/%Y %H:%M:%S'),
                transaction.customer_name or '-',
                transaction.payment_method.title(), items_str,
                float(transaction.total_amount), 'Tersinkron'
                if transaction.sync_status == 'synced' else 'Pending'
            ]

            row_num = ws.max_row + 1
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                cell.border = border

                # Format currency
                if col == 7:  # Total column
                    cell.number_format = 'Rp #,##0'
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left',
                                               vertical='center',
                                               wrap_text=True)

        # Add summary
        ws.append([])
        summary_row = ws.max_row + 1
        ws.merge_cells(f'A{summary_row}:F{summary_row}')
        summary_label = ws[f'A{summary_row}']
        summary_label.value = "TOTAL PENJUALAN"
        summary_label.font = Font(bold=True, size=12)
        summary_label.alignment = Alignment(horizontal='right')

        total_amount = sum(t.total_amount for t in transactions)
        total_cell = ws[f'G{summary_row}']
        total_cell.value = float(total_amount)
        total_cell.number_format = 'Rp #,##0'
        total_cell.font = Font(bold=True, size=12, color='FF6B35')
        total_cell.alignment = Alignment(horizontal='right')
        total_cell.fill = PatternFill(start_color='FFF3E0',
                                      end_color='FFF3E0',
                                      fill_type='solid')

        # Adjust column widths
        column_widths = [5, 25, 20, 20, 18, 50, 15, 12]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # Freeze header rows
        ws.freeze_panes = 'A6'

        # Save to bytes buffer
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Generate filename
        filename = f"Laporan_Kasir_{current_user.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            excel_file,
            mimetype=
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename)

    except Exception as e:
        print(f"[ERROR] Excel export failed: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Gagal mengekspor data: {str(e)}', 'error')
        return redirect(url_for('cashier_transactions'))


# ===========================
# END OFFLINE CASHIER SYSTEM ROUTES
# ===========================

# ===========================
# EXCEL IMPORT FUNCTIONALITY
# ===========================


@app.route('/admin/products/import')
@login_required
def admin_products_import():
    """Excel import interface for bulk product upload"""
    if current_user.role != 'admin':
        flash('Akses ditolak. Hanya admin yang dapat mengimpor produk.',
              'error')
        return redirect(url_for('index'))

    categories = models.Category.query.all()
    return render_template('admin/products_import.html', categories=categories)


@app.route('/admin/products/import/template')
@login_required
def download_import_template():
    """Download Excel template for product import"""
    if current_user.role != 'admin':
        flash('Akses ditolak. Hanya admin yang dapat mengunduh template.',
              'error')
        return redirect(url_for('index'))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Template Produk"

        # Define headers
        headers = [
            'nama_produk', 'deskripsi', 'harga', 'stok_awal', 'brand',
            'kategori', 'berat_gram', 'dimensi_panjang', 'dimensi_lebar',
            'dimensi_tinggi', 'warna', 'material', 'garansi_bulan', 'aktif'
        ]

        # Write headers with styling
        header_fill = PatternFill(start_color='4472C4',
                                  end_color='4472C4',
                                  fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Add instructions in second row
        instructions = [
            'Wajib diisi', 'Opsional', 'Angka (contoh: 150000)',
            'Angka (contoh: 10)', 'Opsional', 'Nama kategori yang ada',
            'Angka dalam gram', 'Angka dalam cm', 'Angka dalam cm',
            'Angka dalam cm', 'Opsional', 'Opsional', 'Angka bulan garansi',
            'true/false'
        ]

        for col, instruction in enumerate(instructions, 1):
            cell = ws.cell(row=2, column=col, value=instruction)
            cell.font = Font(italic=True, color='666666')

        # Add sample data
        sample_data = [
            [
                'Gitar Akustik Yamaha F310',
                'Gitar akustik pemula dengan kualitas suara yang baik',
                1500000, 5, 'Yamaha', 'Gitar', 2500, 103, 38, 11, 'Natural',
                'Kayu Spruce', 6, 'true'
            ],
            [
                'Bass Elektrik Fender Jazz',
                'Bass elektrik 4 senar dengan pickup jazz', 8500000, 2,
                'Fender', 'Bass', 4000, 117, 32, 4, 'Sunburst', 'Kayu Alder',
                12, 'true'
            ],
            [
                'Drum Set Pearl Export',
                'Drum set lengkap 5 piece untuk pemula hingga menengah',
                12000000, 1, 'Pearl', 'Drum', 25000, 140, 120, 150, 'Hitam',
                'Poplar/Mahogany', 24, 'true'
            ]
        ]

        for row_idx, row_data in enumerate(sample_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Adjust column widths
        column_widths = [
            20, 30, 15, 12, 15, 15, 12, 15, 15, 15, 12, 15, 15, 10
        ]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(
                row=1, column=col).column_letter].width = width

        # Add validation sheet
        ws_validation = wb.create_sheet("Instruksi")
        instructions_text = [
            "INSTRUKSI PENGGUNAAN TEMPLATE IMPORT PRODUK", "", "Format File:",
            "- Gunakan file Excel (.xlsx)",
            "- Jangan mengubah nama kolom di baris pertama",
            "- Mulai input data dari baris ke-3 (baris ke-2 adalah contoh)",
            "", "Kolom Wajib:",
            "- nama_produk: Nama produk (maksimal 200 karakter)",
            "- harga: Harga jual dalam rupiah (angka tanpa titik/koma)",
            "- stok_awal: Jumlah stok awal (angka bulat)", "",
            "Kolom Opsional:", "- deskripsi: Deskripsi produk",
            "- brand: Merek produk",
            "- kategori: Nama kategori yang sudah ada di sistem",
            "- berat_gram: Berat produk dalam gram",
            "- dimensi_*: Dimensi dalam cm", "- warna: Warna produk",
            "- material: Material/bahan produk",
            "- garansi_bulan: Lama garansi dalam bulan",
            "- aktif: true untuk produk aktif, false untuk non-aktif", "",
            "Tips:",
            "- Pastikan nama kategori sesuai dengan yang ada di sistem",
            "- Gunakan angka untuk kolom numerik (tanpa satuan)",
            "- Maksimal 1000 produk per file", "- File maksimal 10MB"
        ]

        for row, instruction in enumerate(instructions_text, 1):
            ws_validation.cell(row=row, column=1, value=instruction)

        ws_validation.column_dimensions['A'].width = 60

        # Create downloadable file
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = send_file(
            output,
            mimetype=
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='template_import_produk.xlsx')
        return response

    except Exception as e:
        print(f"Error creating template: {str(e)}")
        flash(f'Error creating template: {str(e)}', 'error')
        return redirect(url_for('admin_products_import'))


@app.route('/admin/products/import/upload', methods=['POST'])
@login_required
def upload_products_excel():
    """Process uploaded Excel file and import products"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Akses ditolak'}), 403

    try:
        if 'excel_file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'File tidak ditemukan'
            }), 400

        file = request.files['excel_file']
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'File tidak dipilih'
            }), 400

        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({
                'success': False,
                'error': 'Format file harus Excel (.xlsx atau .xls)'
            }), 400

        # Check file size (max 10MB)
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning

        if file_size > 10 * 1024 * 1024:  # 10MB
            return jsonify({
                'success': False,
                'error': 'File terlalu besar (maksimal 10MB)'
            }), 400

        from openpyxl import load_workbook

        # Load workbook
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        # Validate headers
        expected_headers = [
            'nama_produk', 'deskripsi', 'harga', 'stok_awal', 'brand',
            'kategori', 'berat_gram', 'dimensi_panjang', 'dimensi_lebar',
            'dimensi_tinggi', 'warna', 'material', 'garansi_bulan', 'aktif'
        ]

        actual_headers = []
        for col in range(1, len(expected_headers) + 1):
            cell_value = ws.cell(row=1, column=col).value
            actual_headers.append(
                cell_value.strip().lower() if cell_value else '')

        if actual_headers != expected_headers:
            return jsonify({
                'success':
                False,
                'error':
                f'Header tidak sesuai template. Expected: {expected_headers}, Got: {actual_headers}'
            }), 400

        # Get all categories for validation
        categories = {
            cat.name.lower(): cat.id
            for cat in models.Category.query.all()
        }

        # Process data rows
        imported_count = 0
        error_count = 0
        errors = []

        max_row = ws.max_row
        if max_row > 1002:  # 1000 data rows + 2 header rows
            return jsonify({
                'success': False,
                'error': 'Maksimal 1000 produk per file'
            }), 400

        for row_num in range(3, max_row +
                             1):  # Start from row 3 (skip headers and example)
            try:
                # Read row data
                row_data = []
                for col in range(1, len(expected_headers) + 1):
                    cell_value = ws.cell(row=row_num, column=col).value
                    row_data.append(cell_value)

                # Skip empty rows
                if not any(row_data):
                    continue

                # Validate required fields
                nama_produk = str(row_data[0]).strip() if row_data[0] else ''
                harga = row_data[2]
                stok_awal = row_data[3]

                if not nama_produk:
                    errors.append(f'Baris {row_num}: Nama produk wajib diisi')
                    error_count += 1
                    continue

                if not harga or not isinstance(harga,
                                               (int, float)) or harga <= 0:
                    errors.append(
                        f'Baris {row_num}: Harga harus berupa angka positif')
                    error_count += 1
                    continue

                if not isinstance(stok_awal, (int, float)) or stok_awal < 0:
                    errors.append(
                        f'Baris {row_num}: Stok awal harus berupa angka non-negatif'
                    )
                    error_count += 1
                    continue

                # Validate category
                kategori_name = str(
                    row_data[5]).strip().lower() if row_data[5] else ''
                category_id = None
                if kategori_name:
                    if kategori_name in categories:
                        category_id = categories[kategori_name]
                    else:
                        errors.append(
                            f'Baris {row_num}: Kategori "{row_data[5]}" tidak ditemukan'
                        )
                        error_count += 1
                        continue

                # Check if product already exists
                existing_product = models.Product.query.filter_by(
                    name=nama_produk).first()
                if existing_product:
                    errors.append(
                        f'Baris {row_num}: Produk "{nama_produk}" sudah ada')
                    error_count += 1
                    continue

                # Parse optional fields
                deskripsi = str(row_data[1]).strip() if row_data[1] else ''
                brand = str(row_data[4]).strip() if row_data[4] else ''
                berat_gram = int(row_data[6]) if row_data[6] and isinstance(
                    row_data[6], (int, float)) else None
                dimensi_panjang = float(
                    row_data[7]) if row_data[7] and isinstance(
                        row_data[7], (int, float)) else None
                dimensi_lebar = float(
                    row_data[8]) if row_data[8] and isinstance(
                        row_data[8], (int, float)) else None
                dimensi_tinggi = float(
                    row_data[9]) if row_data[9] and isinstance(
                        row_data[9], (int, float)) else None
                warna = str(row_data[10]).strip() if row_data[10] else ''
                material = str(row_data[11]).strip() if row_data[11] else ''
                garansi_bulan = int(
                    row_data[12]) if row_data[12] and isinstance(
                        row_data[12], (int, float)) else None

                # Parse aktif field
                aktif_value = str(
                    row_data[13]).strip().lower() if row_data[13] else 'true'
                is_active = aktif_value in ['true', '1', 'yes', 'ya', 'aktif']

                # Create product
                product = models.Product(
                    name=nama_produk,
                    description=deskripsi if deskripsi else None,
                    price=float(harga),
                    stock_quantity=int(stok_awal),
                    brand=brand if brand else None,
                    category_id=category_id,
                    weight_grams=berat_gram,
                    length_cm=dimensi_panjang,
                    width_cm=dimensi_lebar,
                    height_cm=dimensi_tinggi,
                    color=warna if warna else None,
                    material=material if material else None,
                    warranty_months=garansi_bulan,
                    is_active=is_active)

                db.session.add(product)
                imported_count += 1

            except Exception as e:
                errors.append(f'Baris {row_num}: Error - {str(e)}')
                error_count += 1
                continue

        # Commit if there are successful imports
        if imported_count > 0:
            db.session.commit()
            print(
                f"[EXCEL IMPORT] Successfully imported {imported_count} products"
            )
        else:
            db.session.rollback()

        # Prepare response
        response_data = {
            'success': True,
            'imported_count': imported_count,
            'error_count': error_count,
            'errors': errors[:50],  # Limit error messages
            'message':
            f'Import selesai: {imported_count} produk berhasil diimpor'
        }

        if error_count > 0:
            response_data['message'] += f', {error_count} error'

        return jsonify(response_data)

    except Exception as e:
        db.session.rollback()
        print(f"Error importing Excel: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error processing file: {str(e)}'
        }), 500


@app.route('/admin/products/import/validate', methods=['POST'])
@login_required
def validate_excel_file():
    """Validate Excel file before import (preview mode)"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Akses ditolak'}), 403

    try:
        if 'excel_file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'File tidak ditemukan'
            }), 400

        file = request.files['excel_file']
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'File tidak dipilih'
            }), 400

        from openpyxl import load_workbook

        # Load and validate file
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        # Check headers
        expected_headers = [
            'nama_produk', 'deskripsi', 'harga', 'stok_awal', 'brand',
            'kategori', 'berat_gram', 'dimensi_panjang', 'dimensi_lebar',
            'dimensi_tinggi', 'warna', 'material', 'garansi_bulan', 'aktif'
        ]

        actual_headers = []
        for col in range(1, len(expected_headers) + 1):
            cell_value = ws.cell(row=1, column=col).value
            actual_headers.append(
                cell_value.strip().lower() if cell_value else '')

        if actual_headers != expected_headers:
            return jsonify({
                'success': False,
                'error': 'Header tidak sesuai template',
                'expected_headers': expected_headers,
                'actual_headers': actual_headers
            }), 400

        # Count data rows
        data_rows = 0
        for row_num in range(3, ws.max_row + 1):
            row_data = [
                ws.cell(row=row_num, column=col).value
                for col in range(1,
                                 len(expected_headers) + 1)
            ]
            if any(row_data):
                data_rows += 1

        return jsonify({
            'success': True,
            'total_rows': data_rows,
            'message': f'File valid dengan {data_rows} baris data'
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error validating file: {str(e)}'
        }), 500


# ===========================
# END EXCEL IMPORT FUNCTIONALITY
# ===========================


# API endpoint for cart count
@app.route('/api/cart/count')
@login_required
def api_cart_count():
    try:
        if current_user.role != 'buyer':
            return jsonify({'count': 0})

        count = models.CartItem.query.filter_by(
            user_id=current_user.id).count()
        return jsonify({'count': count})
    except Exception as e:
        print(f"Error getting cart count: {e}")
        return jsonify({'count': 0})


# API endpoint for chat service to get product info
@app.route('/api/products/<int:product_id>')
def api_get_product(product_id):
    try:
        product = models.Product.query.get_or_404(product_id)

        # Ensure product has a slug
        if not product.slug:
            product.ensure_slug()
            db.session.commit()

        # Generate product URL using slug if available, otherwise use ID
        if product.slug:
            product_url = url_for('product_detail',
                                  slug=product.slug,
                                  _external=False)
        else:
            product_url = url_for('product_detail_by_id',
                                  product_id=product.id,
                                  _external=False)

        return jsonify({
            'id':
            product.id,
            'name':
            product.name,
            'slug':
            product.slug,
            'url':
            product_url,
            'price':
            float(product.price),
            'image_url':
            product.image_url or '/static/images/placeholder.jpg',
            'brand':
            product.brand or '',
            'description':
            product.description or '',
            'category':
            product.category.name if product.category else '',
            'is_active':
            product.is_active
        })
    except Exception as e:
        print(f"Error getting product {product_id}: {e}")
        return jsonify({'error': 'Product not found'}), 404


# API endpoint for JWT token (for chat service)
@app.route('/api/chat/token')
@login_required
def api_chat_token():
    try:
        # Check if Django service is running
        if not check_django_service():
            print(
                "[WARNING] Django chat service not responding, attempting to start..."
            )
            if not start_django_service():
                print("[ERROR] Failed to start Django service.")
                return jsonify({'error': 'Chat service unavailable'}), 503
            # Wait a bit for the service to start
            import time
            time.sleep(5)  # Give it a few seconds

        token = generate_jwt_token(current_user)
        return jsonify({
            'token': token,
            'expires_in': JWT_ACCESS_TOKEN_LIFETIME,
            'user': {
                'id': current_user.id,
                'name': current_user.name,
                'email': current_user.email,
                'role': current_user.role
            }
        })
    except Exception as e:
        print(f"Error generating chat token: {str(e)}")
        return jsonify({'error': 'Failed to generate token'}), 500


# API endpoint for chat media upload
@app.route('/api/chat/upload-media', methods=['POST'])
@login_required
def api_chat_upload_media():
    try:
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No file provided'
            }), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'No file selected'
            }), 400

        if not allowed_chat_media(file.filename):
            return jsonify({
                'success': False,
                'error': 'File type not allowed'
            }), 400

        filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4()}_{filename}"

        is_video = is_video_file(filename)
        subfolder = 'videos' if is_video else 'images'

        upload_path = os.path.join(app.config['CHAT_MEDIA_FOLDER'], subfolder)
        os.makedirs(upload_path, exist_ok=True)

        filepath = os.path.join(upload_path, unique_filename)
        file.save(filepath)

        if not is_video:
            compress_image(filepath, max_size_mb=2)

        file_url = f"/static/chat_media/{subfolder}/{unique_filename}"

        return jsonify({
            'success': True,
            'file_url': file_url,
            'file_type': 'video' if is_video else 'image',
            'filename': unique_filename
        })

    except Exception as e:
        print(f"Error uploading chat media: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


# Proxy routes for chat service with /chat prefix
@app.route('/chat/<path:path>')
def proxy_chat_service(path):
    """Proxy all /chat requests to Django service"""
    try:
        import requests

        # Build target URL
        target_url = f"http://127.0.0.1:8000/{path}"

        # Forward query parameters
        if request.query_string:
            target_url += f"?{request.query_string.decode()}"

        # Forward the request
        if request.method == 'GET':
            response = requests.get(target_url,
                                    headers=dict(request.headers),
                                    timeout=10)
        elif request.method == 'POST':
            response = requests.post(target_url,
                                     headers=dict(request.headers),
                                     data=request.get_data(),
                                     timeout=10)
        else:
            response = requests.request(request.method,
                                        target_url,
                                        headers=dict(request.headers),
                                        data=request.get_data(),
                                        timeout=10)

        # Return response
        return response.content, response.status_code, dict(response.headers)

    except Exception as e:
        print(f"Error proxying to chat service: {str(e)}")
        return jsonify({'error': 'Chat service unavailable'}), 503


# WebSocket proxy route for chat
@app.route('/ws/<path:path>')
def proxy_websocket(path):
    """Proxy WebSocket requests to Django service"""
    # For WebSocket, we need to redirect to the actual WebSocket endpoint
    # This is just a fallback route - actual WebSocket handling should be done by the server
    return jsonify({'error': 'WebSocket proxy not available via HTTP'}), 400


# Chat service proxy endpoints
@app.route('/api/admin/buyer-rooms/')
@login_required
@admin_required
def proxy_buyer_rooms():
    try:
        search_query = request.args.get('search', '')

        # Generate JWT token for Django service
        jwt_token = generate_jwt_token(current_user)

        # Try multiple endpoints for Django service
        endpoints = [
            'http://127.0.0.1:8000/api/admin/buyer-rooms/',
            'http://localhost:8000/api/admin/buyer-rooms/',
            'http://0.0.0.0:8000/api/admin/buyer-rooms/'
        ]

        if search_query:
            endpoints = [f"{url}?search={search_query}" for url in endpoints]

        headers = {
            'Authorization': f'Bearer {jwt_token}',
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        }

        last_error = None
        for chat_service_url in endpoints:
            try:
                print(f"[DEBUG] Trying Django endpoint: {chat_service_url}")
                response = requests.get(chat_service_url,
                                        headers=headers,
                                        timeout=10)

                if response.status_code == 200:
                    return jsonify(response.json()), response.status_code
                else:
                    last_error = f"Status {response.status_code}: {response.text}"
                    continue

            except requests.exceptions.RequestException as e:
                last_error = str(e)
                continue

        print(f"All Django endpoints failed. Last error: {last_error}")
        return jsonify({
            'error': 'Chat service unavailable',
            'rooms': [],
            'total_count': 0
        }), 503

    except Exception as e:
        print(f"Unexpected error in proxy_buyer_rooms: {str(e)}")
        return jsonify({
            'error': 'Internal server error',
            'rooms': [],
            'total_count': 0
        }), 500


@app.route('/api/rooms/<room_name>/messages/')
@login_required
def proxy_room_messages(room_name):
    try:
        # Generate JWT token for Django service
        jwt_token = generate_jwt_token(current_user)

        # Try multiple endpoints
        endpoints = [
            f"http://127.0.0.1:8000/api/rooms/{room_name}/messages/",
            f"http://localhost:8000/api/rooms/{room_name}/messages/",
            f"http://0.0.0.0:8000/api/rooms/{room_name}/messages/"
        ]

        # Forward query parameters
        query_params = request.args.to_dict()
        if query_params:
            query_string = "&".join(
                [f"{k}={v}" for k, v in query_params.items()])
            endpoints = [f"{url}?{query_string}" for url in endpoints]

        headers = {
            'Authorization': f'Bearer {jwt_token}',
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        }

        last_error = None
        for chat_service_url in endpoints:
            try:
                print(
                    f"[DEBUG] Trying room messages endpoint: {chat_service_url}"
                )
                response = requests.get(chat_service_url,
                                        headers=headers,
                                        timeout=10)

                if response.status_code == 200:
                    try:
                        return jsonify(response.json()), response.status_code
                    except ValueError as e:
                        print(f"JSON decode error: {str(e)}")
                        return jsonify({
                            'error': 'Invalid response from chat service',
                            'results': []
                        }), 502
                else:
                    last_error = f"Status {response.status_code}: {response.text}"
                    continue

            except requests.exceptions.RequestException as e:
                last_error = str(e)
                continue

        print(f"All Django endpoints failed. Last error: {last_error}")
        return jsonify({
            'error': 'Chat service unavailable',
            'results': []
        }), 503

    except Exception as e:
        print(f"Unexpected error in proxy_room_messages: {str(e)}")
        return jsonify({'error': 'Internal server error', 'results': []}), 500


@app.route('/api/rooms/<room_name>/mark-read/', methods=['POST'])
@login_required
def proxy_mark_room_read(room_name):
    try:
        # Check if Django service is running
        if not check_django_service():
            print("[WARNING] Django chat service not responding")
            return jsonify({'error': 'Chat service unavailable'}), 503

        # Try multiple endpoints
        endpoints = [
            f"http://127.0.0.1:8000/api/rooms/{room_name}/mark-read/",
            f"http://localhost:8000/api/rooms/{room_name}/mark-read/",
            f"http://0.0.0.0:8000/api/rooms/{room_name}/mark-read/"
        ]

        headers = {
            'Authorization': f'Bearer {generate_jwt_token(current_user)}',
            'Content-Type': 'application/json'
        }

        last_error = None
        for chat_service_url in endpoints:
            try:
                response = requests.post(chat_service_url,
                                         headers=headers,
                                         timeout=5)

                if response.status_code == 200:
                    try:
                        return jsonify(response.json()), response.status_code
                    except ValueError:
                        return jsonify({'message':
                                        'Messages marked as read'}), 200
                else:
                    last_error = f"Status {response.status_code}: {response.text}"
                    continue

            except requests.exceptions.RequestException as e:
                last_error = str(e)
                continue

        print(f"All Django endpoints failed. Last error: {last_error}")
        return jsonify({'error': 'Chat service unavailable'}), 503

    except Exception as e:
        print(f"Unexpected error in proxy_mark_room_read: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/admin/orders')
@login_required
@staff_required
def admin_orders():
    orders = models.Order.query.order_by(models.Order.created_at.desc()).all()
    return render_template('admin/orders.html', orders=orders)


@app.route('/admin/order/<int:order_id>/update', methods=['POST'])
@login_required
@staff_required
def admin_update_order(order_id):
    order = models.Order.query.get_or_404(order_id)

    new_status = request.form.get('status')
    tracking_number = request.form.get('tracking_number', '').strip()
    courier_service = request.form.get('courier_service', '').strip()

    if new_status not in [
            'pending', 'paid', 'shipped', 'delivered', 'cancelled'
    ]:
        flash('Status tidak valid!', 'error')
        return redirect(url_for('admin_orders'))

    order.status = new_status
    order.tracking_number = tracking_number if tracking_number else None
    order.courier_service = courier_service if courier_service else None
    order.updated_at = datetime.utcnow()

    db.session.commit()

    flash(f'Pesanan #{order.id} berhasil diperbarui!', 'success')
    return redirect(url_for('admin_orders'))


@app.route('/admin/order/<int:order_id>/quick-ship', methods=['POST'])
@login_required
@staff_required
def admin_quick_ship_order(order_id):
    order = models.Order.query.get_or_404(order_id)

    # Pastikan order masih dalam status paid
    if order.status != 'paid':
        flash(f'Pesanan #{order.id} tidak dalam status yang dapat dikirim!',
              'error')
        return redirect(url_for('admin_orders'))

    tracking_number = request.form.get('tracking_number', '').strip()
    courier_service = request.form.get('courier_service', '').strip()

    if not tracking_number:
        tracking_number = generate_tracking_number()

    try:
        order.tracking_number = tracking_number
        order.courier_service = courier_service
        order.status = 'shipped'
        db.session.commit()

        flash(
            f'Pesanan #{order.id} berhasil dikirim dengan nomor resi: {tracking_number}',
            'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error saat mengupdate status pengiriman: {str(e)}', 'error')

    return redirect(url_for('admin_orders'))


@app.route('/admin/order/<int:order_id>/print_professional_label')
@login_required
@staff_required
def print_professional_label(order_id):
    """
    Generate simple thermal label for 120mm printer
    Standard ReportLab format, clean and readable
    """
    order = models.Order.query.get_or_404(order_id)
    store_profile = models.StoreProfile.get_active_profile()

    # Generate tracking number if not exists
    if not order.tracking_number:
        order.tracking_number = generate_tracking_number()
        db.session.commit()

    buffer = io.BytesIO()

    # 120mm thermal printer width (340 points = 120mm)
    width = 340
    height = 480  # Adjustable height
    p = canvas.Canvas(buffer, pagesize=(width, height))

    y_pos = height - 20
    margin = 15

    # Header - Store Name
    p.setFont("Helvetica-Bold", 14)
    store_name = store_profile.store_name if store_profile else "Hurtrock Music Store"
    text_width = p.stringWidth(store_name, "Helvetica-Bold", 14)
    p.drawString((width - text_width) / 2, y_pos, store_name)
    y_pos -= 20

    # Separator line
    p.setLineWidth(1)
    p.line(margin, y_pos, width - margin, y_pos)
    y_pos -= 15

    # Order number
    p.setFont("Helvetica-Bold", 12)
    order_text = f"PESANAN #{order.id:06d}"
    text_width = p.stringWidth(order_text, "Helvetica-Bold", 12)
    p.drawString((width - text_width) / 2, y_pos, order_text)
    y_pos -= 15

    # Tracking number
    if order.tracking_number:
        p.setFont("Helvetica", 10)
        tracking_text = f"Resi: {order.tracking_number}"
        text_width = p.stringWidth(tracking_text, "Helvetica", 10)
        p.drawString((width - text_width) / 2, y_pos, tracking_text)
        y_pos -= 15

    # Date
    p.setFont("Helvetica", 9)
    date_text = f"Tanggal: {order.created_at.strftime('%d/%m/%Y %H:%M')}"
    text_width = p.stringWidth(date_text, "Helvetica", 9)
    p.drawString((width - text_width) / 2, y_pos, date_text)
    y_pos -= 20

    # Separator
    p.line(margin, y_pos, width - margin, y_pos)
    y_pos -= 15

    # DARI (Sender)
    p.setFont("Helvetica-Bold", 10)
    p.drawString(margin, y_pos, "DARI:")
    y_pos -= 12

    p.setFont("Helvetica", 9)
    if store_profile:
        p.drawString(margin, y_pos, store_profile.store_name)
        y_pos -= 10

        # Split address into multiple lines
        address = store_profile.formatted_address
        words = address.split()
        lines = []
        current_line = ""
        for word in words:
            if len(current_line + word) <= 35:
                current_line += word + " "
            else:
                if current_line:
                    lines.append(current_line.strip())
                current_line = word + " "
        if current_line:
            lines.append(current_line.strip())

        for line in lines[:3]:  # Max 3 lines
            p.drawString(margin, y_pos, line)
            y_pos -= 10

        if store_profile.store_phone:
            p.drawString(margin, y_pos, f"Telp: {store_profile.store_phone}")
            y_pos -= 10
    else:
        p.drawString(margin, y_pos, "Hurtrock Music Store")
        y_pos -= 10
        p.drawString(margin, y_pos, "Jakarta, Indonesia")
        y_pos -= 10

    y_pos -= 5

    # KEPADA (Recipient)
    p.setFont("Helvetica-Bold", 10)
    p.drawString(margin, y_pos, "KEPADA:")
    y_pos -= 12

    p.setFont("Helvetica-Bold", 9)
    p.drawString(margin, y_pos, order.user.name.upper())
    y_pos -= 12

    p.setFont("Helvetica", 8)
    if order.user.phone:
        p.drawString(margin, y_pos, f"Telp: {order.user.phone}")
        y_pos -= 10

    # Recipient address
    if order.user.address:
        address = order.user.address.replace('\n', ' ')
        words = address.split()
        lines = []
        current_line = ""
        for word in words:
            if len(current_line + word) <= 35:
                current_line += word + " "
            else:
                if current_line:
                    lines.append(current_line.strip())
                current_line = word + " "
        if current_line:
            lines.append(current_line.strip())

        for line in lines[:4]:  # Max 4 lines
            p.drawString(margin, y_pos, line)
            y_pos -= 10

    y_pos -= 5

    # Separator
    p.line(margin, y_pos, width - margin, y_pos)
    y_pos -= 15

    # Items
    p.setFont("Helvetica-Bold", 9)
    p.drawString(margin, y_pos, "BARANG:")
    y_pos -= 12

    p.setFont("Helvetica", 8)
    for item in order.order_items[:5]:  # Max 5 items
        item_name = item.product.name
        if len(item_name) > 30:
            item_name = item_name[:27] + "..."
        p.drawString(margin, y_pos, f"• {item_name}")
        y_pos -= 9
        p.drawString(margin + 10, y_pos, f"  {item.quantity}pcs")
        y_pos -= 10

    if len(order.order_items) > 5:
        p.drawString(margin, y_pos,
                     f"• +{len(order.order_items) - 5} item lainnya")
        y_pos -= 10

    y_pos -= 5

    # Total and weight
    p.setFont("Helvetica-Bold", 10)
    total_text = f"TOTAL: {order.formatted_total}"
    text_width = p.stringWidth(total_text, "Helvetica-Bold", 10)
    p.drawString((width - text_width) / 2, y_pos, total_text)
    y_pos -= 15

    # Weight
    total_weight = sum(item.quantity * (item.product.weight or 100)
                       for item in order.order_items) / 1000
    p.setFont("Helvetica", 9)
    weight_text = f"Berat: {total_weight:.1f} kg"
    text_width = p.stringWidth(weight_text, "Helvetica", 9)
    p.drawString((width - text_width) / 2, y_pos, weight_text)
    y_pos -= 15

    # Service info
    if order.courier_service:
        service_text = f"Kurir: {order.courier_service}"
        text_width = p.stringWidth(service_text, "Helvetica", 9)
        p.drawString((width - text_width) / 2, y_pos, service_text)
        y_pos -= 12

    # Footer
    p.setFont("Helvetica", 7)
    footer_text = "Terima kasih atas kepercayaan Anda"
    text_width = p.stringWidth(footer_text, "Helvetica", 7)
    p.drawString((width - text_width) / 2, 15, footer_text)

    p.showPage()
    p.save()

    buffer.seek(0)
    return send_file(buffer,
                     as_attachment=True,
                     download_name=f'thermal_label_{order.id}.pdf',
                     mimetype='application/pdf')


@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    users = models.User.query.all()
    return render_template('admin/users.html', users=users)


@app.route('/admin/invoices')
@login_required
@staff_required
def admin_invoices():
    """List all invoices"""
    invoices = models.Invoice.query.order_by(models.Invoice.created_at.desc()).all()
    return render_template('admin/invoices.html', invoices=invoices)


@app.route('/admin/invoice/create/<int:order_id>', methods=['POST'])
@login_required
@staff_required
@csrf.exempt
def admin_create_invoice(order_id):
    """Create invoice from order"""
    try:
        order = models.Order.query.get_or_404(order_id)
        
        # Check if invoice already exists for this order
        existing_invoice = models.Invoice.query.filter_by(order_id=order_id).first()
        if existing_invoice:
            flash('Invoice sudah ada untuk pesanan ini!', 'warning')
            return redirect(url_for('admin_invoices'))
        
        # Generate invoice number
        invoice_number = f"INV-{datetime.utcnow().strftime('%Y%m%d')}-{order.id:06d}"
        
        # Create invoice
        invoice = models.Invoice(
            invoice_number=invoice_number,
            order_id=order_id,
            # Gunakan buyer_name untuk POS (offline), dan nama user untuk online
            customer_name=order.buyer_name or (order.user.name if hasattr(order, 'user') and order.user else 'Walk-in Customer'),
            customer_email=(order.user.email if hasattr(order, 'user') and order.user else ''),
            customer_phone=(order.user.phone if hasattr(order, 'user') and order.user else ''),
            customer_address=order.shipping_address or '',
            subtotal=sum(item.quantity * item.unit_price for item in order.order_items),
            tax_amount=0,
            discount_amount=0,
            shipping_cost=order.shipping_cost or 0,
            total_amount=order.total_amount,
            status='Paid' if order.status in ['paid', 'shipped', 'delivered'] else 'Pending',
            payment_method=order.payment_method or '',
            issued_by=current_user.id
        )
        
        db.session.add(invoice)
        db.session.commit()
        
        flash(f'Invoice {invoice_number} berhasil dibuat!', 'success')
        return redirect(url_for('admin_invoices'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal membuat invoice: {str(e)}', 'error')
        return redirect(url_for('admin_orders'))


@app.route('/admin/invoice/<int:invoice_id>/export-excel')
@login_required
@staff_required
def admin_export_invoice_excel(invoice_id):
    """Export invoice to Excel dengan branding Hurtrock"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage
        import os
        
        invoice = models.Invoice.query.get_or_404(invoice_id)
        order = invoice.order if invoice.order_id else None
        
        # Get store profile from database
        store_profile = models.StoreProfile.get_active_profile()
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Invoice {invoice.invoice_number}"
        
        # Define colors (White background, Black text theme)
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        light_gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Define fonts (Black text)
        title_font = Font(name='Arial', size=20, bold=True, color="000000")
        header_font = Font(name='Arial', size=12, bold=True, color="000000")
        normal_font = Font(name='Arial', size=10, color="000000")
        bold_font = Font(name='Arial', size=10, bold=True, color="000000")
        
        # Define alignment
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        
        # Header - Store Name "Hurtrock Music Store"
        ws.merge_cells('A1:E1')
        cell = ws['A1']
        cell.value = "HURTROCK MUSIC STORE"
        cell.font = title_font
        cell.fill = white_fill
        cell.alignment = center_align
        ws.row_dimensions[1].height = 35
        
        # 5 rows spacing for store profile info
        row = 2
        
        # Store Address
        if store_profile and store_profile.store_address:
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws[f'A{row}']
            cell.value = f"Alamat: {store_profile.store_address}"
            cell.font = normal_font
            cell.alignment = center_align
            ws.row_dimensions[row].height = 18
            row += 1
        
        # Store Phone
        if store_profile and store_profile.store_phone:
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws[f'A{row}']
            cell.value = f"Telp: {store_profile.store_phone}"
            cell.font = normal_font
            cell.alignment = center_align
            ws.row_dimensions[row].height = 18
            row += 1
        
        # Store Email
        if store_profile and store_profile.store_email:
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws[f'A{row}']
            cell.value = f"Email: {store_profile.store_email}"
            cell.font = normal_font
            cell.alignment = center_align
            ws.row_dimensions[row].height = 18
            row += 1
        
        # Additional contact info if available
        if store_profile and hasattr(store_profile, 'whatsapp_number') and store_profile.whatsapp_number:
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws[f'A{row}']
            cell.value = f"WhatsApp: {store_profile.whatsapp_number}"
            cell.font = normal_font
            cell.alignment = center_align
            ws.row_dimensions[row].height = 18
            row += 1
        
        # Website/Social media if available
        if store_profile and hasattr(store_profile, 'website') and store_profile.website:
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws[f'A{row}']
            cell.value = f"Website: {store_profile.website}"
            cell.font = normal_font
            cell.alignment = center_align
            ws.row_dimensions[row].height = 18
            row += 1
        
        # Blank row
        row += 1
        ws.row_dimensions[row].height = 10
        
        # Invoice Title
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "INVOICE"
        cell.font = Font(name='Arial', size=16, bold=True, color="000000")
        cell.alignment = center_align
        ws.row_dimensions[row].height = 25
        invoice_title_row = row
        
        # Invoice details
        row += 2
        ws[f'A{row}'] = "Invoice No:"
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = invoice.invoice_number
        
        row += 1
        ws[f'A{row}'] = "Tanggal:"
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = invoice.created_at.strftime('%d-%m-%Y %H:%M')
        
        row += 1
        ws[f'A{row}'] = "Status:"
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = invoice.status
        
        # Customer details
        row += 2
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "CUSTOMER INFORMATION"
        cell.font = header_font
        cell.fill = light_gray_fill
        cell.alignment = center_align
        ws.row_dimensions[row].height = 25
        
        row += 1
        ws[f'A{row}'] = "Nama:"
        ws[f'A{row}'].font = bold_font
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = invoice.customer_name
        
        row += 1
        ws[f'A{row}'] = "Email:"
        ws[f'A{row}'].font = bold_font
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = invoice.customer_email or '-'
        
        row += 1
        ws[f'A{row}'] = "Telepon:"
        ws[f'A{row}'].font = bold_font
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = invoice.customer_phone or '-'
        
        row += 1
        ws[f'A{row}'] = "Alamat:"
        ws[f'A{row}'].font = bold_font
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = invoice.customer_address or '-'
        
        # Items table
        row += 2
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "ITEMS"
        cell.font = header_font
        cell.fill = light_gray_fill
        cell.alignment = center_align
        ws.row_dimensions[row].height = 25
        
        # Table headers
        row += 1
        headers = ['No', 'Product', 'Qty', 'Price', 'Subtotal']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(name='Arial', size=10, bold=True, color="000000")
            cell.fill = light_gray_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Items data - support both order items and manual invoice items
        if order and order.order_items:
            # Invoice dari order
            for idx, item in enumerate(order.order_items, start=1):
                row += 1
                ws.cell(row=row, column=1, value=idx).alignment = center_align
                ws.cell(row=row, column=2, value=item.product.name).alignment = left_align
                ws.cell(row=row, column=3, value=item.quantity).alignment = center_align
                ws.cell(row=row, column=4, value=f"Rp {item.price:,.0f}").alignment = right_align
                ws.cell(row=row, column=5, value=f"Rp {item.subtotal:,.0f}").alignment = right_align
                
                # Apply borders
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = thin_border
        elif invoice.items:
            # Invoice manual dengan custom items
            for idx, item in enumerate(invoice.items, start=1):
                row += 1
                ws.cell(row=row, column=1, value=idx).alignment = center_align
                ws.cell(row=row, column=2, value=item.item_name).alignment = left_align
                ws.cell(row=row, column=3, value=item.quantity).alignment = center_align
                ws.cell(row=row, column=4, value=f"Rp {item.unit_price:,.0f}").alignment = right_align
                ws.cell(row=row, column=5, value=f"Rp {item.subtotal:,.0f}").alignment = right_align
                
                # Apply borders
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = thin_border
        
        # Totals
        row += 2
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "Subtotal:"
        cell.font = bold_font
        cell.alignment = right_align
        ws[f'E{row}'] = f"Rp {invoice.subtotal:,.0f}"
        ws[f'E{row}'].alignment = right_align
        ws[f'E{row}'].font = normal_font
        
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "Shipping Cost:"
        cell.font = bold_font
        cell.alignment = right_align
        ws[f'E{row}'] = f"Rp {invoice.shipping_cost:,.0f}"
        ws[f'E{row}'].alignment = right_align
        ws[f'E{row}'].font = normal_font
        
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "TOTAL:"
        cell.font = Font(name='Arial', size=12, bold=True, color="000000")
        cell.fill = light_gray_fill
        cell.alignment = right_align
        cell_total = ws[f'E{row}']
        cell_total.value = f"Rp {invoice.total_amount:,.0f}"
        cell_total.font = Font(name='Arial', size=12, bold=True, color="000000")
        cell_total.fill = light_gray_fill
        cell_total.alignment = right_align
        
        # Footer
        row += 3
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "Terima kasih atas kepercayaan Anda kepada Hurtrock Music Store!"
        cell.font = Font(name='Arial', size=10, italic=True)
        cell.alignment = center_align
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'invoice_{invoice.invoice_number}.xlsx'
        )
        
    except Exception as e:
        print(f"[ERROR] Failed to export invoice: {str(e)}")
        flash(f'Gagal export invoice: {str(e)}', 'error')
        return redirect(url_for('admin_invoices'))


@app.route('/admin/invoice/<int:invoice_id>/status', methods=['POST'])
@login_required
@staff_required
def admin_update_invoice_status(invoice_id):
    """Update invoice status"""
    try:
        invoice = models.Invoice.query.get_or_404(invoice_id)
        new_status = request.form.get('status')
        
        if new_status not in ['Paid', 'Pending', 'On Process']:
            flash('Status tidak valid!', 'error')
            return redirect(url_for('admin_invoices'))
        
        invoice.status = new_status
        invoice.updated_at = datetime.utcnow()
        db.session.commit()
        
        flash(f'Status invoice {invoice.invoice_number} berhasil diupdate!', 'success')
        return redirect(url_for('admin_invoices'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal update status: {str(e)}', 'error')
        return redirect(url_for('admin_invoices'))


@app.route('/admin/invoice/create-manual', methods=['POST'])
@login_required
@staff_required
@csrf.exempt
def admin_create_manual_invoice():
    """Create manual invoice with custom items"""
    try:
        # --- Get form data ---
        customer_name = request.form.get('customer_name')
        customer_email = request.form.get('customer_email', '')
        customer_phone = request.form.get('customer_phone', '')
        customer_address = request.form.get('customer_address', '')
        payment_method = request.form.get('payment_method', '')
        status = request.form.get('status', 'Pending')
        notes = request.form.get('notes', '')

        # --- Get items from JSON ---
        items_json = request.form.get('items')
        if not items_json:
            flash('Items tidak boleh kosong!', 'error')
            return redirect(url_for('admin_invoices'))

        items = json.loads(items_json)
        if not items:
            flash('Minimal satu item harus diisi!', 'error')
            return redirect(url_for('admin_invoices'))

        # --- Calculate totals ---
        subtotal = Decimal('0')
        for item in items:
            qty = Decimal(str(item.get('quantity', 0)))
            price = Decimal(str(item.get('unit_price', 0)))
            subtotal += qty * price

        shipping_cost = Decimal(request.form.get('shipping_cost', '0'))
        tax_amount = Decimal(request.form.get('tax_amount', '0'))
        discount_amount = Decimal(request.form.get('discount_amount', '0'))

        total_amount = subtotal + shipping_cost + tax_amount - discount_amount

        # --- Generate invoice number ---
        invoice_number = f"INV-{datetime.utcnow().strftime('%Y%m%d')}-{models.Invoice.query.count() + 1:06d}"

        # --- Create invoice record ---
        invoice = models.Invoice(
            invoice_number=invoice_number,
            order_id=None,  # Manual invoice, no order
            customer_name=customer_name,
            customer_email=customer_email,
            customer_phone=customer_phone,
            customer_address=customer_address,
            subtotal=subtotal,
            tax_amount=tax_amount,
            discount_amount=discount_amount,
            shipping_cost=shipping_cost,
            total_amount=total_amount,
            status=status,
            payment_method=payment_method,
            notes=notes,
            issued_by=current_user.id
        )
        db.session.add(invoice)
        db.session.flush()  # Get invoice ID for invoice items

        # --- Create invoice items ---
        for item in items:
            invoice_item = models.InvoiceItem(
                invoice_id=invoice.id,
                item_name=item.get('item_name', 'Unnamed Item'),
                description=item.get('description', ''),
                quantity=int(item.get('quantity', 0)),
                unit_price=Decimal(str(item.get('unit_price', 0)))
            )
            db.session.add(invoice_item)

        # --- Commit transaction ---
        db.session.commit()

        flash(f'Invoice {invoice_number} berhasil dibuat!', 'success')
        return redirect(url_for('admin_invoices'))

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] Failed to create manual invoice: {str(e)}")
        flash(f'Gagal membuat invoice: {str(e)}', 'error')
        return redirect(url_for('admin_invoices'))

@app.route('/admin/analytics')
@login_required
@admin_required
def admin_analytics():
    from sqlalchemy import func, case
    from decimal import Decimal

    try:
        # Total penjualan hari ini (GABUNGAN ONLINE + KASIR)
        today_start = datetime.utcnow().replace(hour=0,
                                                minute=0,
                                                second=0,
                                                microsecond=0)
        today_end = today_start + timedelta(days=1)

        today_sales = db.session.query(
            func.coalesce(func.sum(models.Order.total_amount), 0)).filter(
                models.Order.created_at >= today_start, models.Order.created_at
                < today_end,
                models.Order.status.in_(['paid', 'shipped', 'delivered'
                                         ])).scalar() or Decimal('0')

        # Breakdown penjualan hari ini per sumber
        today_online_sales = db.session.query(
            func.coalesce(func.sum(models.Order.total_amount), 0)).filter(
                models.Order.created_at >= today_start, models.Order.created_at
                < today_end,
                models.Order.status.in_(['paid', 'shipped', 'delivered']),
                models.Order.source_type == 'online').scalar() or Decimal('0')

        today_offline_sales = db.session.query(
            func.coalesce(func.sum(models.Order.total_amount), 0)).filter(
                models.Order.created_at >= today_start, models.Order.created_at
                < today_end,
                models.Order.status.in_(['paid', 'shipped', 'delivered']),
                models.Order.source_type == 'offline').scalar() or Decimal('0')

        # Total penjualan bulan ini (GABUNGAN ONLINE + KASIR)
        now = datetime.utcnow()
        month_start = now.replace(day=1,
                                  hour=0,
                                  minute=0,
                                  second=0,
                                  microsecond=0)
        next_month = month_start + timedelta(days=32)
        month_end = next_month.replace(day=1)

        monthly_sales = db.session.query(
            func.coalesce(func.sum(models.Order.total_amount), 0)).filter(
                models.Order.created_at >= month_start, models.Order.created_at
                < month_end,
                models.Order.status.in_(['paid', 'shipped', 'delivered'
                                         ])).scalar() or Decimal('0')

        # Breakdown penjualan bulanan per sumber
        monthly_online_sales = db.session.query(
            func.coalesce(func.sum(models.Order.total_amount), 0)).filter(
                models.Order.created_at >= month_start, models.Order.created_at
                < month_end,
                models.Order.status.in_(['paid', 'shipped', 'delivered']),
                models.Order.source_type == 'online').scalar() or Decimal('0')

        monthly_offline_sales = db.session.query(
            func.coalesce(func.sum(models.Order.total_amount), 0)).filter(
                models.Order.created_at >= month_start, models.Order.created_at
                < month_end,
                models.Order.status.in_(['paid', 'shipped', 'delivered']),
                models.Order.source_type == 'offline').scalar() or Decimal('0')

        # Produk terlaris (GABUNGAN ONLINE + KASIR) dengan explicit join
        best_selling_products = db.session.query(
            models.Product.name,
            func.sum(models.OrderItem.quantity).label('total_sold'),
            func.sum(case(
                (models.Order.source_type == 'online', models.OrderItem.quantity),
                else_=0
            )).label('sold_online'),
            func.sum(case(
                (models.Order.source_type == 'offline', models.OrderItem.quantity),
                else_=0
            )).label('sold_offline')
        ).select_from(models.Product)\
        .join(models.OrderItem, models.OrderItem.product_id == models.Product.id)\
        .join(models.Order, models.Order.id == models.OrderItem.order_id)\
        .filter(
            models.Order.status.in_(['paid', 'shipped', 'delivered'])
        ).group_by(models.Product.id, models.Product.name).order_by(
            func.sum(models.OrderItem.quantity).desc()
        ).limit(10).all()

        # Penjualan harian 7 hari terakhir
        seven_days_ago = today_start - timedelta(days=6)
        daily_sales_raw = db.session.query(
            func.date_trunc('day', models.Order.created_at).label('date'),
            func.coalesce(func.sum(models.Order.total_amount), 0).label('total'),
            func.count(models.Order.id).label('orders_count')
        ).filter(
            models.Order.created_at >= seven_days_ago,
            models.Order.status.in_(['paid', 'shipped', 'delivered'])
        ).group_by(func.date_trunc('day', models.Order.created_at))\
        .order_by(func.date_trunc('day', models.Order.created_at)).all()

        daily_sales = [{
            'date': row.date,
            'total': row.total or 0,
            'orders_count': row.orders_count or 0
        } for row in daily_sales_raw]

        # Pelanggan terbaik (Top 5 customers)
        top_customers_raw = db.session.query(
            models.User.name,
            models.User.email,
            func.count(models.Order.id).label('orders_count'),
            func.coalesce(func.sum(models.Order.total_amount), 0).label('total_spent')
        ).select_from(models.User)\
        .join(models.Order, models.Order.user_id == models.User.id)\
        .filter(
            models.Order.status.in_(['paid', 'shipped', 'delivered']),
            models.User.role == 'buyer'
        ).group_by(models.User.id, models.User.name, models.User.email)\
        .order_by(func.coalesce(func.sum(models.Order.total_amount), 0).desc()).limit(5).all()

        top_customers = [{
            'name': row.name,
            'email': row.email,
            'orders_count': row.orders_count or 0,
            'total_spent': row.total_spent or 0
        } for row in top_customers_raw]

        return render_template('admin/analytics.html',
                               today_sales=today_sales,
                               today_online_sales=today_online_sales,
                               today_offline_sales=today_offline_sales,
                               monthly_sales=monthly_sales,
                               monthly_online_sales=monthly_online_sales,
                               monthly_offline_sales=monthly_offline_sales,
                               best_selling_products=best_selling_products,
                               daily_sales=daily_sales,
                               top_customers=top_customers,
                               category_sales=[],
                               current_date=datetime.utcnow().date(),
                               datetime=datetime)
    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] Analytics error: {str(e)}")
        import traceback
        traceback.print_exc()

        # Return template with zero values on error
        return render_template('admin/analytics.html',
                               today_sales=Decimal('0'),
                               today_online_sales=Decimal('0'),
                               today_offline_sales=Decimal('0'),
                               monthly_sales=Decimal('0'),
                               monthly_online_sales=Decimal('0'),
                               monthly_offline_sales=Decimal('0'),
                               best_selling_products=[],
                               daily_sales=[],
                               top_customers=[],
                               category_sales=[],
                               current_date=datetime.utcnow().date(),
                               datetime=datetime)


@app.route('/admin/user/<int:user_id>/change_role', methods=['POST'])
@login_required
@admin_required
def admin_change_user_role(user_id):
    user = models.User.query.get_or_404(user_id)
    new_role = request.form.get('role')

    if new_role not in ['admin', 'staff', 'buyer']:
        flash('Role tidak valid!', 'error')
        return redirect(url_for('admin_users'))

    # Prevent changing own role
    if user.id == current_user.id:
        flash('Tidak bisa mengubah role sendiri!', 'error')
        return redirect(url_for('admin_users'))

    user.role = new_role
    db.session.commit()

    flash(f'Role {user.name} berhasil diubah menjadi {new_role}!', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/user/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_user(user_id):
    user = models.User.query.get_or_404(user_id)

    # Prevent deleting own account
    if user.id == current_user.id:
        flash('Tidak bisa menghapus akun sendiri!', 'error')
        return redirect(url_for('admin_users'))

    # Check if user has orders
    if user.orders:
        flash('Tidak bisa menghapus user yang memiliki riwayat pesanan!',
              'error')
        return redirect(url_for('admin_users'))

    user_name = user.name
    db.session.delete(user)
    db.session.commit()

    flash(f'User {user_name} berhasil dihapus!', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/chat')
@login_required
@staff_required
def admin_chat():
    """Admin chat interface for customer service"""
    return render_template('admin/chat_interface.html')


@app.route('/admin/chat/analytics')
@login_required
@admin_required
def admin_chat_analytics():
    """Real-time chat analytics dashboard"""
    try:
        # Get chat service API stats
        chat_api_url = 'http://127.0.0.1:8000/api/admin/stats/'
        headers = {'Authorization': f'Bearer {generate_jwt_token(current_user)}'}
        
        try:
            response = requests.get(chat_api_url, headers=headers, timeout=5)
            if response.status_code == 200:
                api_stats = response.json()
            else:
                api_stats = {
                    'total_rooms': 0,
                    'active_rooms': 0,
                    'total_messages': 0,
                    'unread_messages': 0,
                    'avg_response_time': 0
                }
        except:
            api_stats = {
                'total_rooms': 0,
                'active_rooms': 0,
                'total_messages': 0,
                'unread_messages': 0,
                'avg_response_time': 0,
                'error': 'Chat service tidak tersedia'
            }
        
        # Get recent chat activity from API
        recent_chats_url = 'http://127.0.0.1:8000/api/admin/rooms/'
        try:
            response = requests.get(recent_chats_url, headers=headers, timeout=5)
            if response.status_code == 200:
                recent_chats = response.json().get('results', [])[:10]
            else:
                recent_chats = []
        except:
            recent_chats = []
        
        return render_template('admin/chat_analytics.html',
                             stats=api_stats,
                             recent_chats=recent_chats)
    except Exception as e:
        print(f"[ERROR] Chat analytics error: {e}")
        flash(f'Error loading chat analytics: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route('/admin/chat/deleted-messages')
@login_required
@admin_required
def admin_chat_deleted_messages():
    """View deleted chat messages for fraud prevention"""
    try:
        # Get deleted messages from chat API
        deleted_api_url = 'http://127.0.0.1:8000/api/messages/?is_deleted=true'
        headers = {'Authorization': f'Bearer {generate_jwt_token(current_user)}'}
        
        try:
            response = requests.get(deleted_api_url, headers=headers, timeout=5)
            if response.status_code == 200:
                deleted_messages = response.json().get('results', [])
            else:
                deleted_messages = []
                flash('Tidak dapat mengambil data pesan yang dihapus', 'warning')
        except Exception as api_error:
            print(f"[ERROR] API error: {api_error}")
            deleted_messages = []
            flash('Chat service tidak tersedia', 'error')
        
        return render_template('admin/chat_deleted_messages.html',
                             deleted_messages=deleted_messages)
    except Exception as e:
        print(f"[ERROR] Deleted messages error: {e}")
        flash(f'Error loading deleted messages: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route('/chat')
@login_required
def chat():
    """Chat interface for buyers - placeholder page"""
    if current_user.is_admin or current_user.is_staff:
        return redirect(url_for('admin_chat'))

    # For now, redirect to a coming soon page or show message
    flash(
        'Fitur chat sedang dalam pengembangan. Silakan hubungi admin melalui WhatsApp.',
        'info')
    return redirect(url_for('index'))


@app.route('/admin/users/add', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_add_user():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        password = request.form['password']
        role = request.form['role']

        # Validate role
        if role not in ['admin', 'staff', 'buyer']:
            flash('Role tidak valid!', 'error')
            return render_template('admin/add_user.html')

        # Check if email already exists
        if models.User.query.filter_by(email=email).first():
            flash('Email sudah terdaftar!', 'error')
            return render_template('admin/add_user.html')

        # Create new user
        hashed_password = generate_password_hash(password)
        user = models.User(name=name,
                           email=email,
                           password_hash=hashed_password,
                           role=role)

        db.session.add(user)
        db.session.commit()

        flash(f'User {name} berhasil ditambahkan dengan role {role}!',
              'success')
        return redirect(url_for('admin_users'))

    return render_template('admin/add_user.html')


@app.route('/admin/user/<int:user_id>/reset_password', methods=['POST'])
@login_required
@admin_required
def admin_reset_password(user_id):
    user = models.User.query.get_or_404(user_id)
    new_password = request.form.get('new_password')

    if not new_password or len(new_password) < 6:
        flash('Password harus minimal 6 karakter!', 'error')
        return redirect(url_for('admin_users'))

    user.password_hash = generate_password_hash(new_password)
    db.session.commit()

    flash(f'Password {user.name} berhasil direset!', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/export/sales/<period>')
@login_required
@staff_required
def export_sales(period):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash(
            'Package openpyxl diperlukan untuk export Excel. Silakan install terlebih dahulu.',
            'error')
        return redirect(url_for('admin_analytics'))

    from sqlalchemy import func
    import io
    import calendar
    from collections import defaultdict

    # Get store profile
    store_profile = models.StoreProfile.get_active_profile()
    store_name = store_profile.store_name if store_profile else "Hurtrock Music Store"

    # Determine date range based on period
    today = datetime.utcnow().date()
    current_month = today.month
    current_year = today.year

    # Handle custom date range
    if period == 'custom':
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        if not start_date_str or not end_date_str:
            flash('Tanggal mulai dan akhir harus diisi untuk custom range!',
                  'error')
            return redirect(url_for('admin_analytics'))

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except:
            flash('Format tanggal tidak valid!', 'error')
            return redirect(url_for('admin_analytics'))

        period_text = f"Custom - {start_date.strftime('%d %B %Y')} s/d {end_date.strftime('%d %B %Y')}"

    elif period == 'daily':
        # Today only - per hari individual
        start_date = today
        end_date = today
        period_text = f"Harian - {today.strftime('%d %B %Y')}"

    elif period == 'weekly':
        # Last 7 days
        start_date = today - timedelta(days=6)
        end_date = today
        period_text = f"Mingguan - {start_date.strftime('%d %B %Y')} s/d {end_date.strftime('%d %B %Y')}"

    elif period == 'monthly':
        # This month
        start_date = today.replace(day=1)
        end_date = today
        period_text = f"Bulanan - {start_date.strftime('%B %Y')}"

    else:
        flash('Periode tidak valid!', 'error')
        return redirect(url_for('admin_analytics'))

    # Get detailed sales data with products
    start_datetime = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())

    try:
        # Get all orders in period
        orders = models.Order.query.filter(
            models.Order.created_at >= start_datetime, models.Order.created_at
            <= end_datetime,
            models.Order.status.in_(['paid', 'shipped',
                                     'delivered'])).order_by(
                                         models.Order.created_at.desc()).all()

        # Compile DETAILED product sales data dengan harga database
        product_sales = defaultdict(
            lambda: {
                'product_name': '',
                'product_price': 0,  # Harga dari database
                'online_qty': 0,
                'offline_qty': 0,
                'online_amount': 0,
                'offline_amount': 0,
                'payment_methods': defaultdict(lambda: {
                    'online': 0,
                    'offline': 0
                })  # Detail per channel
            })

        total_online = 0
        total_offline = 0
        payment_method_totals = defaultdict(lambda: {
            'online': 0,
            'offline': 0
        })

        for order in orders:
            is_online = order.source_type == 'online'
            # Standardize payment method names
            payment_method_raw = (order.payment_method
                                  or 'tidak diketahui').lower()

            # Map payment methods dengan detail channel
            if 'virtual' in payment_method_raw or 'va' in payment_method_raw or 'transfer' in payment_method_raw:
                if is_online:
                    payment_method = 'Virtual Account (Web/Marketplace)'
                else:
                    payment_method = 'Transfer Bank (Kasir)'
            elif payment_method_raw == 'cash':
                payment_method = 'Cash (Kasir Toko)'
            elif payment_method_raw == 'debit':
                payment_method = 'Debit Card (Kasir)'
            elif payment_method_raw == 'qris':
                if is_online:
                    payment_method = 'QRIS (Web/Marketplace)'
                else:
                    payment_method = 'QRIS (Kasir Toko)'
            elif 'ewallet' in payment_method_raw or 'wallet' in payment_method_raw:
                payment_method = 'E-Wallet (OVO/Dana/LinkAja)'
            elif 'credit' in payment_method_raw or 'kartu kredit' in payment_method_raw:
                payment_method = 'Kartu Kredit'
            else:
                payment_method = payment_method_raw.title()

            for item in order.order_items:
                product_id = item.product_id
                product_name = item.product.name if item.product else f"Product ID {product_id}"
                product_price = float(
                    item.product.price) if item.product else float(item.price)

                product_sales[product_id]['product_name'] = product_name
                product_sales[product_id][
                    'product_price'] = product_price  # Harga dari database

                channel = 'online' if is_online else 'offline'

                if is_online:
                    product_sales[product_id]['online_qty'] += item.quantity
                    product_sales[product_id]['online_amount'] += float(
                        item.subtotal)
                    total_online += float(item.subtotal)
                else:
                    product_sales[product_id]['offline_qty'] += item.quantity
                    product_sales[product_id]['offline_amount'] += float(
                        item.subtotal)
                    total_offline += float(item.subtotal)

                # Track payment per channel
                product_sales[product_id]['payment_methods'][payment_method][
                    channel] += float(item.subtotal)
                payment_method_totals[payment_method][channel] += float(
                    item.subtotal)

    except Exception as query_error:
        print(f"[ERROR] Query failed: {query_error}")
        db.session.rollback()
        flash(f'Error mengambil data penjualan: {str(query_error)}', 'error')
        return redirect(url_for('admin_analytics'))

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Laporan {period.title()}"

    # Set page orientation to landscape
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # Define styles - White Background, Black Text Theme
    title_font = Font(name='Arial', size=18, bold=True, color='000000')
    subtitle_font = Font(name='Arial', size=12, bold=True, color='000000')
    normal_font = Font(name='Arial', size=10, color='000000')
    table_header_font = Font(name='Arial', size=11, bold=True, color='000000')
    subtotal_font = Font(name='Arial', size=11, bold=True, color='000000')
    total_font = Font(name='Arial', size=12, bold=True, color='000000')

    # White background, light gray for headers
    white_fill = PatternFill(start_color='FFFFFF',
                              end_color='FFFFFF',
                              fill_type='solid')
    light_gray_fill = PatternFill(start_color='F5F5F5',
                                    end_color='F5F5F5',
                                    fill_type='solid')

    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Company Header - "Hurtrock Music Store"
    ws.merge_cells('A1:I1')
    company_cell = ws['A1']
    company_cell.value = "HURTROCK MUSIC STORE"
    company_cell.font = title_font
    company_cell.fill = white_fill
    company_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Report Title
    ws.merge_cells('A2:I2')
    title_cell = ws['A2']
    title_cell.value = "LAPORAN PENJUALAN DETAIL"
    title_cell.font = Font(name='Arial', size=14, bold=True, color='000000')
    title_cell.fill = white_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # 5 rows for store profile info
    row = 3
    
    # Store Address
    if store_profile and store_profile.store_address:
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = f"Alamat: {store_profile.store_address}"
        cell.font = Font(name='Arial', size=9, color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 18
        row += 1
    
    # Store Phone
    if store_profile and store_profile.store_phone:
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = f"Telp: {store_profile.store_phone}"
        cell.font = Font(name='Arial', size=9, color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 18
        row += 1
    
    # Store Email
    if store_profile and store_profile.store_email:
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = f"Email: {store_profile.store_email}"
        cell.font = Font(name='Arial', size=9, color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 18
        row += 1
    
    # WhatsApp if available
    if store_profile and hasattr(store_profile, 'whatsapp_number') and store_profile.whatsapp_number:
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = f"WhatsApp: {store_profile.whatsapp_number}"
        cell.font = Font(name='Arial', size=9, color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 18
        row += 1
    
    # Website if available
    if store_profile and hasattr(store_profile, 'website') and store_profile.website:
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = f"Website: {store_profile.website}"
        cell.font = Font(name='Arial', size=9, color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 18
        row += 1

    # Period info
    row += 1
    ws.merge_cells(f'A{row}:I{row}')
    period_cell = ws[f'A{row}']
    period_cell.value = f"Periode: {period_text}"
    period_cell.font = subtitle_font
    period_cell.fill = light_gray_fill
    period_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 20

    # Empty row
    row += 1
    ws.row_dimensions[row].height = 5

    # Table headers - Light gray background, black text
    row += 1
    table_header_row = row
    headers = [
        'No', 'Nama Produk', 'Harga Satuan (DB)', 'Qty Online', 'Nilai Online',
        'Qty Kasir', 'Nilai Kasir', 'Total Qty', 'Total Nilai',
        'Metode Pembayaran (Detail per Channel)'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = table_header_font
        cell.fill = light_gray_fill
        cell.alignment = Alignment(horizontal='center',
                                   vertical='center',
                                   wrap_text=True)
        cell.border = border
    ws.row_dimensions[row].height = 35

    # Data rows - Per product detail dengan harga database
    row_num = row + 1
    total_online_amount = 0
    total_offline_amount = 0

    for idx, (product_id, data) in enumerate(
            sorted(
                product_sales.items(),
                key=lambda x: x[1]['online_amount'] + x[1]['offline_amount'],
                reverse=True), 1):
        # No
        ws.cell(row=row_num, column=1, value=idx).border = border
        ws.cell(row=row_num,
                column=1).alignment = Alignment(horizontal='center')

        # Product Name
        ws.cell(row=row_num, column=2,
                value=data['product_name']).border = border
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='left')
        ws.cell(row=row_num, column=2).font = normal_font

        # Harga Satuan (dari database) - Format IDR
        price_cell = ws.cell(row=row_num, column=3)
        price_cell.value = f"Rp {data['product_price']:,.0f}".replace(',', '.')
        price_cell.border = border
        price_cell.alignment = Alignment(horizontal='right')
        price_cell.font = Font(name='Arial',
                               size=10,
                               bold=True,
                               color='000000')

        # Qty Online
        ws.cell(row=row_num, column=4,
                value=data['online_qty']).border = border
        ws.cell(row=row_num,
                column=4).alignment = Alignment(horizontal='center')

        # Nilai Online - Format IDR
        online_cell = ws.cell(row=row_num, column=5)
        online_cell.value = f"Rp {data['online_amount']:,.0f}".replace(
            ',', '.')
        online_cell.border = border
        online_cell.alignment = Alignment(horizontal='right')

        # Qty Kasir
        ws.cell(row=row_num, column=6,
                value=data['offline_qty']).border = border
        ws.cell(row=row_num,
                column=6).alignment = Alignment(horizontal='center')

        # Nilai Kasir - Format IDR
        offline_cell = ws.cell(row=row_num, column=7)
        offline_cell.value = f"Rp {data['offline_amount']:,.0f}".replace(
            ',', '.')
        offline_cell.border = border
        offline_cell.alignment = Alignment(horizontal='right')

        # Total Qty
        total_qty = data['online_qty'] + data['offline_qty']
        ws.cell(row=row_num, column=8, value=total_qty).border = border
        ws.cell(row=row_num,
                column=8).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=8).font = Font(name='Arial',
                                                   size=10,
                                                   bold=True)

        # Total Nilai - Format IDR
        total_amount = data['online_amount'] + data['offline_amount']
        total_cell = ws.cell(row=row_num, column=9)
        total_cell.value = f"Rp {total_amount:,.0f}".replace(',', '.')
        total_cell.border = border
        total_cell.alignment = Alignment(horizontal='right')
        total_cell.font = Font(name='Arial', size=10, bold=True)

        # Payment Methods - DETAIL PER CHANNEL
        payment_details = []
        for method, channels in sorted(
                data['payment_methods'].items(),
                key=lambda x: x[1]['online'] + x[1]['offline'],
                reverse=True):
            if channels['online'] > 0:
                payment_details.append(
                    f"{method} [Online]: Rp {channels['online']:,.0f}".replace(
                        ',', '.'))
            if channels['offline'] > 0:
                payment_details.append(
                    f"{method} [Kasir]: Rp {channels['offline']:,.0f}".replace(
                        ',', '.'))

        payment_text = '; '.join(payment_details) if payment_details else '-'
        ws.cell(row=row_num, column=10, value=payment_text).border = border
        ws.cell(row=row_num,
                column=10).alignment = Alignment(horizontal='left',
                                                 wrap_text=True)
        ws.cell(row=row_num, column=10).font = Font(name='Arial', size=9)

        total_online_amount += data['online_amount']
        total_offline_amount += data['offline_amount']
        row_num += 1

    # Summary Section - Total Online & Offline
    row_num += 1
    ws.merge_cells(f'A{row_num}:C{row_num}')
    ws.cell(row=row_num, column=1,
            value="TOTAL PENJUALAN ONLINE").font = subtotal_font
    ws.cell(row=row_num, column=1).fill = light_gray_fill
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='right')
    ws.cell(row=row_num, column=1).border = border

    ws.merge_cells(f'D{row_num}:E{row_num}')
    online_total_cell = ws.cell(row=row_num, column=4)
    online_total_cell.value = f"Rp {total_online_amount:,.0f}".replace(
        ',', '.')
    online_total_cell.font = subtotal_font
    online_total_cell.fill = light_gray_fill
    online_total_cell.alignment = Alignment(horizontal='right')
    online_total_cell.border = border

    row_num += 1
    ws.merge_cells(f'A{row_num}:C{row_num}')
    ws.cell(row=row_num, column=1,
            value="TOTAL PENJUALAN KASIR").font = subtotal_font
    ws.cell(row=row_num, column=1).fill = light_gray_fill
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='right')
    ws.cell(row=row_num, column=1).border = border

    ws.merge_cells(f'D{row_num}:E{row_num}')
    offline_total_cell = ws.cell(row=row_num, column=4)
    offline_total_cell.value = f"Rp {total_offline_amount:,.0f}".replace(
        ',', '.')
    offline_total_cell.font = subtotal_font
    offline_total_cell.fill = light_gray_fill
    offline_total_cell.alignment = Alignment(horizontal='right')
    offline_total_cell.border = border

    # Grand Total
    row_num += 1
    ws.merge_cells(f'A{row_num}:C{row_num}')
    ws.cell(row=row_num, column=1, value="GRAND TOTAL").font = total_font
    ws.cell(row=row_num, column=1).fill = light_gray_fill
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='right')
    ws.cell(row=row_num, column=1).border = border

    ws.merge_cells(f'D{row_num}:E{row_num}')
    grand_total = total_online_amount + total_offline_amount
    grand_total_cell = ws.cell(row=row_num, column=4)
    grand_total_cell.value = f"Rp {grand_total:,.0f}".replace(',', '.')
    grand_total_cell.font = total_font
    grand_total_cell.fill = light_gray_fill
    grand_total_cell.alignment = Alignment(horizontal='right')
    grand_total_cell.border = border

    # Payment Method Breakdown - DETAIL PER CHANNEL
    row_num += 2
    ws.merge_cells(f'A{row_num}:J{row_num}')
    ws.cell(
        row=row_num,
        column=1,
        value=
        "BREAKDOWN METODE PEMBAYARAN (Online Web/Marketplace & Offline Kasir Toko)"
    ).font = subtitle_font
    ws.cell(row=row_num, column=1).fill = light_gray_fill
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=1).border = border

    row_num += 1
    # Header untuk breakdown
    ws.merge_cells(f'A{row_num}:D{row_num}')
    ws.cell(row=row_num, column=1,
            value="Metode Pembayaran").font = Font(name='Arial',
                                                   size=10,
                                                   bold=True)
    ws.cell(row=row_num, column=1).fill = light_gray_fill
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=1).border = border

    ws.merge_cells(f'E{row_num}:F{row_num}')
    ws.cell(row=row_num, column=5,
            value="Online (Web/Marketplace)").font = Font(name='Arial',
                                                          size=10,
                                                          bold=True)
    ws.cell(row=row_num, column=5).fill = light_gray_fill
    ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=5).border = border

    ws.merge_cells(f'G{row_num}:H{row_num}')
    ws.cell(row=row_num, column=7,
            value="Offline (Kasir Toko)").font = Font(name='Arial',
                                                      size=10,
                                                      bold=True)
    ws.cell(row=row_num, column=7).fill = light_gray_fill
    ws.cell(row=row_num, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=7).border = border

    ws.merge_cells(f'I{row_num}:J{row_num}')
    ws.cell(row=row_num, column=9, value="Total").font = Font(name='Arial',
                                                              size=10,
                                                              bold=True)
    ws.cell(row=row_num, column=9).fill = light_gray_fill
    ws.cell(row=row_num, column=9).alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=9).border = border

    row_num += 1
    for method, channels in sorted(
            payment_method_totals.items(),
            key=lambda x: x[1]['online'] + x[1]['offline'],
            reverse=True):
        ws.merge_cells(f'A{row_num}:D{row_num}')
        ws.cell(row=row_num, column=1, value=method).border = border
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='left')
        ws.cell(row=row_num, column=1).font = normal_font

        # Online amount - Format IDR
        ws.merge_cells(f'E{row_num}:F{row_num}')
        online_payment_cell = ws.cell(row=row_num, column=5)
        online_payment_cell.value = f"Rp {channels['online']:,.0f}".replace(
            ',', '.')
        online_payment_cell.border = border
        online_payment_cell.alignment = Alignment(horizontal='right')

        # Offline amount - Format IDR
        ws.merge_cells(f'G{row_num}:H{row_num}')
        offline_payment_cell = ws.cell(row=row_num, column=7)
        offline_payment_cell.value = f"Rp {channels['offline']:,.0f}".replace(
            ',', '.')
        offline_payment_cell.border = border
        offline_payment_cell.alignment = Alignment(horizontal='right')

        # Total - Format IDR
        total_method = channels['online'] + channels['offline']
        ws.merge_cells(f'I{row_num}:J{row_num}')
        total_payment_cell = ws.cell(row=row_num, column=9)
        total_payment_cell.value = f"Rp {total_method:,.0f}".replace(',', '.')
        total_payment_cell.border = border
        total_payment_cell.alignment = Alignment(horizontal='right')
        total_payment_cell.font = Font(name='Arial', size=10, bold=True)

        row_num += 1

    # Set column widths
    column_widths = [5, 30, 15, 10, 15, 10, 15, 10, 15, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Footer
    row_num += 2
    ws.merge_cells(f'A{row_num}:I{row_num}')
    footer_cell = ws[f'A{row_num}']
    footer_cell.value = f"Dicetak pada: {datetime.utcnow().strftime('%d %B %Y %H:%M:%S')} WIB"
    footer_cell.font = Font(name='Arial', size=9, italic=True, color='666666')
    footer_cell.alignment = Alignment(horizontal='center')

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'laporan_penjualan_{period}_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype=
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/order/<int:order_id>/print_address')
@login_required
@staff_required
def print_order_address(order_id):
    order = models.Order.query.get_or_404(order_id)
    store_profile = models.StoreProfile.get_active_profile()

    # Create PDF alamat pengiriman thermal (120mm width)
    buffer = io.BytesIO()

    # Calculate content
    content_lines = []

    # Header
    content_lines.extend([
        ('header', 'ALAMAT PENGIRIMAN'),
        ('divider', '=' * 35),
    ])

    # Store info (sender)
    content_lines.append(('section', 'DARI:'))
    if store_profile:
        content_lines.extend([
            ('store_name', store_profile.store_name),
            ('store_address', store_profile.formatted_address),
        ])
        if store_profile.store_phone:
            content_lines.append(
                ('store_contact', f"Telp: {store_profile.store_phone}"))
        if store_profile.store_email:
            content_lines.append(
                ('store_contact', f"Email: {store_profile.store_email}"))
    else:
        content_lines.extend([
            ('store_name', 'Hurtrock Music Store'),
            ('store_address', 'Jl. Musik Raya No. 123, Jakarta'),
            ('store_contact', 'Telp: 0821-1555-8035'),
        ])

    content_lines.append(('divider', '-' * 35))

    # Recipient info
    content_lines.extend([
        ('section', 'KEPADA:'),
        ('recipient_name', order.user.name.upper()),
    ])

    if order.user.phone:
        content_lines.append(
            ('recipient_contact', f"Telp: {order.user.phone}"))

    if order.user.address:
        # Split long address
        address = order.user.address.replace('\n', ' ')
        if len(address) > 32:
            words = address.split(' ')
            lines = []
            current_line = ""
            for word in words:
                if len(current_line + word) <= 32:
                    current_line += word + " "
                else:
                    if current_line:
                        lines.append(current_line.strip())
                    current_line = word + " "
            if current_line:
                lines.append(current_line.strip())

            for line in lines:
                content_lines.append(('recipient_address', line))
        else:
            content_lines.append(('recipient_address', address))

    content_lines.append(('divider', '-' * 35))

    # Order info
    content_lines.extend([
        ('order_info', f"Order #{order.id}"),
        ('order_date',
         f"Tanggal: {order.created_at.strftime('%d/%m/%Y %H:%M')}"),
    ])

    if order.tracking_number:
        content_lines.append(('tracking', f"Resi: {order.tracking_number}"))

    if order.courier_service:
        content_lines.append(('courier', f"Kurir: {order.courier_service}"))

    content_lines.extend([
        ('total', f"Total: {order.formatted_total}"),
        ('divider', '=' * 35),
        ('footer', 'Terima kasih atas kepercayaan Anda'),
    ])

    # Calculate dimensions
    width = 340  # 120mm
    line_height = 12
    margin = 20
    total_height = margin * 2 + (len(content_lines) * line_height) + 40

    if total_height < 400:
        total_height = 400

    p = canvas.Canvas(buffer, pagesize=(width, total_height))

    y_pos = total_height - margin

    for line_type, text in content_lines:
        if line_type == 'header':
            p.setFont("Helvetica-Bold", 14)
            text_width = p.stringWidth(text, "Helvetica-Bold", 14)
            x_center = (width - text_width) / 2
            p.drawString(x_center, y_pos, text)
            y_pos -= 18
        elif line_type == 'divider':
            p.setFont("Helvetica", 10)
            text_width = p.stringWidth(text, "Helvetica", 10)
            x_center = (width - text_width) / 2
            p.drawString(x_center, y_pos, text)
            y_pos -= 12
        elif line_type == 'section':
            p.setFont("Helvetica-Bold", 12)
            p.drawString(margin, y_pos, text)
            y_pos -= 14
        elif line_type in ['store_name', 'recipient_name']:
            p.setFont("Helvetica-Bold", 11)
            p.drawString(margin + 5, y_pos, text)
            y_pos -= 13
        elif line_type in [
                'store_address', 'store_contact', 'recipient_address',
                'recipient_contact'
        ]:
            p.setFont("Helvetica", 9)
            p.drawString(margin + 5, y_pos, text)
            y_pos -= 11
        elif line_type in ['order_info', 'order_date', 'tracking', 'courier']:
            p.setFont("Helvetica-Bold", 10)
            p.drawString(margin, y_pos, text)
            y_pos -= 12
        elif line_type == 'total':
            p.setFont("Helvetica-Bold", 12)
            text_width = p.stringWidth(text, "Helvetica-Bold", 12)
            x_center = (width - text_width) / 2
            p.drawString(x_center, y_pos, text)
            y_pos -= 15
        elif line_type == 'footer':
            p.setFont("Helvetica", 8)
            text_width = p.stringWidth(text, "Helvetica", 8)
            x_center = (width - text_width) / 2
            p.drawString(x_center, y_pos, text)
            y_pos -= 10
        else:
            p.setFont("Helvetica", 9)
            p.drawString(margin, y_pos, text)
            y_pos -= 11

    p.showPage()
    p.save()

    buffer.seek(0)

    return send_file(buffer,
                     as_attachment=True,
                     download_name=f'thermal_address_{order.id}.pdf',
                     mimetype='application/pdf')


# Restock Order Management
@app.route('/admin/restock')
@login_required
@admin_required
def admin_restock_orders():
    restock_orders = models.RestockOrder.query.order_by(
        models.RestockOrder.created_at.desc()).all()
    suppliers = models.Supplier.query.filter_by(is_active=True).all()
    products = models.Product.query.filter_by(is_active=True).all()

    # Convert to dict for JSON serialization
    restock_orders_data = []
    for order in restock_orders:
        order_data = {
            'id':
            order.id,
            'supplier': {
                'name': order.supplier.name,
                'contact_person': order.supplier.contact_person
            },
            'status':
            order.status,
            'total_amount':
            float(order.total_amount),
            'formatted_total':
            order.formatted_total,
            'created_at':
            order.created_at.isoformat(),
            'expected_date':
            order.expected_date.isoformat() if order.expected_date else None,
            'received_date':
            order.received_date.isoformat() if order.received_date else None,
            'notes':
            order.notes,
            'items': []
        }
        for item in order.items:
            order_data['items'].append({
                'product': {
                    'name': item.product.name
                },
                'quantity_ordered': item.quantity_ordered,
                'unit_cost': float(item.unit_cost)
            })
        restock_orders_data.append(order_data)

    products_data = []
    for product in products:
        products_data.append({
            'id': product.id,
            'name': product.name,
            'stock_quantity': product.stock_quantity
        })

    return render_template('admin/restock.html',
                           restock_orders=restock_orders,
                           suppliers=suppliers,
                           products=products,
                           restock_orders_data=restock_orders_data,
                           products_data=products_data)


@app.route('/admin/restock/create', methods=['POST'])
@login_required
@admin_required
def admin_create_restock_order():
    try:
        supplier_id = int(request.form['supplier_id'])
        notes = request.form.get('notes', '')
        expected_date_str = request.form.get('expected_date', '')

        expected_date = None
        if expected_date_str:
            expected_date = datetime.strptime(expected_date_str, '%Y-%m-%d')

        restock_order = models.RestockOrder(supplier_id=supplier_id,
                                            notes=notes,
                                            expected_date=expected_date,
                                            created_by=current_user.id)

        db.session.add(restock_order)
        db.session.flush()  # Get the order ID

        # Add items
        product_ids = request.form.getlist('product_ids[]')
        quantities = request.form.getlist('quantities[]')
        unit_costs = request.form.getlist('unit_costs[]')

        total_amount = 0
        for i, product_id in enumerate(product_ids):
            if product_id and quantities[i] and unit_costs[i]:
                quantity = int(quantities[i])
                unit_cost = float(unit_costs[i])

                item = models.RestockOrderItem(
                    restock_order_id=restock_order.id,
                    product_id=int(product_id),
                    quantity_ordered=quantity,
                    unit_cost=unit_cost)
                db.session.add(item)
                total_amount += quantity * unit_cost

        restock_order.total_amount = total_amount
        db.session.commit()

        flash(f'Order restock #{restock_order.id} berhasil dibuat!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_restock_orders'))


@app.route('/admin/restock/<int:order_id>/update_status', methods=['POST'])
@login_required
@admin_required
def admin_update_restock_status(order_id):
    restock_order = models.RestockOrder.query.get_or_404(order_id)

    try:
        new_status = request.form.get('status')
        if new_status not in ['pending', 'ordered', 'received', 'cancelled']:
            flash('Status tidak valid!', 'error')
            return redirect(url_for('admin_restock_orders'))

        restock_order.status = new_status

        if new_status == 'received':
            restock_order.received_date = datetime.utcnow()

            # Update product stock quantities
            for item in restock_order.items:
                item.quantity_received = item.quantity_ordered
                product = item.product
                product.stock_quantity += item.quantity_ordered

        db.session.commit()

        flash(f'Status restock order #{restock_order.id} berhasil diperbarui!',
              'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_restock_orders'))


@app.route('/admin/restock/<int:order_id>/invoice')
@login_required
@admin_required
def admin_generate_restock_invoice(order_id):
    restock_order = models.RestockOrder.query.get_or_404(order_id)

    # Create PDF invoice
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Header
    p.setFont("Helvetica-Bold", 18)
    p.drawString(50, height - 50, "INVOICE RESTOCK ORDER")

    # Store info
    p.setFont("Helvetica", 10)
    p.drawString(50, height - 80, "DARI: Hurtrock Music Store")
    p.drawString(50, height - 95, "Jl. Musik Raya No. 123, Jakarta")
    p.drawString(50, height - 110, "Telp: 0821-1555-8035")

    # Invoice info
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, height - 140,
                 f"Invoice #: RESTOCK-{restock_order.id:05d}")
    p.drawString(50, height - 160,
                 f"Tanggal: {restock_order.created_at.strftime('%d/%m/%Y')}")
    p.drawString(50, height - 180, f"Status: {restock_order.status.upper()}")

    # Supplier info
    p.setFont("Helvetica-Bold", 12)
    p.drawString(350, height - 140, "KEPADA:")
    p.setFont("Helvetica", 10)
    p.drawString(350, height - 160, f"{restock_order.supplier.name}")
    if restock_order.supplier.contact_person:
        p.drawString(350, height - 175,
                     f"PIC: {restock_order.supplier.contact_person}")
    if restock_order.supplier.phone:
        p.drawString(350, height - 190,
                     f"Telp: {restock_order.supplier.phone}")

    # Table header
    y_pos = height - 230
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y_pos, "Produk")
    p.drawString(300, y_pos, "Qty")
    p.drawString(350, y_pos, "Harga Satuan")
    p.drawString(450, y_pos, "Total")

    # Draw line
    p.line(50, y_pos - 5, width - 50, y_pos - 5)

    # Items
    y_pos -= 25
    p.setFont("Helvetica", 9)
    total_amount = 0

    for item in restock_order.items:
        p.drawString(50, y_pos, item.product.name[:30])
        p.drawString(300, y_pos, str(item.quantity_ordered))
        p.drawString(350, y_pos, f"Rp {item.unit_cost:,.0f}".replace(',', '.'))
        p.drawString(450, y_pos, f"Rp {item.subtotal:,.0f}".replace(',', '.'))
        y_pos -= 15
        total_amount += item.subtotal

    # Total
    p.line(50, y_pos - 5, width - 50, y_pos - 5)
    y_pos -= 20
    p.setFont("Helvetica-Bold", 12)
    p.drawString(350, y_pos, "TOTAL:")
    p.drawString(450, y_pos, f"Rp {total_amount:,.0f}".replace(',', '.'))

    # Notes
    if restock_order.notes:
        y_pos -= 40
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, y_pos, "Catatan:")
        y_pos -= 15
        p.setFont("Helvetica", 9)
        p.drawString(50, y_pos, restock_order.notes)

    p.showPage()
    p.save()

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'invoice_restock_{restock_order.id:05d}.pdf',
        mimetype='application/pdf')


# Shipping Services Management
@app.route('/admin/shipping')
@login_required
@admin_required
def admin_shipping_services():
    services = models.ShippingService.query.all()
    # Convert to dict for JSON serialization
    services_data = []
    for service in services:
        services_data.append({
            'id': service.id,
            'name': service.name,
            'code': service.code,
            'base_price': float(service.base_price),
            'price_per_kg': float(service.price_per_kg),
            'price_per_km': float(service.price_per_km),
            'volume_factor': float(service.volume_factor),
            'min_days': service.min_days,
            'max_days': service.max_days,
            'is_active': service.is_active
        })
    return render_template('admin/shipping.html',
                           services=services,
                           services_data=services_data)


@app.route('/admin/shipping/add', methods=['POST'])
@login_required
@admin_required
def admin_add_shipping_service():
    try:
        name = request.form['name']
        code = request.form['code']
        base_price = float(request.form['base_price'])
        price_per_kg = float(request.form['price_per_kg'])
        price_per_km = float(request.form.get('price_per_km', 0))
        volume_factor = float(request.form.get('volume_factor', 5000))
        min_days = int(request.form.get('min_days', 1))
        max_days = int(request.form.get('max_days', 3))

        service = models.ShippingService(name=name,
                                         code=code,
                                         base_price=base_price,
                                         price_per_kg=price_per_kg,
                                         price_per_km=price_per_km,
                                         volume_factor=volume_factor,
                                         min_days=min_days,
                                         max_days=max_days)

        db.session.add(service)
        db.session.commit()

        flash(f'Jasa kirim {name} berhasil ditambahkan!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_shipping_services'))


@app.route('/admin/shipping/<int:service_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_shipping_service(service_id):
    service = models.ShippingService.query.get_or_404(service_id)

    try:
        service.name = request.form['name']
        service.code = request.form['code']
        service.base_price = float(request.form['base_price'])
        service.price_per_kg = float(request.form['price_per_kg'])
        service.price_per_km = float(request.form.get('price_per_km', 0))
        service.volume_factor = float(request.form.get('volume_factor', 5000))
        service.min_days = int(request.form.get('min_days', 1))
        service.max_days = int(request.form.get('max_days', 3))
        service.is_active = request.form.get('is_active') == 'on'

        db.session.commit()

        flash(f'Jasa kirim {service.name} berhasil diperbarui!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_shipping_services'))


@app.route('/admin/shipping/<int:service_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_shipping_service(service_id):
    service = models.ShippingService.query.get_or_404(service_id)

    # Check if service is used in orders
    if service.orders:
        flash(
            'Tidak bisa menghapus jasa kirim yang sedang digunakan di pesanan!',
            'error')
        return redirect(url_for('admin_shipping_services'))

    service_name = service.name
    db.session.delete(service)
    db.session.commit()

    flash(f'Jasa kirim {service_name} berhasil dihapus!', 'success')
    return redirect(url_for('admin_shipping_services'))


# Supplier Management
@app.route('/admin/suppliers')
@login_required
@admin_required
def admin_suppliers():
    suppliers = models.Supplier.query.all()
    # Convert to dict for JSON serialization
    suppliers_data = []
    for supplier in suppliers:
        suppliers_data.append({
            'id': supplier.id,
            'name': supplier.name,
            'contact_person': supplier.contact_person,
            'email': supplier.email,
            'phone': supplier.phone,
            'address': supplier.address,
            'company': supplier.company,
            'notes': supplier.notes,
            'is_active': supplier.is_active
        })
    return render_template('admin/suppliers.html',
                           suppliers=suppliers,
                           suppliers_data=suppliers_data)


@app.route('/admin/suppliers/add', methods=['POST'])
@login_required
@admin_required
def admin_add_supplier():
    try:
        name = request.form['name']
        contact_person = request.form.get('contact_person', '')
        email = request.form.get('email', '')
        phone = request.form.get('phone', '')
        address = request.form.get('address', '')
        company = request.form.get('company', '')
        notes = request.form.get('notes', '')

        supplier = models.Supplier(name=name,
                                   contact_person=contact_person,
                                   email=email,
                                   phone=phone,
                                   address=address,
                                   company=company,
                                   notes=notes)

        db.session.add(supplier)
        db.session.commit()

        flash(f'Supplier {name} berhasil ditambahkan!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_suppliers'))


@app.route('/admin/suppliers/<int:supplier_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_supplier(supplier_id):
    supplier = models.Supplier.query.get_or_404(supplier_id)

    try:
        supplier.name = request.form['name']
        supplier.contact_person = request.form.get('contact_person', '')
        supplier.email = request.form.get('email', '')
        supplier.phone = request.form.get('phone', '')
        supplier.address = request.form.get('address', '')
        supplier.company = request.form.get('company', '')
        supplier.notes = request.form.get('notes', '')
        supplier.is_active = request.form.get('is_active') == 'on'

        db.session.commit()

        flash(f'Supplier {supplier.name} berhasil diperbarui!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_suppliers'))


@app.route('/admin/suppliers/<int:supplier_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_supplier(supplier_id):
    supplier = models.Supplier.query.get_or_404(supplier_id)

    # Check if supplier has products
    if supplier.products:
        flash('Tidak bisa menghapus supplier yang memiliki produk!', 'error')
        return redirect(url_for('admin_suppliers'))

    supplier_name = supplier.name
    db.session.delete(supplier)
    db.session.commit()

    flash(f'Supplier {supplier_name} berhasil dihapus!', 'success')
    return redirect(url_for('admin_suppliers'))


# Store Profile Management
@app.route('/admin/store-profile')
@login_required
@admin_required
def admin_store_profile():
    profile = models.StoreProfile.get_active_profile()
    if not profile:
        # Create default profile if none exists
        profile = models.StoreProfile(
            store_name='Hurtrock Music Store',
            store_tagline='Toko Alat Musik Terpercaya',
            store_address='Jl. Musik Raya No. 123',
            store_city='Jakarta',
            store_postal_code='12345',
            store_phone='0821-1555-8035',
            store_email='info@hurtrock.com',
            logo_url='/static/logo_perusahaan/default_music_icon.jpg')
        db.session.add(profile)
        db.session.commit()

    return render_template('admin/store_profile.html', profile=profile)


@app.route('/admin/store-profile/update', methods=['POST'])
@login_required
@admin_required
def admin_update_store_profile():
    profile = models.StoreProfile.get_active_profile()
    if not profile:
        profile = models.StoreProfile()
        db.session.add(profile)

    try:
        # Basic store information
        profile.store_name = request.form.get('store_name', '').strip()
        profile.store_tagline = request.form.get('store_tagline', '').strip()
        profile.store_address = request.form.get('store_address', '').strip()
        profile.store_city = request.form.get('store_city', '').strip()
        profile.store_postal_code = request.form.get('store_postal_code',
                                                     '').strip()
        profile.store_phone = request.form.get('store_phone', '').strip()
        profile.store_email = request.form.get('store_email', '').strip()
        profile.store_website = request.form.get('store_website', '').strip()

        # Branch information
        profile.branch_name = request.form.get('branch_name', '').strip()
        profile.branch_code = request.form.get('branch_code', '').strip()
        profile.branch_manager = request.form.get('branch_manager', '').strip()

        # Business information
        profile.business_license = request.form.get('business_license',
                                                    '').strip()
        profile.tax_number = request.form.get('tax_number', '').strip()

        # Operating hours
        profile.operating_hours = request.form.get('operating_hours',
                                                   '').strip()

        # Social media
        profile.facebook_url = request.form.get('facebook_url', '').strip()
        profile.instagram_url = request.form.get('instagram_url', '').strip()
        profile.whatsapp_number = request.form.get('whatsapp_number',
                                                   '').strip()

        # Colors
        profile.primary_color = request.form.get('primary_color', '#FF6B35')
        profile.secondary_color = request.form.get('secondary_color',
                                                   '#FF8C42')

        # Handle logo upload
        if 'logo' in request.files and request.files['logo'].filename:
            file = request.files['logo']
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filename = f"logo_{uuid.uuid4()}_{filename}"
                logo_dir = 'static/logo_perusahaan'
                filepath = os.path.join(logo_dir, filename)

                # Create directory if it doesn't exist
                os.makedirs(os.path.dirname(filepath), exist_ok=True)

                file.save(filepath)
                compress_image(filepath)
                profile.logo_url = f"/static/logo_perusahaan/{filename}"

        # Store description/about
        profile.store_description = request.form.get('store_description',
                                                     '').strip()

        profile.updated_at = models.get_utc_time()
        db.session.commit()

        flash('Profil toko berhasil diperbarui!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_store_profile'))


def generate_tracking_number():
    """Generate a tracking number for orders"""
    import random
    import string
    # Generate a simple tracking number format: TR + 8 random characters
    chars = string.ascii_uppercase + string.digits
    return 'TR' + ''.join(random.choices(chars, k=8))


def reset_and_initialize_database():
    """Reset database and create fresh schema - for development use only"""
    print("[WARNING] This will delete ALL data in the database!")
    confirm = input("Are you sure you want to reset the database? (yes/no): ")
    if confirm.lower() != 'yes':
        print("[INFO] Database reset cancelled")
        return False

    with app.app_context():
        try:
            # Rollback any pending transactions first
            db.session.rollback()

            # Drop all tables
            print("[INFO] Dropping all tables...")
            db.drop_all()

            # Create all tables fresh with current schema
            print("[INFO] Creating fresh tables...")
            db.create_all()

            # Create default admin user
            print("[INFO] Creating default admin user...")
            admin_user = models.User(
                email="admin@hurtrock.com",
                password_hash=generate_password_hash("admin123"),
                name="Administrator",
                role="admin")
            db.session.add(admin_user)

            # Create default store profile
            print("[INFO] Creating default store profile...")
            store_profile = models.StoreProfile(
                store_name='Hurtrock Music Store',
                store_tagline='Toko Alat Musik Terpercaya',
                store_address=
                'Jl. Musik Raya No. 123, RT/RW 001/002, Kelurahan Musik, Kecamatan Harmoni',
                store_city='Jakarta Selatan',
                store_postal_code='12345',
                store_phone='0821-1555-8035',
                store_email='info@hurtrock.com',
                store_website='https://hurtrock.com',
                whatsapp_number='6282115558035',
                operating_hours=
                'Senin - Sabtu: 09:00 - 21:00\nMinggu: 10:00 - 18:00',
                branch_name='Cabang Pusat',
                branch_code='HRT-001')
            db.session.add(store_profile)

            # Create sample data
            create_sample_data()

            db.session.commit()

            print("[SUCCESS] Database reset and initialization completed!")
            print(f"[INFO] Default admin login:")
            print(f"        Email: admin@hurtrock.com")
            print(f"        Password: admin123")

            return True

        except Exception as e:
            print(f"[ERROR] Database reset failed: {e}")
            db.session.rollback()
            return False